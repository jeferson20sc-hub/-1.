"""Helpers for sheet construction (column widths, merged headers, tables)."""

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .theme import (
    BORDER_ALL_THIN,
    BRANCO,
    CENTER,
    CINZA_BORDA,
    CINZA_TITULO,
    GRAFITE,
    LEFT,
    PETROLEO,
    PETROLEO_CLARO,
    fill,
)
from openpyxl.styles import Alignment, Border, Font, Side


def set_column_widths(ws, widths: dict) -> None:
    """widths: {'A': 20, 'B': 15, ...}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def set_row_heights(ws, heights: dict) -> None:
    for r, h in heights.items():
        ws.row_dimensions[r].height = h


def title_bar(ws, text: str, subtitle: str, last_col: str = "N") -> None:
    """Render a 2-row executive title bar with petroleo background."""
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = text
    ws["A1"].style = "exa_title"

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].style = "exa_subtitle"


def section_header(ws, cell_range: str, text: str) -> None:
    """Section header with bottom petroleo border."""
    start = cell_range.split(":")[0]
    ws.merge_cells(cell_range)
    ws[start] = text
    ws[start].style = "exa_section"
    # Set row height for the start row
    row = int("".join(c for c in start if c.isdigit()))
    if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < 24:
        ws.row_dimensions[row].height = 26


def kpi_card(ws, label_cell: str, value_cell: str, label: str, value, fmt: str | None = None,
             value_style: str = "exa_kpi_value") -> None:
    """Two-cell KPI card: label on top, value on bottom (or side-by-side)."""
    ws[label_cell] = label
    ws[label_cell].style = "exa_kpi_label"
    ws[value_cell] = value
    ws[value_cell].style = value_style
    if fmt:
        ws[value_cell].number_format = fmt


def make_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> Table:
    """Create a structured table on the sheet.

    Returns the Table object so the caller can attach it.
    """
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    return tbl


def write_table(ws, top_left: str, headers: list[str], rows: list[tuple],
                table_name: str, style: str = "TableStyleMedium2") -> str:
    """Write headers + rows starting at top_left and register a structured table.

    Returns the resulting reference string (e.g. 'A5:F25').
    """
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    col_letter, row = coordinate_from_string(top_left)
    start_col = column_index_from_string(col_letter)

    # headers
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
    # rows
    n_rows = max(len(rows), 1)
    for r_idx, r in enumerate(rows, start=1):
        for c_idx, v in enumerate(r):
            ws.cell(row=row + r_idx, column=start_col + c_idx, value=v)

    end_col = get_column_letter(start_col + len(headers) - 1)
    end_row = row + n_rows
    ref = f"{col_letter}{row}:{end_col}{end_row}"
    make_table(ws, table_name, ref, style)
    return ref


def freeze_header(ws, cell: str = "A3") -> None:
    ws.freeze_panes = cell


def hide_gridlines(ws) -> None:
    ws.sheet_view.showGridLines = False


def block_fill(ws, cell_range: str, color: str) -> None:
    """Fill all cells in a range with a solid color."""
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill(color)


def thin_border_range(ws, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER_ALL_THIN


def page_setup_landscape(ws, fit_to_width: int = 1) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6


def add_data_validation(ws, cell_range: str, formula: str, type_: str = "list",
                        prompt: str | None = None, error: str | None = None) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type=type_, formula1=formula, allow_blank=True)
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    if error:
        dv.error = error
        dv.showErrorMessage = True
    dv.add(cell_range)
    ws.add_data_validation(dv)
