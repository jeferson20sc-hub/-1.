"""Visual theme and number formats for EXAUSTAO 360 ENTERPRISE PRO.

Industrial premium palette: dark teal (azul petroleo), graphite, white, light
gray, plus controlled green/amber/red for status.
"""

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    NamedStyle,
)

# ---------------------------------------------------------------------------
# Color palette (hex without leading #)
# ---------------------------------------------------------------------------
PETROLEO = "FF0B3C5D"          # Azul petroleo escuro - cabecalhos
PETROLEO_CLARO = "FF1F4E5F"    # Variacao para faixas
GRAFITE = "FF2D2D2D"           # Texto principal
BRANCO = "FFFFFFFF"
CINZA_CLARO = "FFF2F2F2"       # Fundo cards
CINZA_BORDA = "FFD9D9D9"       # Bordas
CINZA_TITULO = "FFE6ECEF"      # Titulos de secao
VERDE = "FF2E7D32"             # Positivo / OK
VERDE_FUNDO = "FFE8F5E9"
AMBAR = "FFED9F00"             # Alerta
AMBAR_FUNDO = "FFFFF4E1"
VERMELHO = "FFC62828"          # Critico
VERMELHO_FUNDO = "FFFDECEC"
AZUL_KPI = "FF1976D2"          # KPI neutro

# ---------------------------------------------------------------------------
# Number formats (avoid changing actual cell values - just visual)
# ---------------------------------------------------------------------------
FMT_BRL = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_BRL_MI = '"R$ "#,##0.00,," mi"'
FMT_BRL_K = '"R$ "#,##0,," mil"'
FMT_PCT = '0.0%'
FMT_PCT_INT = '0%'
FMT_INT = '#,##0'
FMT_HORAS = '#,##0.0" h"'
FMT_DATA = 'dd/mm/yyyy'
FMT_DATAHORA = 'dd/mm/yyyy hh:mm'
FMT_TEXTO = '@'

# ---------------------------------------------------------------------------
# Reusable style fragments
# ---------------------------------------------------------------------------
THIN = Side(style="thin", color=CINZA_BORDA)
MEDIUM = Side(style="medium", color=PETROLEO)
BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM_PETROLEO = Border(bottom=Side(style="medium", color=PETROLEO))
BORDER_TOP_PETROLEO = Border(top=Side(style="medium", color=PETROLEO))

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


def fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ---------------------------------------------------------------------------
# Named styles (registered once on workbook)
# ---------------------------------------------------------------------------
def register_named_styles(wb) -> None:
    """Register reusable named styles. Idempotent."""

    def add(style: NamedStyle) -> None:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    # Title bar (top of every executive sheet)
    title = NamedStyle(name="exa_title")
    title.font = Font(name="Calibri", size=20, bold=True, color=BRANCO)
    title.fill = fill(PETROLEO)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    add(title)

    subtitle = NamedStyle(name="exa_subtitle")
    subtitle.font = Font(name="Calibri", size=11, italic=True, color=BRANCO)
    subtitle.fill = fill(PETROLEO_CLARO)
    subtitle.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    add(subtitle)

    # Section header
    section = NamedStyle(name="exa_section")
    section.font = Font(name="Calibri", size=12, bold=True, color=PETROLEO)
    section.fill = fill(CINZA_TITULO)
    section.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    section.border = BORDER_BOTTOM_PETROLEO
    add(section)

    # KPI card label
    kpi_label = NamedStyle(name="exa_kpi_label")
    kpi_label.font = Font(name="Calibri", size=9, bold=True, color=GRAFITE)
    kpi_label.fill = fill(CINZA_CLARO)
    kpi_label.alignment = Alignment(horizontal="center", vertical="center")
    kpi_label.border = BORDER_ALL_THIN
    add(kpi_label)

    # KPI card value
    kpi_value = NamedStyle(name="exa_kpi_value")
    kpi_value.font = Font(name="Calibri", size=18, bold=True, color=PETROLEO)
    kpi_value.fill = fill(BRANCO)
    kpi_value.alignment = CENTER
    kpi_value.border = BORDER_ALL_THIN
    add(kpi_value)

    kpi_value_red = NamedStyle(name="exa_kpi_value_red")
    kpi_value_red.font = Font(name="Calibri", size=18, bold=True, color=VERMELHO)
    kpi_value_red.fill = fill(BRANCO)
    kpi_value_red.alignment = CENTER
    kpi_value_red.border = BORDER_ALL_THIN
    add(kpi_value_red)

    kpi_value_green = NamedStyle(name="exa_kpi_value_green")
    kpi_value_green.font = Font(name="Calibri", size=18, bold=True, color=VERDE)
    kpi_value_green.fill = fill(BRANCO)
    kpi_value_green.alignment = CENTER
    kpi_value_green.border = BORDER_ALL_THIN
    add(kpi_value_green)

    # Table header (used outside structured tables)
    th = NamedStyle(name="exa_th")
    th.font = Font(name="Calibri", size=10, bold=True, color=BRANCO)
    th.fill = fill(PETROLEO)
    th.alignment = CENTER
    th.border = BORDER_ALL_THIN
    add(th)

    # Table body
    td = NamedStyle(name="exa_td")
    td.font = Font(name="Calibri", size=10, color=GRAFITE)
    td.fill = fill(BRANCO)
    td.alignment = LEFT
    td.border = BORDER_ALL_THIN
    add(td)

    # Input field (for forms)
    inp = NamedStyle(name="exa_input")
    inp.font = Font(name="Calibri", size=11, color=GRAFITE)
    inp.fill = fill("FFFFF8E1")
    inp.alignment = LEFT
    inp.border = BORDER_ALL_THIN
    add(inp)

    # Label
    lbl = NamedStyle(name="exa_label")
    lbl.font = Font(name="Calibri", size=10, bold=True, color=PETROLEO)
    lbl.fill = fill(CINZA_CLARO)
    lbl.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    lbl.border = BORDER_ALL_THIN
    add(lbl)

    # Footer/note
    note = NamedStyle(name="exa_note")
    note.font = Font(name="Calibri", size=9, italic=True, color="FF6B6B6B")
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    add(note)
