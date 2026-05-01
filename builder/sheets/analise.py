"""ANALISE sheet: MTBF/MTTR per forno, FMEA, plano de acao."""

from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font

from ..data import FMEA, FORNOS, PLANO_ACAO
from ..theme import (
    AMBAR, AMBAR_FUNDO, BORDER_ALL_THIN, BRANCO, FMT_BRL, FMT_DATA, FMT_HORAS,
    FMT_INT, FMT_PCT, GRAFITE, PETROLEO, VERDE, VERDE_FUNDO, VERMELHO,
    VERMELHO_FUNDO, fill,
)
from ..util import (
    freeze_header,
    hide_gridlines,
    page_setup_landscape,
    section_header,
    set_column_widths,
    set_row_heights,
    title_bar,
    write_table,
)


def build(wb):
    ws = wb.create_sheet("ANALISE")
    hide_gridlines(ws)
    set_column_widths(ws, {
        "A": 2, "B": 10, "C": 22, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 22, "K": 18, "L": 18, "M": 14, "N": 14,
    })
    title_bar(
        ws,
        "EXAUSTAO 360 - ANALISE DE CONFIABILIDADE",
        "MTBF / MTTR por forno, FMEA e Plano de Acao. Atualizar com [Atualizar Tudo].",
        last_col="N",
    )

    # ------------------------------------------------------------------
    # MTBF / MTTR por forno
    # ------------------------------------------------------------------
    section_header(ws, "B4:H4", "MTBF e MTTR por Forno")
    headers = ["Forno", "Eventos", "Duracao Total (h)", "MTBF (h)",
               "MTTR (h)", "Custo Total (R$)", "Status"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=2 + i, value=h)
        c.style = "exa_th"
    ws.row_dimensions[5].height = 22

    for idx, forno in enumerate(FORNOS):
        r = 6 + idx
        forno_id = forno[0]
        ws.cell(row=r, column=2, value=forno_id).style = "exa_td"
        ws.cell(row=r, column=3,
                value=f'=COUNTIF(tbEventos[Forno],B{r})').style = "exa_td"
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f'=SUMIF(tbEventos[Forno],B{r},tbEventos[DuracaoHoras])').style = "exa_td"
        ws.cell(row=r, column=4).number_format = FMT_HORAS
        # MTBF: total operating hours / events
        ws.cell(row=r, column=5,
                value=f'=IFERROR(K_HORAS_PERIODO/MAX(C{r},1),0)').style = "exa_td"
        ws.cell(row=r, column=5).number_format = FMT_HORAS
        # MTTR
        ws.cell(row=r, column=6,
                value=f'=IFERROR(D{r}/MAX(C{r},1),0)').style = "exa_td"
        ws.cell(row=r, column=6).number_format = FMT_HORAS
        ws.cell(row=r, column=7,
                value=f'=SUMIF(tbEventos[Forno],B{r},tbEventos[CustoEstimado])').style = "exa_td"
        ws.cell(row=r, column=7).number_format = FMT_BRL
        ws.cell(row=r, column=8,
                value=f'=IF(E{r}<K_META_MTBF,"Atencao",IF(F{r}>K_META_MTTR,"Revisar","OK"))').style = "exa_td"
        ws.cell(row=r, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    last_mt = 6 + len(FORNOS) - 1
    # Conditional formatting for status
    ws.conditional_formatting.add(
        f"H6:H{last_mt}",
        FormulaRule(formula=[f'$H6="Atencao"'], fill=fill(AMBAR_FUNDO),
                    font=Font(color=AMBAR, bold=True))
    )
    ws.conditional_formatting.add(
        f"H6:H{last_mt}",
        FormulaRule(formula=[f'$H6="Revisar"'], fill=fill(VERMELHO_FUNDO),
                    font=Font(color=VERMELHO, bold=True))
    )
    ws.conditional_formatting.add(
        f"H6:H{last_mt}",
        FormulaRule(formula=[f'$H6="OK"'], fill=fill(VERDE_FUNDO),
                    font=Font(color=VERDE, bold=True))
    )

    # ------------------------------------------------------------------
    # FMEA
    # ------------------------------------------------------------------
    fmea_start = last_mt + 3
    section_header(ws, f"B{fmea_start}:K{fmea_start}", "FMEA - Modos de Falha e Efeitos")
    write_table(
        ws, f"B{fmea_start + 1}",
        ["Forno", "Componente", "ModoFalha", "Efeito", "Controle",
         "Severidade", "Ocorrencia", "Deteccao", "RPN", "Prioridade"],
        FMEA,
        "tbFMEA",
        style="TableStyleMedium3",
    )

    fmea_first_data_row = fmea_start + 2
    fmea_last_data_row = fmea_first_data_row + len(FMEA) - 1
    # Color RPN by value
    ws.conditional_formatting.add(
        f"J{fmea_first_data_row}:J{fmea_last_data_row}",
        FormulaRule(formula=[f"$J{fmea_first_data_row}>=60"],
                    fill=fill(VERMELHO_FUNDO),
                    font=Font(color=VERMELHO, bold=True))
    )
    ws.conditional_formatting.add(
        f"J{fmea_first_data_row}:J{fmea_last_data_row}",
        FormulaRule(formula=[f"AND($J{fmea_first_data_row}>=30,$J{fmea_first_data_row}<60)"],
                    fill=fill(AMBAR_FUNDO),
                    font=Font(color=AMBAR, bold=True))
    )
    ws.conditional_formatting.add(
        f"J{fmea_first_data_row}:J{fmea_last_data_row}",
        FormulaRule(formula=[f"$J{fmea_first_data_row}<30"],
                    fill=fill(VERDE_FUNDO),
                    font=Font(color=VERDE, bold=True))
    )

    # ------------------------------------------------------------------
    # Plano de Acao
    # ------------------------------------------------------------------
    plano_start = fmea_last_data_row + 3
    section_header(ws, f"B{plano_start}:K{plano_start}", "Plano de Acao")
    plano_headers = ["ID", "Acao", "Responsavel", "Prazo", "Status",
                     "Prioridade", "ImpactoFinanceiro", "Risco", "ProximaEtapa"]
    write_table(
        ws, f"B{plano_start + 1}",
        plano_headers,
        PLANO_ACAO,
        "tbPlanoAcao",
        style="TableStyleMedium4",
    )

    plano_first_row = plano_start + 2
    plano_last_row = plano_first_row + len(PLANO_ACAO) - 1
    # Format Prazo as date and ImpactoFinanceiro as currency
    for r in range(plano_first_row, plano_last_row + 1):
        ws.cell(row=r, column=5).number_format = FMT_DATA  # Prazo (col E)
        ws.cell(row=r, column=8).number_format = FMT_BRL   # Impacto (col H)

    # Conditional formatting on Status column (col F)
    ws.conditional_formatting.add(
        f"F{plano_first_row}:F{plano_last_row}",
        FormulaRule(formula=[f'$F{plano_first_row}="Atrasado"'],
                    fill=fill(VERMELHO_FUNDO),
                    font=Font(color=VERMELHO, bold=True))
    )
    ws.conditional_formatting.add(
        f"F{plano_first_row}:F{plano_last_row}",
        FormulaRule(formula=[f'$F{plano_first_row}="Em andamento"'],
                    fill=fill(AMBAR_FUNDO),
                    font=Font(color=AMBAR, bold=True))
    )
    ws.conditional_formatting.add(
        f"F{plano_first_row}:F{plano_last_row}",
        FormulaRule(formula=[f'$F{plano_first_row}="Concluido"'],
                    fill=fill(VERDE_FUNDO),
                    font=Font(color=VERDE, bold=True))
    )

    # Chart: MTTR/MTBF per forno
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "MTBF vs MTTR por Forno"
    chart.y_axis.title = "Horas"
    mtbf_data = Reference(ws, min_col=5, min_row=5, max_col=5, max_row=last_mt)
    mttr_data = Reference(ws, min_col=6, min_row=5, max_col=6, max_row=last_mt)
    cats = Reference(ws, min_col=2, min_row=6, max_col=2, max_row=last_mt)
    chart.add_data(mtbf_data, titles_from_data=True)
    chart.add_data(mttr_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, f"J5")

    freeze_header(ws, "A4")
    page_setup_landscape(ws)
    return ws
