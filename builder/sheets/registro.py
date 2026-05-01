"""REGISTRO sheet: validated entry form for new events."""

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ..theme import (
    BORDER_ALL_THIN,
    BRANCO,
    CINZA_CLARO,
    FMT_BRL,
    FMT_DATA,
    FMT_HORAS,
    GRAFITE,
    PETROLEO,
    fill,
)
from ..util import (
    add_data_validation,
    freeze_header,
    hide_gridlines,
    page_setup_landscape,
    section_header,
    set_column_widths,
    set_row_heights,
    title_bar,
)


def build(wb):
    ws = wb.create_sheet("REGISTRO")
    hide_gridlines(ws)
    set_column_widths(ws, {
        "A": 2, "B": 22, "C": 32, "D": 4, "E": 22, "F": 32, "G": 2,
        "H": 22, "I": 32,
    })
    title_bar(
        ws,
        "EXAUSTAO 360 - REGISTRO DE EVENTO",
        "Preencha os campos abaixo. Campos com (*) sao obrigatorios. Use os botoes para Salvar, Limpar ou Validar.",
        last_col="I",
    )

    # Form section header
    section_header(ws, "B4:I4", "Novo Registro Operacional")

    # Form fields - two columns of label/input pairs
    # left column (B/C), right column (E/F), extras (H/I)
    fields_left = [
        (5,  "Data (*)",          "C5",  FMT_DATA, "=TODAY()"),
        (6,  "Hora (*)",          "C6",  "hh:mm", "=TEXT(NOW(),\"hh:mm\")"),
        (7,  "Operador (*)",      "C7",  None, ""),
        (8,  "Turno",             "C8",  None, ""),
        (9,  "Forno (*)",         "C9",  None, ""),
        (10, "Tipo de Evento (*)","C10", None, ""),
        (11, "Categoria",         "C11", None, ""),
    ]
    fields_right = [
        (5,  "Componente",         "F5",  None, ""),
        (6,  "Criticidade (*)",    "F6",  None, ""),
        (7,  "Duracao (horas)",    "F7",  FMT_HORAS, 0),
        (8,  "Custo Estimado",     "F8",  FMT_BRL, 0),
        (9,  "Status",             "F9",  None, "Aberto"),
        (10, "Numero da OS",       "F10", None, ""),
        (11, "Codigo do Lote",     "F11", None, ""),
    ]

    for row, label, value_cell, fmt, default in fields_left:
        ws.cell(row=row, column=2, value=label).style = "exa_label"
        c = ws[value_cell]
        c.value = default
        c.style = "exa_input"
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[row].height = 22

    for row, label, value_cell, fmt, default in fields_right:
        ws.cell(row=row, column=5, value=label).style = "exa_label"
        c = ws[value_cell]
        c.value = default
        c.style = "exa_input"
        if fmt:
            c.number_format = fmt

    # Observation - large text area
    ws.cell(row=13, column=2, value="Observacao").style = "exa_label"
    ws.merge_cells("C13:I15")
    ws["C13"].style = "exa_input"
    ws["C13"].alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True, indent=1)
    set_row_heights(ws, {13: 22, 14: 22, 15: 22})

    # Validation messages area
    section_header(ws, "B17:I17", "Status do Registro")
    ws.merge_cells("B18:I19")
    ws["B18"] = "Pronto para registro. Clique em VALIDAR para conferir os dados."
    ws["B18"].alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True, indent=2)
    ws["B18"].font = Font(name="Calibri", size=11, italic=True, color="FF6B6B6B")
    ws["B18"].fill = fill(CINZA_CLARO)
    ws.row_dimensions[18].height = 22
    ws.row_dimensions[19].height = 22

    # Buttons placeholder area (real shapes added later by VBA on first run)
    section_header(ws, "B21:I21", "Acoes")
    button_cells = [
        ("C22", "VALIDAR", "ValidarRegistro", "FF1976D2"),
        ("E22", "SALVAR / REGISTRAR", "RegistrarEvento360", "FF2E7D32"),
        ("G22", "LIMPAR", "LimparRegistro", "FFED9F00"),
    ]
    for cell, label, macro, color in button_cells:
        c = ws[cell]
        c.value = label
        c.font = Font(name="Calibri", size=11, bold=True, color=BRANCO)
        c.fill = fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_ALL_THIN
    set_row_heights(ws, {22: 30})

    # Data validations from CONFIG tables
    add_data_validation(
        ws, "C7", "=INDIRECT(\"tbOperadores[Nome]\")",
        prompt="Selecione o operador na lista", error="Operador nao cadastrado em CONFIG.",
    )
    add_data_validation(
        ws, "C8", '"A,B,C"',
        prompt="Turno (A, B ou C)",
    )
    add_data_validation(
        ws, "C9", "=INDIRECT(\"tbFornos[ID]\")",
        prompt="Selecione o forno", error="Forno nao cadastrado em CONFIG.",
    )
    add_data_validation(
        ws, "C10", "=INDIRECT(\"tbTiposEvento[Tipo]\")",
        prompt="Selecione o tipo de evento", error="Tipo nao cadastrado.",
    )
    add_data_validation(
        ws, "C11", "=INDIRECT(\"tbCategorias[Categoria]\")",
        prompt="Selecione a categoria",
    )
    add_data_validation(
        ws, "F5", "=INDIRECT(\"tbComponentes[Nome]\")",
        prompt="Selecione o componente",
    )
    add_data_validation(
        ws, "F6", "=INDIRECT(\"tbCriticidades[Nivel]\")",
        prompt="Selecione a criticidade",
    )
    add_data_validation(
        ws, "F9", '"Aberto,Em analise,Resolvido"',
        prompt="Status do evento",
    )

    # Numeric validations
    dv_dur = DataValidation(type="decimal", operator="greaterThanOrEqual",
                            formula1="0", allow_blank=True,
                            error="Duracao deve ser >= 0",
                            showErrorMessage=True)
    dv_dur.add("F7")
    ws.add_data_validation(dv_dur)

    dv_custo = DataValidation(type="decimal", operator="greaterThanOrEqual",
                              formula1="0", allow_blank=True,
                              error="Custo deve ser >= 0",
                              showErrorMessage=True)
    dv_custo.add("F8")
    ws.add_data_validation(dv_custo)

    # Help footer
    ws.merge_cells("B25:I27")
    ws["B25"] = (
        "INSTRUCOES: 1) Preencha os campos. 2) Clique em VALIDAR. "
        "3) Se OK, clique em SALVAR. O registro e gravado em BASE > tbEventos "
        "e o DASHBOARD e atualizado. Use LIMPAR para iniciar novo registro."
    )
    ws["B25"].alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True, indent=1)
    ws["B25"].style = "exa_note"
    set_row_heights(ws, {25: 18, 26: 18, 27: 18})

    freeze_header(ws, "A4")
    page_setup_landscape(ws)
    return ws
