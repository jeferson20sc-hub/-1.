"""SIMULADOR sheet: scenario analysis and ROI calculator.

Defined names like INPUT_PerdaAnual are registered centrally in build.py.
"""

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font

from ..theme import (
    AMBAR, BORDER_ALL_THIN, BRANCO, CINZA_BORDA, CINZA_CLARO, FMT_BRL,
    FMT_BRL_MI, FMT_INT, FMT_PCT, GRAFITE, PETROLEO, VERDE, VERMELHO, fill,
)
from ..util import (
    hide_gridlines,
    page_setup_landscape,
    section_header,
    set_column_widths,
    set_row_heights,
    title_bar,
)


# Cell coordinates that build.py uses to register INPUT_* defined names
INPUT_CELLS = {
    "INPUT_CustoHoraParada": "C5",
    "INPUT_HorasParadas":    "C6",
    "INPUT_CustoFalha":      "C7",
    "INPUT_NumFalhas":       "C8",
    "INPUT_PerdaAnual":      "C9",
    "INPUT_Investimento":    "C10",
    "INPUT_Desconto":        "C11",
    "INPUT_Horizonte":       "C12",
}


def build(wb):
    ws = wb.create_sheet("SIMULADOR")
    hide_gridlines(ws)
    set_column_widths(ws, {
        "A": 2, "B": 32, "C": 18, "D": 4, "E": 22, "F": 22, "G": 22, "H": 4,
        "I": 28, "J": 18, "K": 18,
    })
    title_bar(
        ws,
        "EXAUSTAO 360 - SIMULADOR FINANCEIRO",
        "Cenarios Conservador / Realista / Agressivo. Os resultados dependem dos parametros editaveis abaixo.",
        last_col="K",
    )

    # ------------------------------------------------------------------
    # Inputs (editaveis)
    # ------------------------------------------------------------------
    section_header(ws, "B4:C4", "Premissas Editaveis")
    inputs = [
        (5,  "Custo medio por hora parada (R$)",        25000,   FMT_BRL),
        (6,  "Horas paradas no periodo de referencia",  120,     FMT_INT),
        (7,  "Custo por falha (R$)",                    18000,   FMT_BRL),
        (8,  "Numero de falhas no periodo",             45,      FMT_INT),
        (9,  "Perda anual estimada hoje (R$)",          2000000, FMT_BRL),
        (10, "Investimento na solucao (R$)",            80000,   FMT_BRL),
        (11, "Taxa de desconto anual (%)",              0.12,    FMT_PCT),
        (12, "Horizonte de analise (anos)",             3,       FMT_INT),
    ]
    for row, label, default, fmt in inputs:
        ws.cell(row=row, column=2, value=label).style = "exa_label"
        c = ws.cell(row=row, column=3, value=default)
        c.style = "exa_input"
        c.number_format = fmt
        ws.row_dimensions[row].height = 20

    # ------------------------------------------------------------------
    # Cenarios
    # ------------------------------------------------------------------
    section_header(ws, "E4:G4", "Cenarios de Reducao de Perdas")

    headers = ["Conservador", "Realista", "Agressivo"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=5 + i, value=h)
        c.style = "exa_th"
    ws.row_dimensions[5].height = 22

    # Reduction percentages (editable input row)
    ws.cell(row=6, column=5, value=0.05).style = "exa_input"
    ws.cell(row=6, column=6, value=0.10).style = "exa_input"
    ws.cell(row=6, column=7, value=0.20).style = "exa_input"
    for col in (5, 6, 7):
        ws.cell(row=6, column=col).number_format = FMT_PCT

    # Side label column for scenario rows
    side_labels = [
        ("% Reducao",          6,  None,                                                          None),
        ("Economia/ano",       7,  '=INPUT_PerdaAnual*{c}',                                       FMT_BRL),
        ("Em milhoes",         8,  '=INPUT_PerdaAnual*{c}',                                       FMT_BRL_MI),
        ("Em horizonte",       9,  '=INPUT_PerdaAnual*{c}*INPUT_Horizonte',                       FMT_BRL_MI),
        ("VPL (NPV)",          10, '=NPV(INPUT_Desconto, INPUT_PerdaAnual*{c}, INPUT_PerdaAnual*{c}, INPUT_PerdaAnual*{c})-INPUT_Investimento', FMT_BRL),
        ("Payback",            11, '=IFERROR(INPUT_Investimento/(INPUT_PerdaAnual*{c}),0)',       '0.00" anos"'),
        ("ROI Anual",          12, '=IFERROR((INPUT_PerdaAnual*{c}-INPUT_Investimento)/INPUT_Investimento,0)', FMT_PCT),
        ("Contrato Sugerido",  13, '=INPUT_PerdaAnual*{c}*0.3',                                   FMT_BRL),
    ]
    for label, row, _, _ in side_labels:
        ws.cell(row=row, column=2, value=label).style = "exa_label"
        ws.row_dimensions[row].height = 20

    # Fill scenario formulas
    for label, row, formula, fmt in side_labels:
        if formula is None:
            continue
        for col in (5, 6, 7):
            scen_cell = ws.cell(row=6, column=col).coordinate
            f = formula.replace("{c}", scen_cell)
            c = ws.cell(row=row, column=col, value=f)
            c.style = "exa_td"
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                c.number_format = fmt

    # Highlight Realista column
    for row in range(5, 14):
        ws.cell(row=row, column=6).fill = fill("FFE6F0FA")

    # ------------------------------------------------------------------
    # Resumo executivo (Realista)
    # ------------------------------------------------------------------
    section_header(ws, "I4:K4", "Resumo Executivo (Realista)")
    summary_items = [
        ("Investimento na solucao", "=INPUT_Investimento", FMT_BRL),
        ("Economia anual (10%)",    "=INPUT_PerdaAnual*0.1", FMT_BRL),
        ("Em milhoes/ano",          "=INPUT_PerdaAnual*0.1", FMT_BRL_MI),
        ("Payback (anos)",          "=IFERROR(INPUT_Investimento/(INPUT_PerdaAnual*0.1),0)", '0.00" anos"'),
        ("ROI Anual",               "=IFERROR((INPUT_PerdaAnual*0.1-INPUT_Investimento)/INPUT_Investimento,0)", FMT_PCT),
        ("Contrato Sugerido",       "=INPUT_PerdaAnual*0.1*0.3", FMT_BRL),
        ("Justificativa",
         '=IF(INPUT_PerdaAnual*0.1>INPUT_Investimento,"Solucao paga-se em menos de 1 ano e protege margem operacional.","Avaliar premissas - economia inferior ao investimento.")',
         None),
    ]
    for i, (label, formula, fmt) in enumerate(summary_items):
        r = 5 + i
        ws.cell(row=r, column=9, value=label).style = "exa_label"
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
        c = ws.cell(row=r, column=10, value=formula)
        c.style = "exa_td"
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 22

    # ------------------------------------------------------------------
    # Action buttons
    # ------------------------------------------------------------------
    section_header(ws, "B16:G16", "Acoes do Simulador")
    btns = [
        ("C17", "SIMULAR CENARIOS", "SimularCenarios", "FF1976D2"),
        ("E17", "SIMULAR MONTE CARLO", "SimularMonteCarlo360", "FF0B3C5D"),
    ]
    for cell, label, _macro, color in btns:
        c = ws[cell]
        c.value = label
        c.font = Font(name="Calibri", size=11, bold=True, color=BRANCO)
        c.fill = fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_ALL_THIN
    ws.row_dimensions[17].height = 30

    # ------------------------------------------------------------------
    # Chart: economy per scenario
    # ------------------------------------------------------------------
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Economia Anual por Cenario"
    chart.y_axis.title = "R$"
    data_ref = Reference(ws, min_col=5, min_row=5, max_col=7, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "B19")

    # Note
    ws.merge_cells("B30:K33")
    ws["B30"] = (
        "NOTA: Os valores de economia sao simulacoes a partir das premissas "
        "editaveis. Substitua os valores por dados reais da sua planta para "
        "obter projecoes confiaveis. O cenario Realista (10%) e o ponto de "
        "calibracao sugerido para apresentacao a diretoria."
    )
    ws["B30"].alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True, indent=1)
    ws["B30"].style = "exa_note"

    page_setup_landscape(ws)
    return ws
