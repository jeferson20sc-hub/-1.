"""DASHBOARD sheet: executive premium panel with KPIs and charts."""

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import (
    CellIsRule, ColorScaleRule, FormulaRule, IconSetRule,
)

from ..theme import (
    AMBAR, AMBAR_FUNDO, AZUL_KPI, BORDER_ALL_THIN, BRANCO, CINZA_BORDA,
    CINZA_CLARO, FMT_BRL, FMT_BRL_MI, FMT_DATAHORA, FMT_HORAS, FMT_INT,
    FMT_PCT, GRAFITE, PETROLEO, PETROLEO_CLARO, VERDE, VERDE_FUNDO,
    VERMELHO, VERMELHO_FUNDO, fill,
)
from ..util import (
    block_fill,
    hide_gridlines,
    page_setup_landscape,
    section_header,
    set_column_widths,
    set_row_heights,
    thin_border_range,
    title_bar,
)


def build(wb):
    ws = wb.create_sheet("DASHBOARD")
    hide_gridlines(ws)

    # Compact, controlled grid - no oversized empty space
    set_column_widths(ws, {
        "A": 2, "B": 18, "C": 18, "D": 4, "E": 18, "F": 18, "G": 4,
        "H": 18, "I": 18, "J": 4, "K": 18, "L": 18, "M": 2,
    })
    title_bar(
        ws,
        "EXAUSTAO 360 - PAINEL EXECUTIVO DE CONFIABILIDADE",
        "Atualize com o botao [Atualizar Dashboard]. Periodo, plantas e parametros sao definidos em CONFIG.",
        last_col="M",
    )

    # ------------------------------------------------------------------
    # Header info row (period + last update)
    # ------------------------------------------------------------------
    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:F3")
    ws["B3"] = "Periodo:"
    ws["B3"].font = Font(name="Calibri", size=10, bold=True, color=PETROLEO)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("H3:L3")
    ws["H3"] = "Ultima atualizacao:"
    ws["H3"].font = Font(name="Calibri", size=10, bold=True, color=PETROLEO)
    ws["H3"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # KPI section
    section_header(ws, "B5:L5", "Indicadores Executivos (KPIs)")

    # Row 6: KPI labels (4 cards x 3 = 12)
    # Row 7: KPI values
    # Row 8: spacer
    # Row 9-10: second KPI row
    # Row 11-12: third KPI row
    set_row_heights(ws, {6: 22, 7: 42, 8: 6, 9: 22, 10: 42, 11: 6, 12: 22, 13: 42})

    kpis = [
        # (label_cell, value_cell, label, formula, fmt, style)
        ("B6",  "B7",  "Disponibilidade",
         '=IFERROR(1-SUMPRODUCT(tbEventos[DuracaoHoras])/(K_HORAS_PERIODO),0)', FMT_PCT, "exa_kpi_value_green"),
        ("E6",  "E7",  "MTBF (horas)",
         '=IFERROR(K_HORAS_PERIODO/MAX(COUNTA(tbEventos[ID]),1),0)', FMT_HORAS, "exa_kpi_value"),
        ("H6",  "H7",  "MTTR (horas)",
         '=IFERROR(AVERAGE(tbEventos[DuracaoHoras]),0)', FMT_HORAS, "exa_kpi_value"),
        ("K6",  "K7",  "Eventos no Periodo",
         '=COUNTA(tbEventos[ID])', FMT_INT, "exa_kpi_value"),

        ("B9",  "B10", "Custo Total Falhas",
         '=SUMPRODUCT(tbEventos[CustoEstimado])', FMT_BRL, "exa_kpi_value_red"),
        ("E9",  "E10", "Custo em Milhoes",
         '=SUMPRODUCT(tbEventos[CustoEstimado])', FMT_BRL_MI, "exa_kpi_value_red"),
        ("H9",  "H10", "Forno Critico",
         '=IFERROR(INDEX(tbEventos[Forno],MATCH(MAX(tbEventos[CustoEstimado]),tbEventos[CustoEstimado],0)),"-")', None, "exa_kpi_value_red"),
        ("K9",  "K10", "Eventos Criticos",
         '=COUNTIF(tbEventos[Criticidade],"Critica")', FMT_INT, "exa_kpi_value_red"),

        ("B12", "B13", "Economia Potencial Anual",
         '=ECONOMIA_ANUAL', FMT_BRL_MI, "exa_kpi_value_green"),
        ("E12", "E13", "ROI Estimado",
         '=ROI_ESTIMADO', FMT_PCT, "exa_kpi_value_green"),
        ("H12", "H13", "Valor Protegido",
         '=VALOR_PROTEGIDO', FMT_BRL_MI, "exa_kpi_value_green"),
        ("K12", "K13", "Semaforo Operacional",
         '=SEMAFORO', None, "exa_kpi_value"),
    ]
    for label_cell, value_cell, label, formula, fmt, style in kpis:
        ws[label_cell] = label
        ws[label_cell].style = "exa_kpi_label"
        ws[value_cell] = formula
        ws[value_cell].style = style
        if fmt:
            ws[value_cell].number_format = fmt

    # ------------------------------------------------------------------
    # Pareto + Ranking
    # ------------------------------------------------------------------
    section_header(ws, "B15:F15", "Pareto de Falhas por Componente")
    section_header(ws, "H15:L15", "Ranking de Fornos por Risco")

    # Pareto data (manually aggregated by component using formulas)
    pareto_headers = ["Componente", "Eventos", "Custo Total", "% Acumulado"]
    # We'll list the 6 main components from CONFIG and reference tbEventos
    componentes = ["Tubulacao", "Bujao Refratario", "Marmita", "Manta Isolante",
                   "Queimador", "Ventilador de Exaustao"]
    for i, h in enumerate(pareto_headers):
        c = ws.cell(row=16, column=2 + i, value=h)
        c.style = "exa_th"
    ws.row_dimensions[16].height = 22

    for idx, comp in enumerate(componentes):
        r = 17 + idx
        ws.cell(row=r, column=2, value=comp).style = "exa_td"
        ws.cell(row=r, column=3,
                value=f'=COUNTIF(tbEventos[Componente],B{r})').style = "exa_td"
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f'=SUMIF(tbEventos[Componente],B{r},tbEventos[CustoEstimado])').style = "exa_td"
        ws.cell(row=r, column=4).number_format = FMT_BRL
        # cumulative pct
        ws.cell(row=r, column=5,
                value=f'=IFERROR(SUM(D$17:D{r})/SUM(D$17:D$22),0)').style = "exa_td"
        ws.cell(row=r, column=5).number_format = FMT_PCT
        ws.row_dimensions[r].height = 18

    # Ranking de Fornos
    rank_headers = ["Forno", "Eventos", "Custo Total", "Risco"]
    for i, h in enumerate(rank_headers):
        c = ws.cell(row=16, column=8 + i, value=h)
        c.style = "exa_th"

    fornos = ["F-101", "F-102", "F-201", "F-202", "F-301", "F-302"]
    for idx, forno in enumerate(fornos):
        r = 17 + idx
        ws.cell(row=r, column=8, value=forno).style = "exa_td"
        ws.cell(row=r, column=9,
                value=f'=COUNTIF(tbEventos[Forno],H{r})').style = "exa_td"
        ws.cell(row=r, column=9).number_format = FMT_INT
        ws.cell(row=r, column=10,
                value=f'=SUMIF(tbEventos[Forno],H{r},tbEventos[CustoEstimado])').style = "exa_td"
        ws.cell(row=r, column=10).number_format = FMT_BRL
        ws.cell(row=r, column=11,
                value=f'=IF(J{r}>500000,"Critico",IF(J{r}>200000,"Alto",IF(J{r}>50000,"Medio","Baixo")))').style = "exa_td"
        ws.cell(row=r, column=11).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    # Conditional formatting for Risco column (K17:K22)
    ws.conditional_formatting.add(
        "K17:K22",
        FormulaRule(formula=['$K17="Critico"'],
                    fill=fill(VERMELHO_FUNDO),
                    font=Font(color=VERMELHO, bold=True))
    )
    ws.conditional_formatting.add(
        "K17:K22",
        FormulaRule(formula=['$K17="Alto"'],
                    fill=fill(AMBAR_FUNDO),
                    font=Font(color=AMBAR, bold=True))
    )
    ws.conditional_formatting.add(
        "K17:K22",
        FormulaRule(formula=['$K17="Medio"'],
                    fill=fill("FFFFFAE6"),
                    font=Font(color="FFB28704", bold=True))
    )
    ws.conditional_formatting.add(
        "K17:K22",
        FormulaRule(formula=['$K17="Baixo"'],
                    fill=fill(VERDE_FUNDO),
                    font=Font(color=VERDE, bold=True))
    )

    # ------------------------------------------------------------------
    # Trend / Recommendations
    # ------------------------------------------------------------------
    section_header(ws, "B24:L24", "Evolucao Mensal e Recomendacoes")

    # Tendencia mensal table (small)
    meses = ["Jan/26", "Fev/26", "Mar/26", "Abr/26", "Mai/26"]
    months_num = [1, 2, 3, 4, 5]
    ws.cell(row=25, column=2, value="Mes").style = "exa_th"
    ws.cell(row=25, column=3, value="Eventos").style = "exa_th"
    ws.cell(row=25, column=4, value="Custo (R$)").style = "exa_th"
    ws.cell(row=25, column=5, value="MTTR (h)").style = "exa_th"
    ws.cell(row=25, column=6, value="Disponib.").style = "exa_th"
    ws.row_dimensions[25].height = 22

    for i, (m, mn) in enumerate(zip(meses, months_num)):
        r = 26 + i
        ws.cell(row=r, column=2, value=m).style = "exa_td"
        ws.cell(row=r, column=3,
                value=f'=SUMPRODUCT((MONTH(tbEventos[Data])={mn})*1)').style = "exa_td"
        ws.cell(row=r, column=3).number_format = FMT_INT
        ws.cell(row=r, column=4,
                value=f'=SUMPRODUCT((MONTH(tbEventos[Data])={mn})*tbEventos[CustoEstimado])').style = "exa_td"
        ws.cell(row=r, column=4).number_format = FMT_BRL
        ws.cell(row=r, column=5,
                value=f'=IFERROR(AVERAGEIFS(tbEventos[DuracaoHoras],tbEventos[Data],">="&DATE(2026,{mn},1),tbEventos[Data],"<"&DATE(2026,{mn}+1,1)),0)').style = "exa_td"
        ws.cell(row=r, column=5).number_format = FMT_HORAS
        ws.cell(row=r, column=6,
                value=f'=IFERROR(1-SUMPRODUCT((MONTH(tbEventos[Data])={mn})*tbEventos[DuracaoHoras])/(24*30),0)').style = "exa_td"
        ws.cell(row=r, column=6).number_format = FMT_PCT
        ws.row_dimensions[r].height = 18

    # Color scale on cost column
    ws.conditional_formatting.add(
        "D26:D30",
        ColorScaleRule(
            start_type="min", start_color="FFE8F5E9",
            mid_type="percentile", mid_value=50, mid_color="FFFFF4E1",
            end_type="max", end_color="FFFDECEC",
        ),
    )

    # Recommendations panel
    ws.cell(row=25, column=8, value="Recomendacoes Automaticas").style = "exa_th"
    ws.merge_cells("H25:L25")

    recommendations = [
        '=IF(B7<K_META_DISPONIBILIDADE,"Acao: Disponibilidade abaixo da meta. Revisar plano preventivo.","OK: Disponibilidade dentro da meta.")',
        '=IF(H7>K_META_MTTR,"Acao: MTTR alto. Treinar equipe e revisar peças sobressalentes.","OK: MTTR dentro da meta.")',
        '=IF(E7<K_META_MTBF,"Acao: MTBF abaixo da meta. Investigar componentes recorrentes.","OK: MTBF dentro da meta.")',
        '=IF(K9>0,"Atencao: existem eventos criticos. Revisar plano de acao em ANALISE.","Sem eventos criticos no periodo.")',
        '=IF(B10>K_LIMIAR_CUSTO,"Acao: Custo total elevado. Avaliar pacote PRO em MODELO_COMERCIAL.","Custo total sob controle.")',
    ]
    for i, formula in enumerate(recommendations):
        r = 26 + i
        ws.merge_cells(f"H{r}:L{r}")
        c = ws.cell(row=r, column=8, value=formula)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.font = Font(name="Calibri", size=10, color=GRAFITE)
        c.fill = fill(BRANCO)
        c.border = BORDER_ALL_THIN
        ws.row_dimensions[r].height = 18

    # ------------------------------------------------------------------
    # Charts
    # ------------------------------------------------------------------
    pareto_chart = BarChart()
    pareto_chart.type = "col"
    pareto_chart.title = "Pareto de Falhas - Componentes"
    pareto_chart.style = 11
    pareto_chart.y_axis.title = "Custo Total (R$)"
    pareto_chart.x_axis.title = "Componente"
    data_ref = Reference(ws, min_col=4, min_row=16, max_col=4, max_row=22)
    cats_ref = Reference(ws, min_col=2, min_row=17, max_col=2, max_row=22)
    pareto_chart.add_data(data_ref, titles_from_data=True)
    pareto_chart.set_categories(cats_ref)
    pareto_chart.height = 7
    pareto_chart.width = 16
    pareto_chart.legend = None
    pareto_chart.dataLabels = DataLabelList(showVal=False)
    ws.add_chart(pareto_chart, "B33")

    # Trend chart (line)
    trend_chart = LineChart()
    trend_chart.title = "Evolucao Mensal - Custo de Falhas"
    trend_chart.style = 12
    trend_chart.y_axis.title = "Custo (R$)"
    trend_chart.x_axis.title = "Mes"
    t_data = Reference(ws, min_col=4, min_row=25, max_col=4, max_row=30)
    t_cats = Reference(ws, min_col=2, min_row=26, max_col=2, max_row=30)
    trend_chart.add_data(t_data, titles_from_data=True)
    trend_chart.set_categories(t_cats)
    trend_chart.height = 7
    trend_chart.width = 16
    ws.add_chart(trend_chart, "H33")

    # Ranking pie chart
    pie = PieChart()
    pie.title = "Distribuicao de Custos por Forno"
    p_data = Reference(ws, min_col=10, min_row=16, max_col=10, max_row=22)
    p_cats = Reference(ws, min_col=8, min_row=17, max_col=8, max_row=22)
    pie.add_data(p_data, titles_from_data=True)
    pie.set_categories(p_cats)
    pie.height = 8
    pie.width = 12
    pie.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(pie, "B49")

    # Footer print area / page setup
    ws.print_area = "A1:M65"
    page_setup_landscape(ws)

    return ws
