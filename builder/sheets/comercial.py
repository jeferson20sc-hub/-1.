"""MODELO_COMERCIAL sheet: pricing packages, ROI argumentation."""

from openpyxl.styles import Alignment, Font

from ..theme import (
    BORDER_ALL_THIN, BRANCO, CINZA_CLARO, FMT_BRL, FMT_BRL_MI, FMT_PCT,
    GRAFITE, PETROLEO, PETROLEO_CLARO, VERDE, VERMELHO, fill,
)
from ..util import (
    hide_gridlines,
    page_setup_landscape,
    section_header,
    set_column_widths,
    set_row_heights,
    title_bar,
)


def build(wb):
    ws = wb.create_sheet("MODELO_COMERCIAL")
    hide_gridlines(ws)
    set_column_widths(ws, {
        "A": 2, "B": 28, "C": 26, "D": 26, "E": 26, "F": 4,
        "G": 32, "H": 22,
    })
    title_bar(
        ws,
        "EXAUSTAO 360 - MODELO COMERCIAL B2B",
        "Proposta de valor, pacotes e calculo de ROI. Edite os campos para gerar uma proposta personalizada.",
        last_col="H",
    )

    # ------------------------------------------------------------------
    # Proposta de valor
    # ------------------------------------------------------------------
    section_header(ws, "B4:H4", "Proposta de Valor")
    ws.merge_cells("B5:H7")
    ws["B5"] = (
        "EXAUSTAO 360 reduz perdas operacionais em sistemas industriais de exaustao "
        "por meio de gestao integrada de eventos, analise de criticidade (FMEA), "
        "MTBF/MTTR e priorizacao de planos de acao com impacto financeiro mensuravel. "
        "Cada R$ investido na solucao busca retornar entre R$ 5 e R$ 25 em economia "
        "por planta industrial - sujeito a disciplina operacional e qualidade de dados."
    )
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=2)
    ws["B5"].font = Font(name="Calibri", size=11, color=GRAFITE)
    ws["B5"].fill = fill(BRANCO)
    ws["B5"].border = BORDER_ALL_THIN
    set_row_heights(ws, {5: 22, 6: 22, 7: 22})

    # ------------------------------------------------------------------
    # Problema resolvido
    # ------------------------------------------------------------------
    section_header(ws, "B9:H9", "Problema Industrial Resolvido")
    problemas = [
        "Falhas em fornos, dutos, bujoes e mantas geram parada nao programada.",
        "Custos invisiveis: cada hora parada custa entre R$ 10.000 e R$ 80.000 dependendo da planta.",
        "Decisoes baseadas em planilhas isoladas, sem rastreabilidade ou auditoria.",
        "Falta de evidencia de ROI dificulta aprovacao de investimentos em confiabilidade.",
        "Equipes de manutencao operam reativamente, sem priorizacao por risco financeiro.",
    ]
    for i, p in enumerate(problemas):
        r = 10 + i
        ws.merge_cells(f"B{r}:H{r}")
        ws.cell(row=r, column=2, value=f"  -  {p}")
        ws.cell(row=r, column=2).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, color=GRAFITE)
        ws.row_dimensions[r].height = 18

    # ------------------------------------------------------------------
    # Pacotes
    # ------------------------------------------------------------------
    section_header(ws, "B17:E17", "Pacotes de Implantacao")

    # Header row
    pkg_headers = ["Item", "Diagnostico", "Pro", "Enterprise"]
    for i, h in enumerate(pkg_headers):
        c = ws.cell(row=18, column=2 + i, value=h)
        c.style = "exa_th"
    ws.row_dimensions[18].height = 24

    # Package rows
    pacotes = [
        ("Escopo", "1 planta - diagnostico inicial",
         "1 planta - implantacao completa",
         "Multiunidade - corporativo"),
        ("Duracao", "30 dias", "60 dias", "120 dias"),
        ("Suporte", "30 dias", "12 meses", "24 meses + SLA"),
        ("Customizacao", "Layout padrao", "Limitada", "Total"),
        ("Treinamento", "1 turma", "3 turmas", "Trilha completa"),
        ("Auditoria/LOG", "Basica", "Completa", "Multi-usuario"),
        ("Modo Supervisor", "Sim", "Sim", "Sim + RBAC"),
        ("Relatorio Executivo", "Mensal", "Mensal", "Quinzenal"),
        ("Atualizacoes", "Nao incluidas", "Inclusas (12m)", "Inclusas (24m)"),
        ("Preco Sugerido (R$)",
         '=INDEX(tbParametros[Valor],MATCH("CustoSolucaoPro",tbParametros[Parametro],0))*0.3',
         '=INDEX(tbParametros[Valor],MATCH("CustoSolucaoPro",tbParametros[Parametro],0))',
         '=INDEX(tbParametros[Valor],MATCH("CustoSolucaoPro",tbParametros[Parametro],0))*3'),
    ]
    for i, row in enumerate(pacotes):
        r = 19 + i
        ws.cell(row=r, column=2, value=row[0]).style = "exa_label"
        for j, v in enumerate(row[1:]):
            c = ws.cell(row=r, column=3 + j, value=v)
            c.style = "exa_td"
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22

    # Format last row (preco) as BRL
    preco_row = 19 + len(pacotes) - 1
    for col in (3, 4, 5):
        ws.cell(row=preco_row, column=col).number_format = FMT_BRL
        ws.cell(row=preco_row, column=col).font = Font(
            name="Calibri", size=12, bold=True, color=PETROLEO
        )

    # Highlight Pro column
    for r in range(18, preco_row + 1):
        ws.cell(row=r, column=4).fill = fill("FFE6F0FA")

    # ------------------------------------------------------------------
    # Calculadora de ROI personalizada
    # ------------------------------------------------------------------
    roi_start = preco_row + 3
    section_header(ws, f"B{roi_start}:H{roi_start}", "Calculadora de ROI Personalizada")
    set_row_heights(ws, {roi_start: 26})

    # Inputs (left side)
    roi_inputs = [
        ("Perda anual atual da planta (R$)",          2000000, FMT_BRL),
        ("% reducao esperada com a solucao",          0.10,    FMT_PCT),
        ("Preco do pacote selecionado (R$)",          80000,   FMT_BRL),
        ("Numero de plantas",                         1,       "0"),
    ]
    for i, (label, default, fmt) in enumerate(roi_inputs):
        r = roi_start + 2 + i
        ws.cell(row=r, column=2, value=label).style = "exa_label"
        c = ws.cell(row=r, column=3, value=default)
        c.style = "exa_input"
        c.number_format = fmt
        ws.row_dimensions[r].height = 22

    # Outputs (right side) referencing the inputs above by relative row
    perda_cell = f"C{roi_start + 2}"
    pct_cell = f"C{roi_start + 3}"
    preco_cell = f"C{roi_start + 4}"
    plantas_cell = f"C{roi_start + 5}"

    roi_outputs = [
        ("Economia anual por planta",
         f"={perda_cell}*{pct_cell}", FMT_BRL),
        ("Economia anual em milhoes",
         f"={perda_cell}*{pct_cell}", FMT_BRL_MI),
        ("Custo total de aquisicao",
         f"={preco_cell}*{plantas_cell}", FMT_BRL),
        ("Payback (anos)",
         f"=IFERROR({preco_cell}/({perda_cell}*{pct_cell}),0)",
         '0.00" anos"'),
        ("ROI Anual",
         f"=IFERROR(({perda_cell}*{pct_cell}-{preco_cell})/{preco_cell},0)",
         FMT_PCT),
        ("Economia em 3 anos",
         f"={perda_cell}*{pct_cell}*3*{plantas_cell}", FMT_BRL_MI),
        ("Argumento Comercial",
         f'=IF({perda_cell}*{pct_cell}>{preco_cell},"Justificavel: economia anual supera o custo da solucao.","Avaliar premissas - economia projetada e inferior ao custo.")',
         None),
    ]
    for i, (label, formula, fmt) in enumerate(roi_outputs):
        r = roi_start + 2 + i
        ws.cell(row=r, column=7, value=label).style = "exa_label"
        c = ws.cell(row=r, column=8, value=formula)
        c.style = "exa_td"
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.font = Font(name="Calibri", size=11, bold=True, color=PETROLEO)
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[r].height = 22

    # ------------------------------------------------------------------
    # Argumento de venda (exemplo)
    # ------------------------------------------------------------------
    arg_start = roi_start + 2 + max(len(roi_inputs), len(roi_outputs)) + 2
    section_header(ws, f"B{arg_start}:H{arg_start}", "Exemplo de Argumento Comercial")
    ws.merge_cells(f"B{arg_start + 1}:H{arg_start + 4}")
    ws[f"B{arg_start + 1}"] = (
        "Se uma planta perde R$ 2.000.000/ano com falhas de exaustao e a solucao "
        "reduz apenas 10% dessas perdas, o valor recuperado e R$ 200.000/ano. "
        "Um contrato de R$ 40.000 a R$ 80.000 pode ser justificavel se houver "
        "evidencia operacional de melhoria.\n\n"
        "IMPORTANTE: a economia real depende dos dados da planta, da disciplina "
        "operacional e da qualidade do registro. EXAUSTAO 360 fornece a "
        "infraestrutura para mensurar e comprovar essa economia."
    )
    ws[f"B{arg_start + 1}"].alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True, indent=2
    )
    ws[f"B{arg_start + 1}"].font = Font(
        name="Calibri", size=11, italic=True, color=GRAFITE
    )
    ws[f"B{arg_start + 1}"].fill = fill(CINZA_CLARO)
    ws[f"B{arg_start + 1}"].border = BORDER_ALL_THIN
    for r in range(arg_start + 1, arg_start + 5):
        ws.row_dimensions[r].height = 22

    page_setup_landscape(ws)
    return ws
