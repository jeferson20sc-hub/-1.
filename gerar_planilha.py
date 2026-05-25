"""Gera Gestao_Estoque_LIS.xlsx — modelo para o SharePoint/OneDrive online.

Estrutura compatível com o fluxo do Power Automate:
  - Aba 'Base de lancamentos' com a tabela 'Tabela'
  - Colunas: ID, Hora, Data, Operador, Codigo, Descricao, Tipo, Quantidade,
             Saldo, EstoqueSeg, PontoPedido, Status, Origem
  - Aba 'Dashboards' com KPIs + 4 gráficos atualizados em tempo real conforme
    o Power Automate insere linhas (usa SUMIFS/COUNTIFS sobre colunas inteiras).
  - Aba 'Itens' com catálogo base.

Como atualizar o online: rodar este script, baixar o .xlsx e substituir o
arquivo no OneDrive. Em seguida, no passo 'Adicionar uma linha em uma tabela'
do Power Automate, reapontar Arquivo + Tabela (Tabela).
"""
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

# Layout exato esperado pelo fluxo (campos minúsculos do gatilho mapeados
# para essas colunas no passo 'Adicionar uma linha em uma tabela').
COLUNAS = [
    ("ID", 24),
    ("Hora", 14),
    ("Data", 22),
    ("Operador", 16),
    ("Codigo", 10),
    ("Descricao", 42),
    ("Tipo", 12),
    ("Quantidade", 12),
    ("Saldo", 12),
    ("EstoqueSeg", 13),
    ("PontoPedido", 13),
    ("Status", 18),
    ("Origem", 12),
]
EXEMPLO = [
    "LIS-EXEMPLO-001", "12:00:00", "19/05/2026, 12:00:00",
    "Jeferson", "001", "Filtro de óleo (loja de peças)", "Entrada",
    10, 130, 40, 140, "🟢 OK", "Site",
]
ITENS = [
    ("001", "Filtro de óleo (loja de peças)", "un", "Não", 40, 140),
    ("002", "Arroz (restaurante)", "kg", "Sim", 10, 55),
    ("003", "Chapas de MDF (fábrica de móveis)", "un", "Não", 100, 300),
    ("004", "Polivitamínico (farmácia)", "un", "Sim", 30, 100),
    ("005", "Resmas de papel A4 (papelaria)", "un", "Não", 150, 350),
    ("006", "Essência cosmética (cosméticos)", "L", "Sim", 50, 150),
    ("007", "Componente crítico (peças críticas)", "un", "Não", 150, 450),
    ("008", "Laptops (varejo eletrônicos)", "un", "Não", 50, 200),
    ("009", "Kits de luvas hospitalar (NR-32)", "cx", "Sim", 2000, 7000),
    ("010", "Insumo industrial", "un", "Não", 200, 600),
]

azul = PatternFill("solid", fgColor="00529B")
verde = PatternFill("solid", fgColor="107C41")
amarelo = PatternFill("solid", fgColor="FCD34D")
vermelho = PatternFill("solid", fgColor="DC2626")
cinza = PatternFill("solid", fgColor="F1F5F9")
cinzaEsc = PatternFill("solid", fgColor="E5E7EB")
branco = Font(color="FFFFFF", bold=True, size=11)
titulo = Font(color="00529B", bold=True, size=16)
sub = Font(color="334155", bold=True, size=11)
borda = Border(*[Side(style="thin", color="D9E2EC")] * 4)


def cabec(cell, cor=azul):
    cell.fill = cor
    cell.font = branco
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda


wb = Workbook()

# ========================= ABA 1: Base de lancamentos =========================
ws = wb.active
ws.title = "Base de lancamentos"
ws.sheet_view.showGridLines = False
ws.append([c[0] for c in COLUNAS])
ws.append(EXEMPLO)
for ci, (nome, larg) in enumerate(COLUNAS, 1):
    L = get_column_letter(ci)
    ws.column_dimensions[L].width = larg
    cabec(ws.cell(row=1, column=ci))
    d = ws.cell(row=2, column=ci)
    d.border = borda
    d.alignment = Alignment(
        horizontal="center" if nome != "Descricao" else "left", vertical="center"
    )
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"
ref = f"A1:{get_column_letter(len(COLUNAS))}2"
tab = Table(displayName="Tabela", ref=ref)
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False,
)
ws.add_table(tab)
ws.auto_filter.ref = ref

# ============================ ABA 2: Itens (catálogo) ========================
wi = wb.create_sheet("Itens")
wi.sheet_view.showGridLines = False
itens_cols = [
    ("Codigo", 10), ("Descricao", 42), ("Unidade", 12),
    ("VenceRapido", 13), ("EstoqueSeg", 13), ("PontoPedido", 13),
]
wi.append([c[0] for c in itens_cols])
for it in ITENS:
    wi.append(list(it))
for ci, (nome, larg) in enumerate(itens_cols, 1):
    L = get_column_letter(ci)
    wi.column_dimensions[L].width = larg
    cabec(wi.cell(row=1, column=ci), cor=verde)
wi.row_dimensions[1].height = 24
wi.freeze_panes = "A2"
itab = Table(displayName="Itens", ref=f"A1:F{1+len(ITENS)}")
itab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
wi.add_table(itab)
wi.auto_filter.ref = f"A1:F{1+len(ITENS)}"

# =============================== ABA 3: Dashboards ===========================
wd = wb.create_sheet("Dashboards")
wd.sheet_view.showGridLines = False
wd["B2"] = "DASHBOARD — GESTÃO DE ESTOQUE LIS"
wd["B2"].font = titulo
wd.merge_cells("B2:K2")
wd["B3"] = "Atualiza automaticamente conforme o Power Automate insere linhas em 'Base de lancamentos'."
wd["B3"].font = sub
wd.merge_cells("B3:K3")

# --- KPIs (linha 5) ---
kpis = [
    ("📋 Movimentações",
     '=COUNTA(Tabela[ID])',
     verde),
    ("➕ Total Entradas",
     '=SUMIFS(Tabela[Quantidade],Tabela[Tipo],"Entrada")',
     azul),
    ("➖ Total Saídas",
     '=SUMIFS(Tabela[Quantidade],Tabela[Tipo],"Saída")',
     amarelo),
    ("👥 Operadores",
     '=SUMPRODUCT((Tabela[Operador]<>"")/COUNTIF(Tabela[Operador],Tabela[Operador]&""))',
     azul),
    ("🟡 Itens p/ pedir",
     '=COUNTIF(Tabela[Status],"*PEDIR*")',
     amarelo),
    ("🔴 Itens críticos",
     '=COUNTIF(Tabela[Status],"*CRÍTICO*")',
     vermelho),
]
for i, (rot, frm, cor) in enumerate(kpis):
    c = 2 + i
    rotCell = wd.cell(row=5, column=c, value=rot)
    rotCell.fill = cor
    rotCell.font = Font(color="FFFFFF", bold=True, size=11)
    rotCell.alignment = Alignment(horizontal="center", vertical="center")
    rotCell.border = borda
    valCell = wd.cell(row=6, column=c, value=frm)
    valCell.fill = cinza
    valCell.font = Font(bold=True, size=18, color="0F172A")
    valCell.alignment = Alignment(horizontal="center", vertical="center")
    valCell.border = borda
    wd.column_dimensions[get_column_letter(c)].width = 18
wd.row_dimensions[5].height = 26
wd.row_dimensions[6].height = 36

# --- Helper area 1: últimos 14 dias (linhas 9..22, M..O) ---
wd["M8"] = "Data"; wd["N8"] = "Entradas"; wd["O8"] = "Saídas"
for c in ("M8", "N8", "O8"):
    cabec(wd[c])
for i in range(14):
    r = 9 + i
    wd.cell(row=r, column=13, value=f"=TODAY()-{13-i}").number_format = "dd/mm"
    # Faz match começando "dd/mm/yyyy" (o site grava "dd/mm/yyyy, HH:MM:SS").
    wd.cell(row=r, column=14,
            value=('=SUMPRODUCT((LEFT(Tabela[Data],10)=TEXT($M' + str(r) +
                   ',"dd/mm/yyyy"))*(Tabela[Tipo]="Entrada")*Tabela[Quantidade])'))
    wd.cell(row=r, column=15,
            value=('=SUMPRODUCT((LEFT(Tabela[Data],10)=TEXT($M' + str(r) +
                   ',"dd/mm/yyyy"))*(Tabela[Tipo]="Saída")*Tabela[Quantidade])'))
for col in ("M", "N", "O"):
    wd.column_dimensions[col].width = 14

# --- Helper area 2: saídas por item (linhas 9..18, Q..R) ---
wd["Q8"] = "Item"; wd["R8"] = "Saídas"
for c in ("Q8", "R8"):
    cabec(wd[c])
for i, it in enumerate(ITENS):
    r = 9 + i
    wd.cell(row=r, column=17, value=it[0])
    wd.cell(row=r, column=18,
            value=f'=SUMIFS(Tabela[Quantidade],Tabela[Codigo],$Q{r},Tabela[Tipo],"Saída")')
wd.column_dimensions["Q"].width = 10
wd.column_dimensions["R"].width = 12

# --- Helper area 3: saldo atual por item (linhas 9..18, T..U) ---
wd["T8"] = "Item"; wd["U8"] = "Saldo"
for c in ("T8", "U8"):
    cabec(wd[c])
for i, it in enumerate(ITENS):
    r = 9 + i
    wd.cell(row=r, column=20, value=it[0])
    wd.cell(row=r, column=21, value=(
        f'=SUMIFS(Tabela[Quantidade],Tabela[Codigo],$T{r},Tabela[Tipo],"Entrada")'
        f'-SUMIFS(Tabela[Quantidade],Tabela[Codigo],$T{r},Tabela[Tipo],"Saída")'
    ))
wd.column_dimensions["T"].width = 10
wd.column_dimensions["U"].width = 12

# --- Helper area 4: acumulado de movimentações por dia (V..W) ---
wd["V8"] = "Data"; wd["W8"] = "Mov. acumulado"
for c in ("V8", "W8"):
    cabec(wd[c])
for i in range(14):
    r = 9 + i
    wd.cell(row=r, column=22, value=f"=TODAY()-{13-i}").number_format = "dd/mm"
    wd.cell(row=r, column=23, value=(
        '=SUMPRODUCT((LEFT(Tabela[Data],10)<=TEXT($V' + str(r) + ',"dd/mm/yyyy"))*1)'
    ))
wd.column_dimensions["V"].width = 14
wd.column_dimensions["W"].width = 16

# ========================== GRÁFICOS ==========================
# Gráfico 1: barras entradas vs saídas — últimos 14 dias
bar = BarChart()
bar.type = "col"
bar.style = 11
bar.grouping = "clustered"
bar.title = "Entradas vs Saídas — últimos 14 dias"
bar.y_axis.title = "Quantidade"
bar.x_axis.title = "Dia"
bar.add_data(Reference(wd, min_col=14, max_col=15, min_row=8, max_row=22),
             titles_from_data=True)
bar.set_categories(Reference(wd, min_col=13, min_row=9, max_row=22))
bar.height = 9; bar.width = 18
wd.add_chart(bar, "B9")

# Gráfico 2: pizza saídas por item
pie = PieChart()
pie.title = "Saídas por item"
pie.add_data(Reference(wd, min_col=18, min_row=8, max_row=18),
             titles_from_data=True)
pie.set_categories(Reference(wd, min_col=17, min_row=9, max_row=18))
pie.dataLabels = DataLabelList(showPercent=True)
pie.height = 9; pie.width = 13
wd.add_chart(pie, "B27")

# Gráfico 3: barras horizontais saldo atual por item
bar2 = BarChart()
bar2.type = "bar"; bar2.style = 12
bar2.title = "Saldo atual por item (Entrada − Saída)"
bar2.y_axis.title = "Item"
bar2.x_axis.title = "Saldo"
bar2.add_data(Reference(wd, min_col=21, min_row=8, max_row=18),
              titles_from_data=True)
bar2.set_categories(Reference(wd, min_col=20, min_row=9, max_row=18))
bar2.height = 9; bar2.width = 18
wd.add_chart(bar2, "H27")

# Gráfico 4: linha acumulada de movimentações
line = LineChart()
line.title = "Total de movimentações — acumulado por dia"
line.style = 13
line.y_axis.title = "Total acumulado"
line.x_axis.title = "Dia"
line.add_data(Reference(wd, min_col=23, min_row=8, max_row=22),
              titles_from_data=True)
line.set_categories(Reference(wd, min_col=22, min_row=9, max_row=22))
line.height = 9; line.width = 18
wd.add_chart(line, "H9")

# Larguras das colunas vazias do dashboard (visual)
for col, larg in [("A", 2), ("B", 16), ("C", 16), ("D", 16),
                  ("E", 16), ("F", 16), ("G", 16), ("H", 16),
                  ("I", 16), ("J", 16), ("K", 16)]:
    wd.column_dimensions[col].width = larg

wb.save("/home/user/senai-cba/Gestao_Estoque_LIS.xlsx")
print("OK: workbook gerado")
print("  Abas:", wb.sheetnames)
print("  Tabela principal: 'Tabela' em 'Base de lancamentos' (13 colunas)")
print("  Dashboards: 6 KPIs + 4 gráficos automáticos")
