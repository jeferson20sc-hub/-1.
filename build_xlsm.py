"""
Gerador do arquivo Controle_Fornos.xlsm

Estrategia de PERFORMANCE:
- Apenas 3 abas visiveis (Dashboard, Registro, Filtros) + 1 oculta (_Dados)
- Tabela Excel nativa (ListObject) - referencia estruturada, rapida
- Formulas COUNTIFS/SUMIFS (nao volateis, otimizadas pelo motor do Excel)
- Sem OFFSET/INDIRECT/NOW (volateis - recalculam a cada acao)
- 150 linhas para resultado de filtro (suficiente, nao trava o arquivo)
- Charts apontando para tabelas auxiliares pequenas
- Application.Calculation=Manual durante VBA Register (impede recalculo intermediario)
- VBA minimo (3 macros: Registrar, LimparCampos, LimparFiltros)
- Modo de operacao alternativo SEM VBA: digitar direto na tabela em _Dados
"""

import os
import struct
import zipfile
import shutil
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation


OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(OUT_DIR, "Controle_Fornos.xlsx")
OUT_XLSM = os.path.join(OUT_DIR, "Controle_Fornos.xlsm")
OUT_BAS = os.path.join(OUT_DIR, "modFornos.bas")

TIPOS = ["Operacao", "Manutencao", "Falha", "Inspecao"]
LADOS = ["A", "B"]

# Cores
COR_PRIMARIA = "1F4E78"
COR_SECUNDARIA = "2E75B6"
COR_HEADER_TXT = "FFFFFF"
COR_DESTAQUE = "C00000"
COR_OK = "548235"

CORES_TIPOS = {
    "Operacao":    "2E75B6",
    "Manutencao":  "BF9000",
    "Falha":       "C00000",
    "Inspecao":    "548235",
}

# Limites de performance
MAX_FILTRO_LINHAS = 150     # linhas no resultado da aba Filtros
MAX_REGISTROS_FORNO = 20    # linhas em cada uma das 4 tabelas por tipo
MAX_RESUMO_GERAL = 25       # linhas no resumo geral por forno
MAX_HELPER_ROWS = 200       # linhas das formulas auxiliares em _Dados


# =============================================================================
# ESTILOS
# =============================================================================

def thin_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border(color=COR_PRIMARIA):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=COR_HEADER_TXT)
    cell.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def style_input_label(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="595959")
    cell.fill = PatternFill("solid", fgColor="EAEAEA")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def style_input_field(cell):
    cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
    cell.fill = PatternFill("solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border()

def style_button(cell, color=COR_OK):
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border(color)

def style_data(cell, bold=False):
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()


# =============================================================================
# WORKBOOK
# =============================================================================

def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    ws_dash = wb.create_sheet("Dashboard")
    ws_reg = wb.create_sheet("Registro")
    ws_filt = wb.create_sheet("Filtros")
    ws_dados = wb.create_sheet("_Dados")

    build_dados(ws_dados)
    add_dados_helpers(ws_dados)
    build_registro(ws_reg)
    build_dashboard(ws_dash)
    build_filtros(ws_filt)

    ws_dados.sheet_state = "hidden"
    wb.active = wb.index(ws_dash)

    # CodeNames precisam casar com os modulos VBA (Sheet1..Sheet4 + ThisWorkbook).
    # Excel usa codeName para vincular modulo VBA <-> aba.
    wb.code_name = "ThisWorkbook"
    ws_dash.sheet_properties.codeName = "Sheet1"
    ws_reg.sheet_properties.codeName = "Sheet2"
    ws_filt.sheet_properties.codeName = "Sheet3"
    ws_dados.sheet_properties.codeName = "Sheet4"

    return wb


# =============================================================================
# ABA _Dados (oculta) - tabela mestre
# =============================================================================

def build_dados(ws):
    ws.sheet_view.showGridLines = False

    headers = ["ID", "Data", "Forno", "Tipo", "Lado"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)

    # Linhas de exemplo (para a tabela ter referencia inicial)
    exemplos = [
        (1,  datetime(2026, 1, 5),  101, "Operacao",   "A"),
        (2,  datetime(2026, 1, 5),  102, "Manutencao", "B"),
        (3,  datetime(2026, 1, 6),  101, "Falha",      "A"),
        (4,  datetime(2026, 2, 10), 103, "Inspecao",   "B"),
        (5,  datetime(2026, 2, 12), 102, "Operacao",   "A"),
        (6,  datetime(2026, 3, 3),  104, "Manutencao", "A"),
        (7,  datetime(2026, 3, 15), 101, "Operacao",   "B"),
        (8,  datetime(2026, 3, 15), 105, "Falha",      "A"),
        (9,  datetime(2026, 4, 2),  103, "Operacao",   "B"),
        (10, datetime(2026, 4, 18), 102, "Inspecao",   "A"),
        (11, datetime(2026, 5, 1),  104, "Operacao",   "A"),
        (12, datetime(2026, 5, 1),  105, "Manutencao", "B"),
    ]
    for row in exemplos:
        ws.append(row)

    last_row = 1 + len(exemplos)
    for r in range(2, last_row + 1):
        for c in range(1, 6):
            style_data(ws.cell(row=r, column=c))
        ws.cell(row=r, column=2).number_format = "dd/mm/yyyy"

    # Tabela Excel
    tab = Table(displayName="tblFornos", ref=f"A1:E{last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tab)

    for i, w in enumerate([8, 14, 10, 14, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =============================================================================
# Formulas auxiliares em _Dados
# =============================================================================

def add_dados_helpers(ws):
    # Coluna G: lista TIPOS
    ws.cell(row=1, column=7, value="Tipos").font = Font(bold=True)
    for i, t in enumerate(TIPOS, start=2):
        ws.cell(row=i, column=7, value=t)

    # Coluna H: lista LADOS
    ws.cell(row=1, column=8, value="Lados").font = Font(bold=True)
    for i, l in enumerate(LADOS, start=2):
        ws.cell(row=i, column=8, value=l)

    # Coluna J/K/L: registros por mes (para grafico)
    ws.cell(row=1, column=10, value="Mes").font = Font(bold=True)
    ws.cell(row=1, column=11, value="MesNum").font = Font(bold=True)
    ws.cell(row=1, column=12, value="Qtd").font = Font(bold=True)
    meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                   "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    for i, mn in enumerate(meses_nomes):
        r = 2 + i
        ws.cell(row=r, column=10, value=mn)
        ws.cell(row=r, column=11, value=i + 1)
        ws.cell(row=r, column=12, value=(
            f'=COUNTIFS(tblFornos[Data],">="&DATE(YEAR(TODAY()),{i+1},1),'
            f'tblFornos[Data],"<"&DATE(YEAR(TODAY()),{i+2},1))'
        ))

    # Coluna N/O: top 10 dias (data, qtd)
    ws.cell(row=1, column=14, value="Dia").font = Font(bold=True)
    ws.cell(row=1, column=15, value="Qtd").font = Font(bold=True)
    # Para os top 10 dias, contamos quantas vezes cada data aparece e pegamos
    # as 10 maiores. Como precisa de datas distintas, fazemos via formula:
    for i in range(10):
        r = 2 + i
        # Pegamos a i-esima data com maior contagem unica
        # Usando MATCH+LARGE+IF (formula leve, so 12 linhas iniciais de exemplo,
        # cresce com dados mas com cap em 200)
        # Como queremos datas DISTINTAS, usamos uma abordagem por counting
        ws.cell(row=r, column=14, value=(
            f'=IFERROR(INDEX(tblFornos[Data],'
            f'MATCH(LARGE(IF(COUNTIF($N$1:N{r-1},tblFornos[Data])=0,'
            f'COUNTIFS(tblFornos[Data],tblFornos[Data])),1),'
            f'IF(COUNTIF($N$1:N{r-1},tblFornos[Data])=0,'
            f'COUNTIFS(tblFornos[Data],tblFornos[Data])),0)),"")'
        ))
        ws.cell(row=r, column=14).number_format = "dd/mm"
        ws.cell(row=r, column=15, value=(
            f'=IF(N{r}="","",COUNTIFS(tblFornos[Data],N{r}))'
        ))

    # Coluna Q/R: por tipo (para grafico pizza)
    ws.cell(row=1, column=17, value="Tipo").font = Font(bold=True)
    ws.cell(row=1, column=18, value="Qtd").font = Font(bold=True)
    for i, t in enumerate(TIPOS):
        r = 2 + i
        ws.cell(row=r, column=17, value=t)
        ws.cell(row=r, column=18, value=f'=COUNTIFS(tblFornos[Tipo],"{t}")')

    # Coluna T/U/V/W: top fornos (para tabelas por tipo na aba Registro)
    # Para cada tipo, listar os fornos mais ativos
    for k, tipo in enumerate(TIPOS):
        col_forno = 20 + k * 2  # T=20, V=22, X=24, Z=26
        col_qtd = col_forno + 1
        ws.cell(row=1, column=col_forno, value=f"Forno_{tipo}").font = Font(bold=True)
        ws.cell(row=1, column=col_qtd, value=f"Qtd_{tipo}").font = Font(bold=True)
        for i in range(MAX_REGISTROS_FORNO):
            r = 2 + i
            col_letter = get_column_letter(col_forno)
            # Pegar i-esimo forno mais ativo do tipo (forno distinto)
            ws.cell(row=r, column=col_forno, value=(
                f'=IFERROR(INDEX(tblFornos[Forno],'
                f'MATCH(LARGE(IF((tblFornos[Tipo]="{tipo}")*'
                f'(COUNTIF(${col_letter}$1:{col_letter}{r-1},tblFornos[Forno])=0),'
                f'COUNTIFS(tblFornos[Forno],tblFornos[Forno],tblFornos[Tipo],"{tipo}")),1),'
                f'IF((tblFornos[Tipo]="{tipo}")*'
                f'(COUNTIF(${col_letter}$1:{col_letter}{r-1},tblFornos[Forno])=0),'
                f'COUNTIFS(tblFornos[Forno],tblFornos[Forno],tblFornos[Tipo],"{tipo}")),0)),"")'
            ))
            ws.cell(row=r, column=col_qtd, value=(
                f'=IF({col_letter}{r}="","",'
                f'COUNTIFS(tblFornos[Forno],{col_letter}{r},tblFornos[Tipo],"{tipo}"))'
            ))

    # Coluna AB-AG: resultado dos filtros (para aba Filtros)
    # AB=passa, AC=ID, AD=Data, AE=Forno, AF=Tipo, AG=Lado
    ws.cell(row=1, column=28, value="passa").font = Font(bold=True)  # AB
    ws.cell(row=1, column=29, value="fID").font = Font(bold=True)    # AC
    ws.cell(row=1, column=30, value="fData").font = Font(bold=True)  # AD
    ws.cell(row=1, column=31, value="fForno").font = Font(bold=True) # AE
    ws.cell(row=1, column=32, value="fTipo").font = Font(bold=True)  # AF
    ws.cell(row=1, column=33, value="fLado").font = Font(bold=True)  # AG

    for i in range(MAX_HELPER_ROWS):
        r = 2 + i
        # Coluna AB: 1 se a linha i da tabela passa nos filtros
        ws.cell(row=r, column=28, value=(
            f'=IF(ROW()-1>ROWS(tblFornos[ID]),0,'
            f'IF(AND('
            f'OR(Filtros!$B$5="",INDEX(tblFornos[Data],ROW()-1)>=Filtros!$B$5),'
            f'OR(Filtros!$C$5="",INDEX(tblFornos[Data],ROW()-1)<=Filtros!$C$5),'
            f'OR(Filtros!$D$5="",INDEX(tblFornos[Forno],ROW()-1)=Filtros!$D$5),'
            f'OR(Filtros!$E$5="",INDEX(tblFornos[Tipo],ROW()-1)=Filtros!$E$5),'
            f'OR(Filtros!$F$5="",INDEX(tblFornos[Lado],ROW()-1)=Filtros!$F$5),'
            f'OR(Filtros!$G$5="",DAY(INDEX(tblFornos[Data],ROW()-1))=Filtros!$G$5),'
            f'OR(Filtros!$H$5="",MONTH(INDEX(tblFornos[Data],ROW()-1))=Filtros!$H$5),'
            f'OR(Filtros!$I$5="",YEAR(INDEX(tblFornos[Data],ROW()-1))=Filtros!$I$5),'
            f'OR(Filtros!$C$7="",ISNUMBER(SEARCH(Filtros!$C$7,'
            f'INDEX(tblFornos[Forno],ROW()-1)&" "&INDEX(tblFornos[Tipo],ROW()-1)&" "&'
            f'INDEX(tblFornos[Lado],ROW()-1)&" "&TEXT(INDEX(tblFornos[Data],ROW()-1),"dd/mm/yyyy"))))'
            f'),1,0))'
        ))

        # AC..AG: o k-esimo registro que passou no filtro
        # Usa AGGREGATE para pegar os indices ordenados por linha original
        ws.cell(row=r, column=29, value=(
            f'=IFERROR(INDEX(tblFornos[ID],'
            f'AGGREGATE(15,6,(ROW($AB$2:$AB${1+MAX_HELPER_ROWS})-1)/'
            f'($AB$2:$AB${1+MAX_HELPER_ROWS}=1),ROWS($AC$2:AC{r}))),"")'
        ))
        ws.cell(row=r, column=30, value=(
            f'=IF($AC{r}="","",INDEX(tblFornos[Data],MATCH($AC{r},tblFornos[ID],0)))'
        ))
        ws.cell(row=r, column=30).number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=31, value=(
            f'=IF($AC{r}="","",INDEX(tblFornos[Forno],MATCH($AC{r},tblFornos[ID],0)))'
        ))
        ws.cell(row=r, column=32, value=(
            f'=IF($AC{r}="","",INDEX(tblFornos[Tipo],MATCH($AC{r},tblFornos[ID],0)))'
        ))
        ws.cell(row=r, column=33, value=(
            f'=IF($AC{r}="","",INDEX(tblFornos[Lado],MATCH($AC{r},tblFornos[ID],0)))'
        ))


# =============================================================================
# ABA REGISTRO
# =============================================================================

def build_registro(ws):
    ws.sheet_view.showGridLines = False

    larguras = {
        "A": 2,  "B": 13, "C": 13, "D": 13, "E": 13, "F": 3,
        "G": 13, "H": 11, "I": 3,
        "J": 13, "K": 11, "L": 3,
        "M": 13, "N": 11, "O": 3,
        "P": 13, "Q": 11,
    }
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    # Titulo
    ws.merge_cells("B2:Q2")
    c = ws["B2"]
    c.value = "REGISTRO RAPIDO DE FORNOS"
    c.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    # ===== AREA DE INPUT =====
    for ref, txt in [("B4", "DATA"), ("C4", "FORNO"), ("D4", "TIPO"), ("E4", "LADO")]:
        cell = ws[ref]
        cell.value = txt
        style_input_label(cell)

    ws["B5"] = datetime(2026, 5, 9).date()
    ws["B5"].number_format = "dd/mm/yyyy"
    ws["C5"] = ""
    ws["D5"] = TIPOS[0]
    ws["E5"] = LADOS[0]
    for ref in ["B5", "C5", "D5", "E5"]:
        style_input_field(ws[ref])
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[4].height = 18

    # Botao Registrar (com macro Registrar)
    ws.merge_cells("G4:H5")
    btn = ws["G4"]
    btn.value = "REGISTRAR"
    style_button(btn, COR_OK)

    # Botao Limpar (com macro LimparCampos)
    ws.merge_cells("J4:K5")
    btn2 = ws["J4"]
    btn2.value = "LIMPAR"
    style_button(btn2, "808080")

    # Instrucoes
    ws.merge_cells("M4:Q5")
    info = ws["M4"]
    info.value = ("Preencha os 4 campos e clique em REGISTRAR. "
                  "Sem macros: digite uma nova linha direto na tabela "
                  "(via Dados/Conexoes ou edite a aba _Dados).")
    info.font = Font(name="Calibri", size=9, italic=True, color="595959")
    info.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    info.fill = PatternFill("solid", fgColor="F2F2F2")
    info.border = thin_border()

    # ===== DATA VALIDATION =====
    dv_data = DataValidation(type="date", operator="between",
                             formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)",
                             allow_blank=False)
    dv_data.add("B5")
    ws.add_data_validation(dv_data)

    dv_forno = DataValidation(type="whole", operator="between",
                              formula1=1, formula2=99999, allow_blank=False)
    dv_forno.add("C5")
    ws.add_data_validation(dv_forno)

    dv_tipo = DataValidation(type="list", formula1='"Operacao,Manutencao,Falha,Inspecao"',
                             allow_blank=False)
    dv_tipo.add("D5")
    ws.add_data_validation(dv_tipo)

    dv_lado = DataValidation(type="list", formula1='"A,B"', allow_blank=False)
    dv_lado.add("E5")
    ws.add_data_validation(dv_lado)

    # ===== STATUS =====
    ws.merge_cells("B7:Q7")
    s = ws["B7"]
    s.value = ('=IF(COUNTA(tblFornos[ID])>0,'
               '"Ultimo registro: ID "&MAX(tblFornos[ID])&'
               '"  -  Total na base: "&COUNTA(tblFornos[ID])&" registros",'
               '"Sem registros ainda.")')
    s.font = Font(name="Calibri", size=11, bold=True, color=COR_PRIMARIA)
    s.fill = PatternFill("solid", fgColor="DEEBF7")
    s.alignment = Alignment(horizontal="center", vertical="center")
    s.border = thin_border(COR_PRIMARIA)
    ws.row_dimensions[7].height = 22

    # ===== TITULO TABELAS =====
    ws.merge_cells("B9:Q9")
    hdr = ws["B9"]
    hdr.value = "REGISTROS POR TIPO  -  Forno + Quantidade. Totais ao final."
    hdr.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor=COR_SECUNDARIA)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 24

    # ===== 4 TABELAS POR TIPO =====
    layout_tipos = [
        ("Operacao",   "B", "C", 20, 21),  # T,U
        ("Manutencao", "E", "F", 22, 23),  # V,W
        ("Falha",      "H", "I", 24, 25),  # X,Y
        ("Inspecao",   "K", "L", 26, 27),  # Z,AA
    ]

    cabecalho_row = 11
    inicio_dados = 12

    for tipo, col_forno, col_qtd, dados_forno_col, dados_qtd_col in layout_tipos:
        # Titulo do tipo
        ws.merge_cells(f"{col_forno}10:{col_qtd}10")
        title = ws[f"{col_forno}10"]
        title.value = tipo.upper()
        title.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor=CORES_TIPOS[tipo])
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[10].height = 22

        # Cabecalhos
        ws[f"{col_forno}{cabecalho_row}"] = "Forno"
        ws[f"{col_qtd}{cabecalho_row}"] = "Qtd"
        for ref in [f"{col_forno}{cabecalho_row}", f"{col_qtd}{cabecalho_row}"]:
            style_header(ws[ref])

        # Dados (linhas 12 a 12+MAX-1) buscando do _Dados
        dados_forno_letter = get_column_letter(dados_forno_col)
        dados_qtd_letter = get_column_letter(dados_qtd_col)
        for i in range(MAX_REGISTROS_FORNO):
            r = inicio_dados + i
            cf = ws[f"{col_forno}{r}"]
            cq = ws[f"{col_qtd}{r}"]
            cf.value = f'=IFERROR(_Dados!{dados_forno_letter}{2+i},"")'
            cq.value = f'=IFERROR(_Dados!{dados_qtd_letter}{2+i},"")'
            style_data(cf)
            style_data(cq, bold=True)

        # Total
        total_row = inicio_dados + MAX_REGISTROS_FORNO
        ws[f"{col_forno}{total_row}"] = "TOTAL"
        ws[f"{col_qtd}{total_row}"] = f'=COUNTIFS(tblFornos[Tipo],"{tipo}")'
        for ref in [f"{col_forno}{total_row}", f"{col_qtd}{total_row}"]:
            ws[ref].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            ws[ref].fill = PatternFill("solid", fgColor=CORES_TIPOS[tipo])
            ws[ref].alignment = Alignment(horizontal="center", vertical="center")
            ws[ref].border = medium_border(CORES_TIPOS[tipo])

    # ===== RESUMO POR FORNO (N:Q) =====
    ws.merge_cells("N10:Q10")
    rs = ws["N10"]
    rs.value = "RESUMO POR FORNO"
    rs.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    rs.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    rs.alignment = Alignment(horizontal="center", vertical="center")

    # Cabecalho - usaremos lista de fornos unicos com totais por tipo
    ws["N11"] = "Forno"
    ws["O11"] = "O / M / F / I"
    ws["P11"] = "Total"
    ws["Q11"] = "% Falha"
    for ref in ["N11", "O11", "P11", "Q11"]:
        style_header(ws[ref])

    # Lista de fornos unicos (top MAX_RESUMO_GERAL fornos mais ativos)
    # Pega o k-esimo forno distinto mais ativo
    for i in range(MAX_RESUMO_GERAL):
        r = 12 + i
        # Forno (k-esimo distinto)
        ws[f"N{r}"] = (
            f'=IFERROR(INDEX(tblFornos[Forno],'
            f'MATCH(LARGE(IF(COUNTIF($N$11:N{r-1},tblFornos[Forno])=0,'
            f'COUNTIFS(tblFornos[Forno],tblFornos[Forno])),1),'
            f'IF(COUNTIF($N$11:N{r-1},tblFornos[Forno])=0,'
            f'COUNTIFS(tblFornos[Forno],tblFornos[Forno])),0)),"")'
        )
        # Concatenacao de contagens por tipo
        ws[f"O{r}"] = (
            f'=IF(N{r}="","",'
            f'COUNTIFS(tblFornos[Forno],N{r},tblFornos[Tipo],"Operacao")&"/"&'
            f'COUNTIFS(tblFornos[Forno],N{r},tblFornos[Tipo],"Manutencao")&"/"&'
            f'COUNTIFS(tblFornos[Forno],N{r},tblFornos[Tipo],"Falha")&"/"&'
            f'COUNTIFS(tblFornos[Forno],N{r},tblFornos[Tipo],"Inspecao"))'
        )
        # Total
        ws[f"P{r}"] = f'=IF(N{r}="","",COUNTIFS(tblFornos[Forno],N{r}))'
        # % Falha
        ws[f"Q{r}"] = (
            f'=IF(OR(N{r}="",P{r}=0),"",'
            f'COUNTIFS(tblFornos[Forno],N{r},tblFornos[Tipo],"Falha")/P{r})'
        )
        ws[f"Q{r}"].number_format = "0.0%"
        for ref in [f"N{r}", f"O{r}", f"P{r}", f"Q{r}"]:
            style_data(ws[ref])
            ws[ref].font = Font(name="Calibri", size=10)

    # Total geral
    total_r = 12 + MAX_RESUMO_GERAL
    ws[f"N{total_r}"] = "TOTAL GERAL"
    ws[f"O{total_r}"] = (
        '=COUNTIFS(tblFornos[Tipo],"Operacao")&"/"&'
        'COUNTIFS(tblFornos[Tipo],"Manutencao")&"/"&'
        'COUNTIFS(tblFornos[Tipo],"Falha")&"/"&'
        'COUNTIFS(tblFornos[Tipo],"Inspecao")'
    )
    ws[f"P{total_r}"] = '=COUNTA(tblFornos[ID])'
    ws[f"Q{total_r}"] = ('=IF(P{0}=0,"",'
                        'COUNTIFS(tblFornos[Tipo],"Falha")/P{0})').format(total_r)
    ws[f"Q{total_r}"].number_format = "0.0%"
    for ref in [f"N{total_r}", f"O{total_r}", f"P{total_r}", f"Q{total_r}"]:
        ws[ref].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[ref].fill = PatternFill("solid", fgColor=COR_PRIMARIA)
        ws[ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[ref].border = medium_border(COR_PRIMARIA)

    # Conditional formatting: % Falha em vermelho se >0
    ws.conditional_formatting.add(
        f"Q12:Q{12+MAX_RESUMO_GERAL-1}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="C00000")
    )

    ws.freeze_panes = "B9"


# =============================================================================
# ABA DASHBOARD
# =============================================================================

def build_dashboard(ws):
    ws.sheet_view.showGridLines = False

    for col, w in [
        ("A", 2), ("B", 17), ("C", 17), ("D", 17), ("E", 17),
        ("F", 3), ("G", 17), ("H", 17), ("I", 17), ("J", 17),
        ("K", 3), ("L", 17), ("M", 17),
    ]:
        ws.column_dimensions[col].width = w

    # Header
    ws.merge_cells("B2:M2")
    c = ws["B2"]
    c.value = "DASHBOARD - CONTROLE DE FORNOS"
    c.font = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 42

    ws.merge_cells("B3:M3")
    s = ws["B3"]
    s.value = '=" Atualizado em "&TEXT(TODAY(),"dd/mm/yyyy")&"  -  Total: "&COUNTA(tblFornos[ID])&" registros"'
    s.font = Font(name="Calibri", size=10, italic=True, color="595959")
    s.alignment = Alignment(horizontal="center", vertical="center")
    s.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[3].height = 18

    # ===== KPI CARDS =====
    kpi_blocks = [
        ("B5:C5", "B6:C7", "TOTAL REGISTROS", "=COUNTA(tblFornos[ID])", COR_PRIMARIA),
        ("D5:E5", "D6:E7", "FORNOS DISTINTOS",
         '=SUMPRODUCT((tblFornos[Forno]<>"")/COUNTIF(tblFornos[Forno],tblFornos[Forno]&""))',
         COR_SECUNDARIA),
        ("G5:H5", "G6:H7", "FALHAS", '=COUNTIFS(tblFornos[Tipo],"Falha")', COR_DESTAQUE),
        ("I5:J5", "I6:J7", "OPERACOES", '=COUNTIFS(tblFornos[Tipo],"Operacao")', COR_OK),
        ("L5:M5", "L6:M7", "MES + ATIVO",
         '=IFERROR(INDEX(_Dados!J2:J13,MATCH(MAX(_Dados!L2:L13),_Dados!L2:L13,0)),"-")',
         COR_PRIMARIA),
    ]
    for label_rng, val_rng, label, formula, cor in kpi_blocks:
        ws.merge_cells(label_rng)
        ws.merge_cells(val_rng)
        lbl = label_rng.split(":")[0]
        val = val_rng.split(":")[0]
        ws[lbl] = label
        ws[lbl].font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        ws[lbl].fill = PatternFill("solid", fgColor=cor)
        ws[lbl].alignment = Alignment(horizontal="center", vertical="center")
        ws[lbl].border = medium_border(cor)

        ws[val] = formula
        ws[val].font = Font(name="Calibri", size=20, bold=True, color=cor)
        ws[val].fill = PatternFill("solid", fgColor="FFFFFF")
        ws[val].alignment = Alignment(horizontal="center", vertical="center")
        ws[val].border = medium_border(cor)

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 24
    ws.row_dimensions[7].height = 24

    # Titulos das areas de grafico
    ws.merge_cells("B9:F9"); t = ws["B9"]; t.value = "REGISTROS POR MES"
    t.font = Font(name="Calibri", size=12, bold=True, color=COR_PRIMARIA)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", fgColor="DEEBF7")
    t.border = thin_border()

    ws.merge_cells("G9:J9"); t = ws["G9"]; t.value = "TOP 10 DIAS COM MAIS REGISTROS"
    t.font = Font(name="Calibri", size=12, bold=True, color=COR_PRIMARIA)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", fgColor="DEEBF7")
    t.border = thin_border()

    ws.merge_cells("L9:M9"); t = ws["L9"]; t.value = "POR TIPO"
    t.font = Font(name="Calibri", size=12, bold=True, color=COR_PRIMARIA)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", fgColor="DEEBF7")
    t.border = thin_border()
    ws.row_dimensions[9].height = 22

    # ===== GRAFICO 1: Registros por Mes (col chart) =====
    ch1 = BarChart()
    ch1.type = "col"
    ch1.style = 11
    ch1.legend = None
    ch1.height = 9
    ch1.width = 16
    data_ref = Reference(ws, range_string="_Dados!$L$1:$L$13")
    cats_ref = Reference(ws, range_string="_Dados!$J$2:$J$13")
    ch1.add_data(data_ref, titles_from_data=True)
    ch1.set_categories(cats_ref)
    ch1.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(ch1, "B10")

    # ===== GRAFICO 2: Top 10 dias (bar chart horizontal) =====
    ch2 = BarChart()
    ch2.type = "bar"
    ch2.style = 12
    ch2.legend = None
    ch2.height = 9
    ch2.width = 13
    d2 = Reference(ws, range_string="_Dados!$O$1:$O$11")
    c2 = Reference(ws, range_string="_Dados!$N$2:$N$11")
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(c2)
    ch2.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(ch2, "G10")

    # ===== GRAFICO 3: Por Tipo (pie chart) =====
    ch3 = PieChart()
    ch3.height = 9
    ch3.width = 8
    d3 = Reference(ws, range_string="_Dados!$R$1:$R$5")
    c3 = Reference(ws, range_string="_Dados!$Q$2:$Q$5")
    ch3.add_data(d3, titles_from_data=True)
    ch3.set_categories(c3)
    ch3.dataLabels = DataLabelList(showPercent=True, showCatName=True)
    ws.add_chart(ch3, "L10")

    # ===== TABELA RESUMO POR MES x TIPO =====
    ws.merge_cells("B26:M26")
    h = ws["B26"]
    h.value = "RESUMO POR MES E TIPO"
    h.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=COR_SECUNDARIA)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[26].height = 22

    ws["B27"] = "Mes"
    ws.merge_cells("C27:D27"); ws["C27"] = "Operacao"
    ws.merge_cells("E27:F27"); ws["E27"] = "Manutencao"
    ws.merge_cells("G27:H27"); ws["G27"] = "Falha"
    ws.merge_cells("I27:J27"); ws["I27"] = "Inspecao"
    ws.merge_cells("L27:M27"); ws["L27"] = "TOTAL"
    for ref in ["B27", "C27", "E27", "G27", "I27", "L27"]:
        style_header(ws[ref])

    meses_nomes = ["Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    for i, mn in enumerate(meses_nomes):
        r = 28 + i
        m_num = i + 1
        ws[f"B{r}"] = mn
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"] = (f'=COUNTIFS(tblFornos[Tipo],"Operacao",'
                      f'tblFornos[Data],">="&DATE(YEAR(TODAY()),{m_num},1),'
                      f'tblFornos[Data],"<"&DATE(YEAR(TODAY()),{m_num+1},1))')
        ws.merge_cells(f"E{r}:F{r}")
        ws[f"E{r}"] = (f'=COUNTIFS(tblFornos[Tipo],"Manutencao",'
                      f'tblFornos[Data],">="&DATE(YEAR(TODAY()),{m_num},1),'
                      f'tblFornos[Data],"<"&DATE(YEAR(TODAY()),{m_num+1},1))')
        ws.merge_cells(f"G{r}:H{r}")
        ws[f"G{r}"] = (f'=COUNTIFS(tblFornos[Tipo],"Falha",'
                      f'tblFornos[Data],">="&DATE(YEAR(TODAY()),{m_num},1),'
                      f'tblFornos[Data],"<"&DATE(YEAR(TODAY()),{m_num+1},1))')
        ws.merge_cells(f"I{r}:J{r}")
        ws[f"I{r}"] = (f'=COUNTIFS(tblFornos[Tipo],"Inspecao",'
                      f'tblFornos[Data],">="&DATE(YEAR(TODAY()),{m_num},1),'
                      f'tblFornos[Data],"<"&DATE(YEAR(TODAY()),{m_num+1},1))')
        ws.merge_cells(f"L{r}:M{r}")
        ws[f"L{r}"] = f'=C{r}+E{r}+G{r}+I{r}'
        for ref in [f"B{r}", f"C{r}", f"E{r}", f"G{r}", f"I{r}", f"L{r}"]:
            style_data(ws[ref])
            if ref == f"B{r}":
                ws[ref].alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws[ref].font = Font(name="Calibri", size=10, bold=True, color="595959")

    # Total geral
    total_r = 40
    ws[f"B{total_r}"] = "TOTAL"
    ws.merge_cells(f"C{total_r}:D{total_r}"); ws[f"C{total_r}"] = '=COUNTIFS(tblFornos[Tipo],"Operacao")'
    ws.merge_cells(f"E{total_r}:F{total_r}"); ws[f"E{total_r}"] = '=COUNTIFS(tblFornos[Tipo],"Manutencao")'
    ws.merge_cells(f"G{total_r}:H{total_r}"); ws[f"G{total_r}"] = '=COUNTIFS(tblFornos[Tipo],"Falha")'
    ws.merge_cells(f"I{total_r}:J{total_r}"); ws[f"I{total_r}"] = '=COUNTIFS(tblFornos[Tipo],"Inspecao")'
    ws.merge_cells(f"L{total_r}:M{total_r}"); ws[f"L{total_r}"] = '=COUNTA(tblFornos[ID])'
    for ref in [f"B{total_r}", f"C{total_r}", f"E{total_r}", f"G{total_r}", f"I{total_r}", f"L{total_r}"]:
        ws[ref].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws[ref].fill = PatternFill("solid", fgColor=COR_PRIMARIA)
        ws[ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[ref].border = medium_border(COR_PRIMARIA)

    # Color scale na coluna Total
    ws.conditional_formatting.add(
        f"L28:M{total_r-1}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color=COR_PRIMARIA)
    )

    ws.freeze_panes = "B5"


# =============================================================================
# ABA FILTROS
# =============================================================================

def build_filtros(ws):
    ws.sheet_view.showGridLines = False

    larguras = {
        "A": 2, "B": 13, "C": 13, "D": 12, "E": 13, "F": 8,
        "G": 8, "H": 8, "I": 8, "J": 12,
    }
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    # Titulo
    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value = "FILTROS E ANALISE DETALHADA"
    c.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_PRIMARIA)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    # Labels dos filtros
    labels = [
        ("B4", "Data inicial"),
        ("C4", "Data final"),
        ("D4", "Forno"),
        ("E4", "Tipo"),
        ("F4", "Lado"),
        ("G4", "Dia"),
        ("H4", "Mes"),
        ("I4", "Ano"),
    ]
    for ref, txt in labels:
        cell = ws[ref]
        cell.value = txt
        style_input_label(cell)

    # Campos
    ws["B5"] = ""
    ws["C5"] = ""
    ws["B5"].number_format = "dd/mm/yyyy"
    ws["C5"].number_format = "dd/mm/yyyy"
    ws["D5"] = ""
    ws["E5"] = ""
    ws["F5"] = ""
    ws["G5"] = ""
    ws["H5"] = ""
    ws["I5"] = ""
    for ref in ["B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5"]:
        style_input_field(ws[ref])
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[4].height = 18

    # Validations
    dv_d1 = DataValidation(type="date", allow_blank=True,
                           formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)",
                           operator="between")
    dv_d1.add("B5"); dv_d1.add("C5")
    ws.add_data_validation(dv_d1)

    dv_t = DataValidation(type="list", formula1='"Operacao,Manutencao,Falha,Inspecao"', allow_blank=True)
    dv_t.add("E5")
    ws.add_data_validation(dv_t)

    dv_l = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    dv_l.add("F5")
    ws.add_data_validation(dv_l)

    dv_dia = DataValidation(type="whole", operator="between", formula1=1, formula2=31, allow_blank=True)
    dv_dia.add("G5")
    ws.add_data_validation(dv_dia)

    dv_mes = DataValidation(type="whole", operator="between", formula1=1, formula2=12, allow_blank=True)
    dv_mes.add("H5")
    ws.add_data_validation(dv_mes)

    dv_ano = DataValidation(type="whole", operator="between", formula1=2000, formula2=2100, allow_blank=True)
    dv_ano.add("I5")
    ws.add_data_validation(dv_ano)

    # Botao Limpar
    ws.merge_cells("J4:J5")
    btn = ws["J4"]
    btn.value = "LIMPAR"
    style_button(btn, "808080")

    # Busca livre
    ws["B7"] = "Busca livre"
    style_input_label(ws["B7"])
    ws.merge_cells("C7:I7")
    ws["C7"] = ""
    style_input_field(ws["C7"])
    ws["C7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["J7"] = "(forno, tipo, lado)"
    ws["J7"].font = Font(size=9, italic=True, color="808080")
    ws["J7"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    # Status
    ws.merge_cells("B9:J9")
    s = ws["B9"]
    s.value = '=" Resultados encontrados: "&SUMPRODUCT(--(_Dados!AC2:AC' + str(1 + MAX_HELPER_ROWS) + '<>""))&"  registros"'
    s.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    s.fill = PatternFill("solid", fgColor=COR_SECUNDARIA)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 24

    # Tabela de resultados
    ws["B11"] = "ID"; ws["C11"] = "Data"; ws["D11"] = "Forno"; ws["E11"] = "Tipo"
    ws["F11"] = "Lado"; ws["G11"] = "Dia"; ws["H11"] = "Mes"; ws["I11"] = "Ano"; ws["J11"] = "Semana"
    for ref in ["B11", "C11", "D11", "E11", "F11", "G11", "H11", "I11", "J11"]:
        style_header(ws[ref])

    for i in range(MAX_FILTRO_LINHAS):
        r = 12 + i
        # Sourcing das colunas AC-AG do _Dados
        ws[f"B{r}"] = f'=IFERROR(_Dados!AC{2+i},"")'
        ws[f"C{r}"] = f'=IFERROR(_Dados!AD{2+i},"")'
        ws[f"C{r}"].number_format = "dd/mm/yyyy"
        ws[f"D{r}"] = f'=IFERROR(_Dados!AE{2+i},"")'
        ws[f"E{r}"] = f'=IFERROR(_Dados!AF{2+i},"")'
        ws[f"F{r}"] = f'=IFERROR(_Dados!AG{2+i},"")'
        ws[f"G{r}"] = f'=IF(_Dados!AD{2+i}="","",DAY(_Dados!AD{2+i}))'
        ws[f"H{r}"] = f'=IF(_Dados!AD{2+i}="","",MONTH(_Dados!AD{2+i}))'
        ws[f"I{r}"] = f'=IF(_Dados!AD{2+i}="","",YEAR(_Dados!AD{2+i}))'
        ws[f"J{r}"] = f'=IF(_Dados!AD{2+i}="","",WEEKNUM(_Dados!AD{2+i}))'
        for ref in [f"B{r}", f"C{r}", f"D{r}", f"E{r}", f"F{r}", f"G{r}", f"H{r}", f"I{r}", f"J{r}"]:
            style_data(ws[ref])

    # Conditional formatting: linhas alternadas
    ws.conditional_formatting.add(
        f"B12:J{11+MAX_FILTRO_LINHAS}",
        FormulaRule(formula=[f'=AND(MOD(ROW(),2)=0,$B12<>"")'],
                    fill=PatternFill("solid", fgColor="F8F8F8"))
    )

    # Cor por tipo na coluna E
    for tipo, cor in CORES_TIPOS.items():
        ws.conditional_formatting.add(
            f"E12:E{11+MAX_FILTRO_LINHAS}",
            FormulaRule(formula=[f'$E12="{tipo}"'], font=Font(color=cor, bold=True))
        )

    ws.freeze_panes = "B12"


# =============================================================================
# VBA - Codigo do modulo
# =============================================================================

VBA_MODULE_CODE = '''Attribute VB_Name = "modFornos"
Option Explicit

' =====================================================================
' Modulo de macros - Controle de Fornos
' Macros: RegistrarForno, LimparCampos, LimparFiltros
' =====================================================================

Public Sub RegistrarForno()
    Dim ws As Worksheet, wsD As Worksheet
    Dim lo As ListObject
    Dim novaLinha As ListRow
    Dim dt As Variant, fn As Variant, tp As Variant, ld As Variant
    Dim novoID As Long

    Set ws = ThisWorkbook.Worksheets("Registro")
    Set wsD = ThisWorkbook.Worksheets("_Dados")
    Set lo = wsD.ListObjects("tblFornos")

    dt = ws.Range("B5").Value
    fn = ws.Range("C5").Value
    tp = ws.Range("D5").Value
    ld = ws.Range("E5").Value

    If Not IsDate(dt) Then
        MsgBox "Informe uma data valida.", vbExclamation, "Registro"
        ws.Range("B5").Select: Exit Sub
    End If
    If Not IsNumeric(fn) Or fn = "" Then
        MsgBox "Informe o numero do forno.", vbExclamation, "Registro"
        ws.Range("C5").Select: Exit Sub
    End If
    If Trim(CStr(tp)) = "" Then
        MsgBox "Selecione o tipo.", vbExclamation, "Registro"
        ws.Range("D5").Select: Exit Sub
    End If
    If Trim(CStr(ld)) = "" Then
        MsgBox "Selecione o lado.", vbExclamation, "Registro"
        ws.Range("E5").Select: Exit Sub
    End If

    Application.EnableEvents = False
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual

    If lo.ListRows.Count = 0 Then
        novoID = 1
    Else
        novoID = Application.WorksheetFunction.Max(lo.ListColumns("ID").DataBodyRange) + 1
    End If

    Set novaLinha = lo.ListRows.Add
    novaLinha.Range.Cells(1, 1).Value = novoID
    novaLinha.Range.Cells(1, 2).Value = CDate(dt)
    novaLinha.Range.Cells(1, 2).NumberFormat = "dd/mm/yyyy"
    novaLinha.Range.Cells(1, 3).Value = CLng(fn)
    novaLinha.Range.Cells(1, 4).Value = CStr(tp)
    novaLinha.Range.Cells(1, 5).Value = CStr(ld)

    ws.Range("C5").Value = ""
    ws.Range("C5").Select

    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    Application.EnableEvents = True
    Application.StatusBar = "Forno " & fn & " (" & tp & ", lado " & ld & ") registrado em " & Format(dt, "dd/mm/yyyy") & "."
End Sub

Public Sub LimparCampos()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Registro")
    Application.EnableEvents = False
    ws.Range("B5").Value = Date
    ws.Range("C5").Value = ""
    ws.Range("D5").Value = "Operacao"
    ws.Range("E5").Value = "A"
    ws.Range("C5").Select
    Application.EnableEvents = True
End Sub

Public Sub LimparFiltros()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("Filtros")
    Application.EnableEvents = False
    Application.ScreenUpdating = False
    ws.Range("B5").Value = ""
    ws.Range("C5").Value = ""
    ws.Range("D5").Value = ""
    ws.Range("E5").Value = ""
    ws.Range("F5").Value = ""
    ws.Range("G5").Value = ""
    ws.Range("H5").Value = ""
    ws.Range("I5").Value = ""
    ws.Range("C7").Value = ""
    Application.ScreenUpdating = True
    Application.EnableEvents = True
End Sub

Public Sub IrParaDashboard()
    ThisWorkbook.Worksheets("Dashboard").Activate
End Sub

Public Sub IrParaRegistro()
    ThisWorkbook.Worksheets("Registro").Activate
    ThisWorkbook.Worksheets("Registro").Range("C5").Select
End Sub

Public Sub IrParaFiltros()
    ThisWorkbook.Worksheets("Filtros").Activate
End Sub
'''


# Codigo VBA para o modulo Sheet2 (Registro): clicar em REGISTRAR/LIMPAR aciona macros
SHEET_REGISTRO_CODE = '''Attribute VB_Name = "Sheet2"
Attribute VB_Base = "0{00020820-0000-0000-C000-000000000046}"
Attribute VB_GlobalNameSpace = False
Attribute VB_Creatable = False
Attribute VB_PredeclaredId = True
Attribute VB_Exposed = True
Attribute VB_TemplateDerived = False
Attribute VB_Customizable = True
Option Explicit

Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    If Target.Cells.Count > 4 Then Exit Sub
    ' REGISTRAR (G4:H5)
    If Not Application.Intersect(Target, Me.Range("G4:H5")) Is Nothing Then
        Application.EnableEvents = False
        Me.Range("C5").Select
        Application.EnableEvents = True
        Call modFornos.RegistrarForno
        Exit Sub
    End If
    ' LIMPAR (J4:K5)
    If Not Application.Intersect(Target, Me.Range("J4:K5")) Is Nothing Then
        Application.EnableEvents = False
        Me.Range("C5").Select
        Application.EnableEvents = True
        Call modFornos.LimparCampos
        Exit Sub
    End If
End Sub
'''


# Codigo VBA para o modulo Sheet3 (Filtros): clicar em LIMPAR aciona macro
SHEET_FILTROS_CODE = '''Attribute VB_Name = "Sheet3"
Attribute VB_Base = "0{00020820-0000-0000-C000-000000000046}"
Attribute VB_GlobalNameSpace = False
Attribute VB_Creatable = False
Attribute VB_PredeclaredId = True
Attribute VB_Exposed = True
Attribute VB_TemplateDerived = False
Attribute VB_Customizable = True
Option Explicit

Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    If Target.Cells.Count > 4 Then Exit Sub
    If Not Application.Intersect(Target, Me.Range("J4:J5")) Is Nothing Then
        Application.EnableEvents = False
        Me.Range("B5").Select
        Application.EnableEvents = True
        Call modFornos.LimparFiltros
    End If
End Sub
'''


# =============================================================================
# OLE2 / CFB writer (para vbaProject.bin)
# =============================================================================

SECTOR_SIZE = 512
MINI_SECTOR_SIZE = 64
MINI_STREAM_CUTOFF = 4096
FREESECT = 0xFFFFFFFF
ENDOFCHAIN = 0xFFFFFFFE
FATSECT = 0xFFFFFFFD
NOSTREAM = 0xFFFFFFFF


def pad_to(data, size):
    if len(data) >= size:
        return data[:size]
    return data + b"\x00" * (size - len(data))


def make_unicode_name(name):
    encoded = name.encode("utf-16-le")
    return pad_to(encoded + b"\x00\x00", 64)


def dir_entry(name, type_, color=1, left=NOSTREAM, right=NOSTREAM, child=NOSTREAM,
              start_sector=ENDOFCHAIN, size=0, clsid=b"\x00" * 16):
    name_bytes = make_unicode_name(name)
    name_len = (len(name) + 1) * 2 if name else 0
    entry = b""
    entry += pad_to(name_bytes, 64)
    entry += struct.pack("<H", name_len)
    entry += struct.pack("<B", type_)
    entry += struct.pack("<B", color)
    entry += struct.pack("<I", left)
    entry += struct.pack("<I", right)
    entry += struct.pack("<I", child)
    entry += pad_to(clsid, 16)
    entry += struct.pack("<I", 0)
    entry += struct.pack("<Q", 0)
    entry += struct.pack("<Q", 0)
    entry += struct.pack("<I", start_sector)
    entry += struct.pack("<Q", size)
    return pad_to(entry, 128)


def vba_compress(data):
    """MS-OVBA RLE-style compression. Use uncompressed chunks (compression flag=0)."""
    out = bytearray()
    out.append(0x01)  # SignatureByte
    pos = 0
    while pos < len(data):
        chunk = data[pos:pos + 4096]
        chunk_len = len(chunk)
        # ChunkSize = chunk_len + 2 (header is 2 bytes); ChunkHeader value = ChunkSize - 3
        chunk_header_size = chunk_len + 2 - 3
        # bits 0-11: chunk_header_size
        # bits 12-14: signature 0b011
        # bit 15: compression flag = 0 (uncompressed)
        header = (chunk_header_size & 0x0FFF) | (0b011 << 12)
        out += struct.pack("<H", header)
        out += chunk
        pos += 4096
    return bytes(out)


def build_vba_dir_stream():
    """Build VBA/dir stream content (pre-compression)."""
    out = bytearray()

    # PROJECTSYSKIND: id=0x0001, size=4, kind=1 (Win32)
    out += struct.pack("<HII", 0x0001, 4, 1)
    # PROJECTLCID: id=0x0002, size=4
    out += struct.pack("<HII", 0x0002, 4, 0x0409)
    # PROJECTLCIDINVOKE: id=0x0014
    out += struct.pack("<HII", 0x0014, 4, 0x0409)
    # PROJECTCODEPAGE: id=0x0003, size=2, codepage=1252
    out += struct.pack("<HIH", 0x0003, 2, 1252)
    # PROJECTNAME: id=0x0004
    pname = b"VBAProject"
    out += struct.pack("<HI", 0x0004, len(pname)) + pname
    # PROJECTDOCSTRING: id=0x0005, size=0
    out += struct.pack("<HI", 0x0005, 0)
    out += struct.pack("<HI", 0x0040, 0)
    # PROJECTHELPFILEPATH: id=0x0006
    out += struct.pack("<HI", 0x0006, 0)
    out += struct.pack("<HI", 0x003D, 0)
    # PROJECTHELPCONTEXT: id=0x0007
    out += struct.pack("<HII", 0x0007, 4, 0)
    # PROJECTLIBFLAGS: id=0x0008
    out += struct.pack("<HII", 0x0008, 4, 0)
    # PROJECTVERSION: id=0x0009 - 4 bytes major + 2 bytes minor
    out += struct.pack("<HIIH", 0x0009, 6, 0x000F4DB6, 0x0000)
    # PROJECTCONSTANTS: id=0x000C
    out += struct.pack("<HI", 0x000C, 0)
    out += struct.pack("<HI", 0x003C, 0)

    # Reference: stdole
    refn = b"stdole"
    out += struct.pack("<HI", 0x0016, len(refn)) + refn
    out += struct.pack("<HI", 0x003E, len(refn) * 2) + refn.decode().encode("utf-16-le")
    libid = b"*\\G{00020430-0000-0000-C000-000000000046}#2.0#0#C:\\Windows\\System32\\stdole2.tlb#OLE Automation"
    inner = struct.pack("<I", len(libid)) + libid + struct.pack("<I", 0) + struct.pack("<H", 0)
    out += struct.pack("<HI", 0x000D, len(inner)) + inner

    # PROJECTMODULES: id=0x000F, size=2, count
    sheet_codenames = ["Sheet1", "Sheet2", "Sheet3", "Sheet4"]
    n_modules = 1 + 1 + len(sheet_codenames)  # modFornos + ThisWorkbook + sheets
    out += struct.pack("<HIH", 0x000F, 2, n_modules)
    # PROJECTCOOKIE: id=0x0013
    out += struct.pack("<HIH", 0x0013, 2, 0xFFFF)

    def module_block(name, kind=1):
        b = bytearray()
        nb = name.encode("ascii")
        nbu = name.encode("utf-16-le")
        # MODULENAME (id=0x0019)
        b += struct.pack("<HI", 0x0019, len(nb)) + nb
        # MODULENAMEUNICODE (id=0x0047)
        b += struct.pack("<HI", 0x0047, len(nbu)) + nbu
        # MODULESTREAMNAME (id=0x001A)
        b += struct.pack("<HI", 0x001A, len(nb)) + nb
        # MODULESTREAMNAMEUNICODE (id=0x0032)
        b += struct.pack("<HI", 0x0032, len(nbu)) + nbu
        # MODULEDOCSTRING (id=0x001C)
        b += struct.pack("<HI", 0x001C, 0)
        # MODULEDOCSTRINGUNICODE (id=0x0048)
        b += struct.pack("<HI", 0x0048, 0)
        # MODULEOFFSET (id=0x0031, size=4)
        b += struct.pack("<HII", 0x0031, 4, 0)
        # MODULEHELPCONTEXT (id=0x001E)
        b += struct.pack("<HII", 0x001E, 4, 0)
        # MODULECOOKIE (id=0x002C)
        b += struct.pack("<HIH", 0x002C, 2, 0xFFFF)
        # MODULETYPE: 0x0021 (procedural) or 0x0022 (document)
        if kind == 2:
            b += struct.pack("<HI", 0x0022, 0)
        else:
            b += struct.pack("<HI", 0x0021, 0)
        # Terminator (id=0x002B)
        b += struct.pack("<HI", 0x002B, 0)
        return bytes(b)

    out += module_block("modFornos", kind=1)
    out += module_block("ThisWorkbook", kind=2)
    for sn in sheet_codenames:
        out += module_block(sn, kind=2)

    # Final terminator (id=0x0010 + reserved 4 bytes)
    out += struct.pack("<HII", 0x0010, 0, 0)

    return bytes(out)


def build_module_source_stream(code):
    return vba_compress(code.encode("cp1252", errors="replace"))


def build_thisworkbook_stream():
    code = (
        'Attribute VB_Name = "ThisWorkbook"\r\n'
        'Attribute VB_Base = "0{00020819-0000-0000-C000-000000000046}"\r\n'
        'Attribute VB_GlobalNameSpace = False\r\n'
        'Attribute VB_Creatable = False\r\n'
        'Attribute VB_PredeclaredId = True\r\n'
        'Attribute VB_Exposed = True\r\n'
        'Attribute VB_TemplateDerived = False\r\n'
        'Attribute VB_Customizable = True\r\n'
        '\r\n'
    )
    return vba_compress(code.encode("cp1252"))


def build_sheet_stream(name):
    code = (
        f'Attribute VB_Name = "{name}"\r\n'
        f'Attribute VB_Base = "0{{00020820-0000-0000-C000-000000000046}}"\r\n'
        f'Attribute VB_GlobalNameSpace = False\r\n'
        f'Attribute VB_Creatable = False\r\n'
        f'Attribute VB_PredeclaredId = True\r\n'
        f'Attribute VB_Exposed = True\r\n'
        f'Attribute VB_TemplateDerived = False\r\n'
        f'Attribute VB_Customizable = True\r\n'
        f'\r\n'
    )
    return vba_compress(code.encode("cp1252"))


def build_vba_project_header():
    out = bytearray()
    out += struct.pack("<H", 0x61CC)
    out += struct.pack("<H", 0x00FF)
    out += b"\x00"
    out += struct.pack("<H", 0x0001)
    out += b"\x00" * 2042
    return bytes(out)


def build_project_stream():
    text = (
        "ID=\"{00000000-0000-0000-0000-000000000000}\"\r\n"
        "Document=ThisWorkbook/&H00000000\r\n"
        "Document=Sheet1/&H00000000\r\n"
        "Document=Sheet2/&H00000000\r\n"
        "Document=Sheet3/&H00000000\r\n"
        "Document=Sheet4/&H00000000\r\n"
        "Module=modFornos\r\n"
        "Name=\"VBAProject\"\r\n"
        "HelpContextID=0\r\n"
        "VersionCompatible32=\"393222000\"\r\n"
        "CMG=\"00000000000000000000\"\r\n"
        "DPB=\"00000000000000000000\"\r\n"
        "GC=\"00000000000000000000\"\r\n"
        "\r\n"
        "[Host Extender Info]\r\n"
        "&H00000001={3832D640-CF90-11CF-8E43-00A0C911005A};VBE;&H00000000\r\n"
        "\r\n"
    )
    return text.encode("cp1252")


def build_projectwm_stream():
    names = ["modFornos", "ThisWorkbook", "Sheet1", "Sheet2", "Sheet3", "Sheet4"]
    out = bytearray()
    for n in names:
        out += n.encode("ascii") + b"\x00"
        out += n.encode("utf-16-le") + b"\x00\x00"
    out += b"\x00"
    return bytes(out)


class CFBWriter:
    def __init__(self):
        self.streams = []

    def add_stream(self, path, data):
        self.streams.append((path, data))

    def build(self):
        root_streams = []
        vba_streams = []
        for path, data in self.streams:
            if path.startswith("/VBA/"):
                vba_streams.append((path[len("/VBA/"):], data))
            else:
                root_streams.append((path.lstrip("/"), data))

        sector_data = bytearray()
        mini_data = bytearray()
        main_chains = []
        mini_chains = []
        stream_alloc = {}

        def alloc_main(data):
            start = len(sector_data) // SECTOR_SIZE
            n = (len(data) + SECTOR_SIZE - 1) // SECTOR_SIZE
            sector_data.extend(data + b"\x00" * (n * SECTOR_SIZE - len(data)))
            return start, n

        def alloc_mini(data):
            start = len(mini_data) // MINI_SECTOR_SIZE
            n = (len(data) + MINI_SECTOR_SIZE - 1) // MINI_SECTOR_SIZE
            mini_data.extend(data + b"\x00" * (n * MINI_SECTOR_SIZE - len(data)))
            return start, n

        # Allocate root streams
        for name, data in root_streams:
            if len(data) < MINI_STREAM_CUTOFF:
                start, n = alloc_mini(data)
                stream_alloc[("root", name)] = (start, len(data), True)
                mini_chains.append((start, n))
            else:
                start, n = alloc_main(data)
                stream_alloc[("root", name)] = (start, len(data), False)
                main_chains.append((start, n))

        # Allocate VBA streams
        for name, data in vba_streams:
            if len(data) < MINI_STREAM_CUTOFF:
                start, n = alloc_mini(data)
                stream_alloc[("vba", name)] = (start, len(data), True)
                mini_chains.append((start, n))
            else:
                start, n = alloc_main(data)
                stream_alloc[("vba", name)] = (start, len(data), False)
                main_chains.append((start, n))

        # Allocate mini stream container
        if len(mini_data) > 0:
            mini_container_start, mini_container_n = alloc_main(bytes(mini_data))
            main_chains.append((mini_container_start, mini_container_n))
        else:
            mini_container_start = ENDOFCHAIN
            mini_container_n = 0

        # Build mini-FAT
        per_main = SECTOR_SIZE // 4
        mini_n_sectors = len(mini_data) // MINI_SECTOR_SIZE
        if mini_n_sectors > 0:
            mini_fat_array = [FREESECT] * mini_n_sectors
            for start, n in mini_chains:
                for k in range(n):
                    mini_fat_array[start + k] = (start + k + 1) if k < n - 1 else ENDOFCHAIN
            while len(mini_fat_array) % per_main != 0:
                mini_fat_array.append(FREESECT)
            mini_fat_data = b"".join(struct.pack("<I", v) for v in mini_fat_array)
            mini_fat_start, mini_fat_n_sectors = alloc_main(mini_fat_data)
            main_chains.append((mini_fat_start, mini_fat_n_sectors))
        else:
            mini_fat_start = ENDOFCHAIN
            mini_fat_n_sectors = 0

        # Build directory entries
        entries = []
        # Root
        entries.append({
            "name": "Root Entry", "type": 5, "color": 1,
            "left": NOSTREAM, "right": NOSTREAM, "child": NOSTREAM,
            "start_sector": mini_container_start,
            "size": len(mini_data), "clsid": b"\x00" * 16,
        })
        root_idxs = []
        for name, data in root_streams:
            start, sz, _ = stream_alloc[("root", name)]
            idx = len(entries)
            root_idxs.append(idx)
            entries.append({
                "name": name, "type": 2, "color": 1,
                "left": NOSTREAM, "right": NOSTREAM, "child": NOSTREAM,
                "start_sector": start, "size": sz, "clsid": b"\x00" * 16,
            })
        # VBA storage
        vba_idx = len(entries)
        root_idxs.append(vba_idx)
        entries.append({
            "name": "VBA", "type": 1, "color": 1,
            "left": NOSTREAM, "right": NOSTREAM, "child": NOSTREAM,
            "start_sector": ENDOFCHAIN, "size": 0, "clsid": b"\x00" * 16,
        })
        vba_idxs = []
        for name, data in vba_streams:
            start, sz, _ = stream_alloc[("vba", name)]
            idx = len(entries)
            vba_idxs.append(idx)
            entries.append({
                "name": name, "type": 2, "color": 1,
                "left": NOSTREAM, "right": NOSTREAM, "child": NOSTREAM,
                "start_sector": start, "size": sz, "clsid": b"\x00" * 16,
            })

        def ole_sort_key(name):
            return (len(name), name.upper())

        def build_tree(indexes):
            if not indexes:
                return NOSTREAM
            mid = len(indexes) // 2
            r = indexes[mid]
            entries[r]["left"] = build_tree(indexes[:mid])
            entries[r]["right"] = build_tree(indexes[mid+1:])
            return r

        root_idxs.sort(key=lambda i: ole_sort_key(entries[i]["name"]))
        vba_idxs.sort(key=lambda i: ole_sort_key(entries[i]["name"]))
        entries[0]["child"] = build_tree(root_idxs)
        entries[vba_idx]["child"] = build_tree(vba_idxs)

        dir_data = b""
        for e in entries:
            dir_data += dir_entry(
                name=e["name"], type_=e["type"], color=e["color"],
                left=e["left"], right=e["right"], child=e["child"],
                start_sector=e["start_sector"], size=e["size"],
                clsid=e["clsid"],
            )
        while len(dir_data) % SECTOR_SIZE != 0:
            dir_data += dir_entry("", 0, 0)
        dir_start, dir_n = alloc_main(dir_data)
        main_chains.append((dir_start, dir_n))

        # Compute FAT
        n_data_sectors = len(sector_data) // SECTOR_SIZE
        n_fat_sectors = 0
        while True:
            total = n_data_sectors + n_fat_sectors
            needed = (total + per_main - 1) // per_main
            if needed <= n_fat_sectors:
                break
            n_fat_sectors = needed

        fat_positions = list(range(n_data_sectors, n_data_sectors + n_fat_sectors))
        fat = [FREESECT] * (n_fat_sectors * per_main)
        for start, n in main_chains:
            for k in range(n):
                fat[start + k] = (start + k + 1) if k < n - 1 else ENDOFCHAIN
        for fp in fat_positions:
            fat[fp] = FATSECT

        # Header
        header = bytearray()
        header += b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
        header += b"\x00" * 16
        header += struct.pack("<H", 0x003E)
        header += struct.pack("<H", 0x0003)
        header += struct.pack("<H", 0xFFFE)
        header += struct.pack("<H", 9)
        header += struct.pack("<H", 6)
        header += b"\x00" * 6
        header += struct.pack("<I", 0)
        header += struct.pack("<I", n_fat_sectors)
        header += struct.pack("<I", dir_start)
        header += struct.pack("<I", 0)
        header += struct.pack("<I", MINI_STREAM_CUTOFF)
        header += struct.pack("<I", mini_fat_start if mini_fat_n_sectors > 0 else ENDOFCHAIN)
        header += struct.pack("<I", mini_fat_n_sectors)
        header += struct.pack("<I", ENDOFCHAIN)
        header += struct.pack("<I", 0)
        for i in range(109):
            if i < n_fat_sectors:
                header += struct.pack("<I", fat_positions[i])
            else:
                header += struct.pack("<I", FREESECT)
        header = pad_to(bytes(header), SECTOR_SIZE)

        all_data = bytearray(sector_data)
        fat_data = b"".join(struct.pack("<I", v) for v in fat)
        all_data.extend(fat_data)

        return bytes(header) + bytes(all_data)


def build_vba_project_bin(module_code):
    cfb = CFBWriter()
    cfb.add_stream("/PROJECT", build_project_stream())
    cfb.add_stream("/PROJECTwm", build_projectwm_stream())
    cfb.add_stream("/VBA/_VBA_PROJECT", build_vba_project_header())
    cfb.add_stream("/VBA/dir", vba_compress(build_vba_dir_stream()))
    cfb.add_stream("/VBA/modFornos", build_module_source_stream(module_code))
    cfb.add_stream("/VBA/ThisWorkbook", build_thisworkbook_stream())
    cfb.add_stream("/VBA/Sheet1", build_sheet_stream("Sheet1"))
    # Sheet2 (Registro) e Sheet3 (Filtros) tem handlers de evento
    cfb.add_stream("/VBA/Sheet2", vba_compress(SHEET_REGISTRO_CODE.encode("cp1252")))
    cfb.add_stream("/VBA/Sheet3", vba_compress(SHEET_FILTROS_CODE.encode("cp1252")))
    cfb.add_stream("/VBA/Sheet4", build_sheet_stream("Sheet4"))
    return cfb.build()


# =============================================================================
# Conversao xlsx -> xlsm
# =============================================================================

def convert_xlsx_to_xlsm(xlsx_path, xlsm_path, vba_bin):
    tmp_path = xlsm_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    txt = data.decode("utf-8")
                    if "vbaProject" not in txt:
                        ins = '<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/>'
                        txt = txt.replace(
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                            "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                        )
                        txt = txt.replace("</Types>", ins + "</Types>")
                    data = txt.encode("utf-8")
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    txt = data.decode("utf-8")
                    if "vbaProject" not in txt:
                        import re
                        ids = re.findall(r'Id="rId(\d+)"', txt)
                        max_id = max([int(x) for x in ids], default=0)
                        new_id = max_id + 1
                        rel = (f'<Relationship Id="rId{new_id}" '
                               f'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                               f'Target="vbaProject.bin"/>')
                        txt = txt.replace("</Relationships>", rel + "</Relationships>")
                    data = txt.encode("utf-8")
                zout.writestr(item, data)
            zout.writestr("xl/vbaProject.bin", vba_bin)
    shutil.move(tmp_path, xlsm_path)


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("==> Construindo workbook...")
    wb = build_workbook()

    print(f"==> Salvando {OUT_XLSX}...")
    wb.save(OUT_XLSX)

    print("==> Construindo vbaProject.bin...")
    vba_bin = build_vba_project_bin(VBA_MODULE_CODE)
    print(f"   vbaProject.bin tem {len(vba_bin):,} bytes")

    print(f"==> Convertendo para .xlsm em {OUT_XLSM}...")
    convert_xlsx_to_xlsm(OUT_XLSX, OUT_XLSM, vba_bin)

    print(f"==> Salvando codigo VBA em {OUT_BAS}...")
    with open(OUT_BAS, "w", encoding="cp1252") as f:
        f.write(VBA_MODULE_CODE)

    # Salvar tambem os codigos dos modulos das planilhas (em txt, para colar)
    sheet2_path = os.path.join(OUT_DIR, "Sheet2_Registro.txt")
    sheet3_path = os.path.join(OUT_DIR, "Sheet3_Filtros.txt")
    # Remover linhas Attribute antes de salvar (essas linhas existem so dentro do .bas)
    def strip_attrs(code):
        return "\n".join(l for l in code.split("\n") if not l.startswith("Attribute "))
    with open(sheet2_path, "w", encoding="cp1252") as f:
        f.write(strip_attrs(SHEET_REGISTRO_CODE))
    with open(sheet3_path, "w", encoding="cp1252") as f:
        f.write(strip_attrs(SHEET_FILTROS_CODE))

    print("==> Pronto!")
    print(f"   xlsx: {OUT_XLSX} ({os.path.getsize(OUT_XLSX):,} bytes)")
    print(f"   xlsm: {OUT_XLSM} ({os.path.getsize(OUT_XLSM):,} bytes)")
    print(f"   bas:  {OUT_BAS}  ({os.path.getsize(OUT_BAS):,} bytes)")
    print(f"   txt:  {sheet2_path}, {sheet3_path}")


if __name__ == "__main__":
    main()
