"""
Build the upgraded Curva ABC Reggae workbook (v6 - 10/10).

Improvements over v5:
  1. Live Pareto chart on each ABC sheet (sorted bars by V.T. + cumulative %
     on a secondary 0-100% axis), matching the reference image.
  2. Helper columns P:V driven by formulas so the chart updates automatically
     when Qtd. / Custo Unit. / limits are edited.
  3. % Acumulado rewritten with a tie-safe progressive sum (handles equal V.T.).
  4. Conditional formatting on data rows: green for A, yellow for B, red for C.
  5. Data validation on the A/B limits (B9, B10) - decimal between 0 and 1.
  6. Cell comments on the editable Qtd. and Custo Unit. columns.
  7. New 🏠 DASHBOARD sheet at the front with consolidated KPIs and a pie chart
     of V.T. by class across all three datasets.
"""

from copy import copy

from openpyxl import load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.comments import Comment
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice as DrawColor
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SRC = "/root/.claude/uploads/af5f62b9-abe8-4996-80a8-9ffafd174255/0cb5b4c3-curva_abc_reggae_5.xlsx"
DST = "output/curva_abc_reggae_6.xlsx"


# Reggae palette - vivid bands (per user choice) + cyan bars + red line
GREEN = "1FAE4D"
GREEN_LIGHT = "9FE3A5"   # vivid reggae green band (A region)
GREEN_VERY_LIGHT = "EAF7EB"
YELLOW = "F4C81E"
YELLOW_LIGHT = "FFE066"  # vivid reggae yellow band (B region)
YELLOW_VERY_LIGHT = "FFF6D3"
RED = "D4322B"
RED_LIGHT = "FFB07A"     # vivid orange/red band (C region)
RED_VERY_LIGHT = "FBD7D5"
BLUE = "3FB3E5"          # cyan bar fill
BLUE_DARK = "1A78A0"
LINE_RED = "C0392B"
DARK = "1B1B1B"
CREAM = "FFF8E7"


SHEETS = [
    # (sheet_name, first_data_row, last_data_row)
    ("🟢 Primeira", 12, 21),
    ("🟡 01",       12, 31),
    ("🔴 02",       12, 41),
]


def make_outer_shadow(color="000000", dist=38100, blur=50000):
    """Drawing-ML EffectList with a soft drop shadow."""
    from openpyxl.drawing.effect import EffectList, OuterShadow
    shadow = OuterShadow(
        blurRad=blur,
        dist=dist,
        dir=2700000,        # 270° = below
        algn="ctr",
        rotWithShape=False,
        srgbClr=color,
    )
    return EffectList(outerShdw=shadow)


def thin_border(color="888888"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_solid(color):
    gp = GraphicalProperties(solidFill=color)
    gp.line = LineProperties(solidFill="333333")
    return gp


# ---------------------------------------------------------------------------
# Build helper columns with formulas (sorted by V.T. desc, with tiebreaker)
# ---------------------------------------------------------------------------
def write_helper_columns(ws, first, last):
    """
    Helper layout (to the right of the data area). Drives the live Pareto chart.
      P  (16) Rank (1..N)
      Q  (17) SKU @ rank            - INDEX/MATCH against tie-broken score
      R  (18) V.T. @ rank           - same INDEX/MATCH on G
      S  (19) Classe @ rank         - A/B/C from cumulative %
      T  (20) % Acumulado @ rank    - progressive sum (sorted)
      U  (21) % Individual @ rank   - R / SUM(G)
      V  (22) Barra A (% indiv if S=A else NA)  - bar series, label "A"
      W  (23) Barra B (% indiv if S=B else NA)
      X  (24) Barra C (% indiv if S=C else NA)
      Y  (25) Banda A (1 if S=A else 0)         - area band fill
      Z  (26) Banda B
      AA (27) Banda C
      AB (28) score - hidden tie-breaker = G + (last-ROW)/1e9
    """
    n = last - first + 1
    g_range = f"$G${first}:$G${last}"

    # Headers row 11
    header_map = {
        16: "Rank", 17: "SKU", 18: "V.T. (R$)", 19: "Classe",
        20: "% Acumulado", 21: "% Individual",
        22: "A", 23: "B", 24: "C",
        25: "Banda A", 26: "Banda B", 27: "Banda C",
        28: "score",
    }
    for col, text in header_map.items():
        ws.cell(row=11, column=col, value=text)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=DARK)
    for col in range(16, 29):
        c = ws.cell(row=11, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # AB : tie-broken score = G + (last-ROW)/1e9
    for r in range(first, last + 1):
        ws.cell(row=r, column=28,
                value=f"=IFERROR($G{r}+({last}-ROW())/1000000000,0)")

    # P : rank 1..N
    for i in range(n):
        ws.cell(row=first + i, column=16, value=i + 1)

    score_range = f"$AB${first}:$AB${last}"
    sku_range   = f"$B${first}:$B${last}"
    total_g     = f"SUM({g_range})"

    for i in range(n):
        r = first + i
        rank_cell = f"$P{r}"

        # Q : SKU at this rank
        ws.cell(row=r, column=17,
                value=f"=IFERROR(INDEX({sku_range},MATCH(LARGE({score_range},{rank_cell}),{score_range},0)),\"\")")

        # R : V.T. at this rank
        ws.cell(row=r, column=18,
                value=f"=IFERROR(INDEX({g_range},MATCH(LARGE({score_range},{rank_cell}),{score_range},0)),0)")

        # T : cumulative % (progressive over sorted R)
        if i == 0:
            ws.cell(row=r, column=20, value=f"=IFERROR($R{r}/{total_g},0)")
        else:
            ws.cell(row=r, column=20, value=f"=IFERROR($T{r-1}+$R{r}/{total_g},0)")

        # S : class based on cumulative %
        ws.cell(row=r, column=19,
                value=f"=IF($T{r}<=$B$9,\"A\",IF($T{r}<=$B$10,\"B\",\"C\"))")

        # U : % individual at this rank
        ws.cell(row=r, column=21, value=f"=IFERROR($R{r}/{total_g},0)")

        # V/W/X : per-class bar values (% individual) — NA() so unmatched
        # classes leave a gap and their data label is skipped.
        ws.cell(row=r, column=22, value=f"=IF($S{r}=\"A\",$U{r},NA())")
        ws.cell(row=r, column=23, value=f"=IF($S{r}=\"B\",$U{r},NA())")
        ws.cell(row=r, column=24, value=f"=IF($S{r}=\"C\",$U{r},NA())")

        # Y/Z/AA : background bands - shape follows the cumulative curve
        # (technique from the YouTube tutorial). Each is the cumulative value
        # masked by class, with NA() outside so the area chart leaves no fill.
        #   Y  (Banda A) only for class-A rows
        #   Z  (Banda B) for class-A AND class-B rows (so the yellow area
        #                continues from where green ends)
        #   AA (Banda C) for every row (orange covers everything)
        # Drawn order in the chart is C -> B -> A so smaller layers stay on top.
        ws.cell(row=r, column=25, value=f"=IF($S{r}=\"A\",$T{r},NA())")
        ws.cell(row=r, column=26, value=f"=IF(OR($S{r}=\"A\",$S{r}=\"B\"),$T{r},NA())")
        ws.cell(row=r, column=27, value=f"=$T{r}")

        # Number formats
        ws.cell(row=r, column=18).number_format = '#,##0'
        ws.cell(row=r, column=20).number_format = '0.0%'
        ws.cell(row=r, column=21).number_format = '0.0%'
        ws.cell(row=r, column=22).number_format = '0.0%'
        ws.cell(row=r, column=23).number_format = '0.0%'
        ws.cell(row=r, column=24).number_format = '0.0%'
        ws.cell(row=r, column=25).number_format = '0.0%'
        ws.cell(row=r, column=26).number_format = '0.0%'
        ws.cell(row=r, column=27).number_format = '0.0%'

    # Hide score column AB
    ws.column_dimensions['AB'].hidden = True
    # Narrower helper area
    for col in 'PQRSTUVWXYZ':
        ws.column_dimensions[col].width = 11
    ws.column_dimensions['AA'].width = 11


# ---------------------------------------------------------------------------
# Fix the % Acumulado formula in column I (tie-safe progressive sum)
# ---------------------------------------------------------------------------
def fix_cumulative_formula(ws, first, last):
    """
    Replace the SUMIF(">="&G) approach (which gives equal values for ties)
    with a position-aware progressive sum:
      cumul% = (sum of G where G > G_this) + (sum of G where G = G_this AND
                row <= row_this) / total
    """
    g_range = f"$G${first}:$G${last}"
    row_range = f"ROW({g_range})"
    for r in range(first, last + 1):
        formula = (
            f"=IFERROR("
            f"(SUMPRODUCT(({g_range}>$G{r})*{g_range})"
            f"+SUMPRODUCT(({g_range}=$G{r})*({row_range}<=ROW($G{r}))*{g_range}))"
            f"/SUM({g_range}),0)"
        )
        ws.cell(row=r, column=9, value=formula)
        ws.cell(row=r, column=9).number_format = '0.0%'


# ---------------------------------------------------------------------------
# Conditional formatting: paint data rows green/yellow/red by class.
# ---------------------------------------------------------------------------
def add_conditional_formatting(ws, first, last):
    rng = f"A{first}:I{last}"

    fill_a = PatternFill("solid", fgColor=GREEN_LIGHT)
    fill_b = PatternFill("solid", fgColor=YELLOW_LIGHT)
    fill_c = PatternFill("solid", fgColor=RED_LIGHT)

    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'$A{first}="A"'], stopIfTrue=False, fill=fill_a))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'$A{first}="B"'], stopIfTrue=False, fill=fill_b))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'$A{first}="C"'], stopIfTrue=False, fill=fill_c))


# ---------------------------------------------------------------------------
# Data validation on the A/B limits (B9, B10)
# ---------------------------------------------------------------------------
def add_data_validation(ws):
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=0,
        formula2=1,
        allow_blank=False,
        showErrorMessage=True,
        showInputMessage=True,
    )
    dv.error = "Digite um percentual entre 0 e 1 (ex.: 0,8 para 80%)."
    dv.errorTitle = "Limite invalido"
    dv.prompt = "Percentual entre 0 e 1 (0,8 = 80%)."
    dv.promptTitle = "Limite ABC"
    dv.add("B9:B10")
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Cell comments on editable columns (E, F)
# ---------------------------------------------------------------------------
def add_editable_comments(ws, first, last):
    note_qty = Comment(
        "Edite apenas as colunas E (Qtd.) e F (Custo Unit.).\n"
        "V.T., % e Classe ABC sao recalculados automaticamente.",
        "Curva ABC",
    )
    note_qty.width = 220
    note_qty.height = 80
    # Drop a single, prominent comment on the first editable row (E12, F12).
    ws.cell(row=first, column=5).comment = Comment(note_qty.text, note_qty.author)
    ws.cell(row=first, column=6).comment = Comment(note_qty.text, note_qty.author)


# ---------------------------------------------------------------------------
# Build the Pareto chart on each ABC sheet
# ---------------------------------------------------------------------------
def add_pareto_chart(ws, sheet_name, first, last):
    """
    Composite Pareto chart (matches the YouTube reference look).

      Layer 1 (back)  : AreaChart with 3 series cut by class on the
                        cumulative curve - paints reggae-coloured zones
                        behind the bars, with the top edge following T.
                          AA (col 27) - Banda C - laranja - draws first (back)
                          Z  (col 26) - Banda B - amarelo
                          Y  (col 25) - Banda A - verde   - draws last (front)
      Layer 2         : BarChart, stacked (only one series non-NA per bar)
                        all painted the same cyan blue. Series names are
                        "A"/"B"/"C" so showSerName puts the class letter
                        above the bar.
                          V (22) Bar A  W (23) Bar B  X (24) Bar C
      Layer 3 (front) : LineChart - thick red smooth cumulative line.
                          T (20)
      Y-axis          : single 0-100% scale, no decimals. The secondary
                        axes openpyxl creates when combining are kept at
                        the same 0..1 scale but their labels are hidden.
      Categories      : Q (col 17) - sorted SKUs.
    """
    # --- Background bands ---------------------------------------------------
    area = AreaChart()
    area.title = "CURVA ABC"
    area.height = 12
    area.width = 28
    area.legend = None
    area.y_axis.scaling.min = 0
    area.y_axis.scaling.max = 1
    area.y_axis.number_format = "0%"
    area.y_axis.majorUnit = 0.1
    area.y_axis.majorTickMark = "out"
    area.x_axis.title = None
    area.y_axis.title = None

    cats_ref = Reference(ws, min_col=17, min_row=first, max_row=last)

    # Order matters: first series drawn BEHIND, last drawn IN FRONT.
    # We want green (A, smallest) on top -> add C then B then A.
    band_specs = [
        (27, RED_LIGHT),     # AA - Banda C (orange) BACK
        (26, YELLOW_LIGHT),  # Z  - Banda B (yellow) MID
        (25, GREEN_LIGHT),   # Y  - Banda A (green)  FRONT
    ]
    for col_idx, _ in band_specs:
        ref = Reference(ws, min_col=col_idx, min_row=11, max_row=last)
        area.add_data(ref, titles_from_data=True)
    area.set_categories(cats_ref)

    for s, (_, color) in zip(area.series, band_specs):
        s.graphicalProperties = GraphicalProperties(solidFill=color)
        s.graphicalProperties.line = LineProperties(noFill=True)

    # --- Blue bars (% Individual, labelled A/B/C) ---------------------------
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.gapWidth = 60  # thicker bars per the video
    for col_idx in (22, 23, 24):  # V W X
        ref = Reference(ws, min_col=col_idx, min_row=11, max_row=last)
        bar.add_data(ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    for s in bar.series:
        s.graphicalProperties = GraphicalProperties(solidFill=BLUE)
        s.graphicalProperties.line = LineProperties(solidFill=BLUE_DARK, w=6000)
        s.graphicalProperties.effectLst = make_outer_shadow(dist=25000, blur=40000)
        s.dLbls = DataLabelList(showSerName=True)
        s.dLbls.position = "t"

    # --- Red cumulative line ------------------------------------------------
    line = LineChart()
    cum_ref = Reference(ws, min_col=20, min_row=11, max_row=last)
    line.add_data(cum_ref, titles_from_data=True)
    s = line.series[0]
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=LINE_RED, w=44450)  # 3.5pt
    s.graphicalProperties.effectLst = make_outer_shadow(dist=20000, blur=30000)
    s.smooth = True

    # --- Combine ------------------------------------------------------------
    area += bar
    area += line

    # The += pushes bar/line onto a secondary axis. Keep the same 0..1 scale
    # so visually they line up with the area's primary axis, and hide the
    # secondary axis labels so the user only sees ONE % scale on the left.
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 1
    bar.y_axis.delete = True
    line.y_axis.scaling.min = 0
    line.y_axis.scaling.max = 1
    line.y_axis.delete = True

    # Anchor below the data table.
    anchor = f"A{last + 3}"
    ws.add_chart(area, anchor)


# ---------------------------------------------------------------------------
# Clean up old static helper labels P12:P{last} (legacy text categories).
# ---------------------------------------------------------------------------
def clear_legacy_helper(ws, first, last):
    for r in range(first, last + 1):
        for col in range(16, 30):
            ws.cell(row=r, column=col).value = None


# ---------------------------------------------------------------------------
# Remove old embedded chart from sheet by stripping its drawing reference.
# openpyxl re-emits charts only for ones it created or loaded, so we want a
# clean slate. We do this by clearing ws._charts and ws._drawing.
# ---------------------------------------------------------------------------
def reset_sheet_charts(ws):
    ws._charts = []
    # Comments cleared too -- we re-add the editable comments later.
    ws._comments = []


# ---------------------------------------------------------------------------
# Dashboard sheet: consolidated KPIs and pie chart of V.T. by class.
# ---------------------------------------------------------------------------
def build_dashboard(wb):
    if "🏠 DASHBOARD" in wb.sheetnames:
        del wb["🏠 DASHBOARD"]
    ws = wb.create_sheet("🏠 DASHBOARD", 0)
    ws.sheet_properties.tabColor = "1FAE4D"

    # Column widths
    widths = {"A": 2, "B": 28, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18,
              "H": 18, "I": 18, "J": 18, "K": 18, "L": 2}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    # ----- Title banner -----
    ws.merge_cells("B2:K3")
    title = ws.cell(row=2, column=2,
                    value="🏠 DASHBOARD ABC — One Love, One Excel 🌿")
    title.font = Font(name="Segoe UI", size=22, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, 4):
        for c in range(2, 12):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN)

    ws.merge_cells("B4:K4")
    sub = ws.cell(row=4, column=2,
                  value="Resumo consolidado das 3 curvas ABC  •  KPIs ao vivo  •  Visao por classe")
    sub.font = Font(name="Segoe UI", size=11, italic=True, color="333333")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    sub.fill = PatternFill("solid", fgColor=CREAM)

    # ----- KPI cards -----
    # Row 6 headers (per-sheet) and row 7 values.
    kpi_label_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    kpi_val_font = Font(name="Segoe UI", size=14, bold=True, color="1B1B1B")

    # KPI definitions per dataset
    sheets_info = [
        ("🟢 Primeira", "Primeira",  12, 21, GREEN),
        ("🟡 01",       "Aba 01",    12, 31, YELLOW),
        ("🔴 02",       "Aba 02",    12, 41, RED),
    ]

    # Per-sheet KPI rows
    start_row = 6
    for i, (sheet, alias, first, last, color) in enumerate(sheets_info):
        row = start_row + i * 3
        # Header strip
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        h = ws.cell(row=row, column=2, value=f"📊  {alias}  —  intervalo {first}:{last}")
        h.font = kpi_label_font
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(2, 12):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=color)

        # Card labels (row+1) and values (row+2)
        cards = [
            ("Total Itens",     f"=COUNTA('{sheet}'!D{first}:D{last})",                 "0"),
            ("V.T. Total",      f"=SUM('{sheet}'!G{first}:G{last})",                    "R$ #,##0"),
            ("V.T. Classe A",   f"=SUMIF('{sheet}'!A{first}:A{last},\"A\",'{sheet}'!G{first}:G{last})", "R$ #,##0"),
            ("V.T. Classe B",   f"=SUMIF('{sheet}'!A{first}:A{last},\"B\",'{sheet}'!G{first}:G{last})", "R$ #,##0"),
            ("V.T. Classe C",   f"=SUMIF('{sheet}'!A{first}:A{last},\"C\",'{sheet}'!G{first}:G{last})", "R$ #,##0"),
            ("Itens em A",      f"=COUNTIF('{sheet}'!A{first}:A{last},\"A\")",          "0"),
            ("% Itens A",       f"=IFERROR(COUNTIF('{sheet}'!A{first}:A{last},\"A\")/COUNTA('{sheet}'!D{first}:D{last}),0)", "0.0%"),
            ("Maior V.T.",      f"=MAX('{sheet}'!G{first}:G{last})",                    "R$ #,##0"),
        ]
        # Layout: cards across columns B..I (8 cards per row), 2 columns per card pair.
        # We'll place each card in its own column (B..I).
        for idx, (lbl, formula, fmt) in enumerate(cards):
            col = 2 + idx
            lbl_cell = ws.cell(row=row + 1, column=col, value=lbl)
            lbl_cell.font = Font(name="Segoe UI", size=9, bold=True, color="555555")
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.fill = PatternFill("solid", fgColor="F4F4F4")
            lbl_cell.border = thin_border("CCCCCC")

            val_cell = ws.cell(row=row + 2, column=col, value=formula)
            val_cell.font = kpi_val_font
            val_cell.number_format = fmt
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            val_cell.border = thin_border("CCCCCC")

        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row + 1].height = 16
        ws.row_dimensions[row + 2].height = 26

    # ----- Grand totals -----
    gt_row = start_row + len(sheets_info) * 3 + 1
    ws.merge_cells(start_row=gt_row, start_column=2, end_row=gt_row, end_column=11)
    h = ws.cell(row=gt_row, column=2, value="🌍 TOTAIS CONSOLIDADOS (3 abas somadas)")
    h.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 12):
        ws.cell(row=gt_row, column=c).fill = PatternFill("solid", fgColor=DARK)

    # Grand-total cards
    refs = [
        ("Total Itens",
            "=COUNTA('🟢 Primeira'!D12:D21)+COUNTA('🟡 01'!D12:D31)+COUNTA('🔴 02'!D12:D41)",
            "0"),
        ("V.T. Total",
            "=SUM('🟢 Primeira'!G12:G21)+SUM('🟡 01'!G12:G31)+SUM('🔴 02'!G12:G41)",
            "R$ #,##0"),
        ("V.T. Classe A",
            "=SUMIF('🟢 Primeira'!A12:A21,\"A\",'🟢 Primeira'!G12:G21)"
            "+SUMIF('🟡 01'!A12:A31,\"A\",'🟡 01'!G12:G31)"
            "+SUMIF('🔴 02'!A12:A41,\"A\",'🔴 02'!G12:G41)",
            "R$ #,##0"),
        ("V.T. Classe B",
            "=SUMIF('🟢 Primeira'!A12:A21,\"B\",'🟢 Primeira'!G12:G21)"
            "+SUMIF('🟡 01'!A12:A31,\"B\",'🟡 01'!G12:G31)"
            "+SUMIF('🔴 02'!A12:A41,\"B\",'🔴 02'!G12:G41)",
            "R$ #,##0"),
        ("V.T. Classe C",
            "=SUMIF('🟢 Primeira'!A12:A21,\"C\",'🟢 Primeira'!G12:G21)"
            "+SUMIF('🟡 01'!A12:A31,\"C\",'🟡 01'!G12:G31)"
            "+SUMIF('🔴 02'!A12:A41,\"C\",'🔴 02'!G12:G41)",
            "R$ #,##0"),
        ("Itens em A",
            "=COUNTIF('🟢 Primeira'!A12:A21,\"A\")"
            "+COUNTIF('🟡 01'!A12:A31,\"A\")"
            "+COUNTIF('🔴 02'!A12:A41,\"A\")",
            "0"),
        ("Itens em B",
            "=COUNTIF('🟢 Primeira'!A12:A21,\"B\")"
            "+COUNTIF('🟡 01'!A12:A31,\"B\")"
            "+COUNTIF('🔴 02'!A12:A41,\"B\")",
            "0"),
        ("Itens em C",
            "=COUNTIF('🟢 Primeira'!A12:A21,\"C\")"
            "+COUNTIF('🟡 01'!A12:A31,\"C\")"
            "+COUNTIF('🔴 02'!A12:A41,\"C\")",
            "0"),
    ]
    for idx, (lbl, formula, fmt) in enumerate(refs):
        col = 2 + idx
        lbl_cell = ws.cell(row=gt_row + 1, column=col, value=lbl)
        lbl_cell.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.fill = PatternFill("solid", fgColor="333333")
        lbl_cell.border = thin_border("777777")

        val_cell = ws.cell(row=gt_row + 2, column=col, value=formula)
        val_cell.font = Font(name="Segoe UI", size=14, bold=True, color=GREEN)
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        val_cell.border = thin_border("777777")

    ws.row_dimensions[gt_row].height = 22
    ws.row_dimensions[gt_row + 1].height = 16
    ws.row_dimensions[gt_row + 2].height = 28

    # ----- Pie chart of V.T. by class (consolidated) -----
    pie_row = gt_row + 5
    ws.cell(row=pie_row, column=2, value="Distribuicao do V.T. consolidado por classe ABC").font = \
        Font(name="Segoe UI", size=12, bold=True, color=DARK)

    # Helper area for pie data (small, kept on same sheet to the right)
    ws.cell(row=pie_row + 1, column=11, value="Classe")
    ws.cell(row=pie_row + 1, column=12, value="V.T.")
    ws.cell(row=pie_row + 2, column=11, value="A")
    ws.cell(row=pie_row + 2, column=12,
            value="=SUMIF('🟢 Primeira'!A12:A21,\"A\",'🟢 Primeira'!G12:G21)"
                  "+SUMIF('🟡 01'!A12:A31,\"A\",'🟡 01'!G12:G31)"
                  "+SUMIF('🔴 02'!A12:A41,\"A\",'🔴 02'!G12:G41)")
    ws.cell(row=pie_row + 3, column=11, value="B")
    ws.cell(row=pie_row + 3, column=12,
            value="=SUMIF('🟢 Primeira'!A12:A21,\"B\",'🟢 Primeira'!G12:G21)"
                  "+SUMIF('🟡 01'!A12:A31,\"B\",'🟡 01'!G12:G31)"
                  "+SUMIF('🔴 02'!A12:A41,\"B\",'🔴 02'!G12:G41)")
    ws.cell(row=pie_row + 4, column=11, value="C")
    ws.cell(row=pie_row + 4, column=12,
            value="=SUMIF('🟢 Primeira'!A12:A21,\"C\",'🟢 Primeira'!G12:G21)"
                  "+SUMIF('🟡 01'!A12:A31,\"C\",'🟡 01'!G12:G31)"
                  "+SUMIF('🔴 02'!A12:A41,\"C\",'🔴 02'!G12:G41)")
    for r in range(pie_row + 1, pie_row + 5):
        ws.cell(row=r, column=12).number_format = "R$ #,##0"
    # Hide helper columns K:L by narrowing... actually keep them visible because
    # the pie sits next to them; just style them subtly.

    pie = PieChart()
    pie.title = "V.T. por Classe (consolidado)"
    labels = Reference(ws, min_col=11, min_row=pie_row + 2, max_row=pie_row + 4)
    data = Reference(ws, min_col=12, min_row=pie_row + 1, max_row=pie_row + 4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 9
    pie.width = 14
    pie.dataLabels = DataLabelList(showPercent=True)

    # Slice colours: green / yellow / red
    slice_colors = [GREEN, YELLOW, RED]
    pts = []
    for i, c in enumerate(slice_colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties = GraphicalProperties(solidFill=c)
        pts.append(dp)
    pie.series[0].data_points = pts

    ws.add_chart(pie, f"B{pie_row + 1}")

    # ----- Footer note -----
    foot_row = pie_row + 22
    ws.merge_cells(start_row=foot_row, start_column=2, end_row=foot_row, end_column=11)
    f = ws.cell(row=foot_row, column=2,
                value="🌿  One Love, One Excel  •  Dashboard atualizado automaticamente com base nas 3 abas ABC.")
    f.font = Font(name="Segoe UI", size=10, italic=True, color="FFFFFF")
    f.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 12):
        ws.cell(row=foot_row, column=c).fill = PatternFill("solid", fgColor=GREEN)

    # Tab order: dashboard first
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Run.
# ---------------------------------------------------------------------------
def main():
    wb = load_workbook(SRC)

    for sheet_name, first, last in SHEETS:
        ws = wb[sheet_name]
        reset_sheet_charts(ws)
        clear_legacy_helper(ws, first, last)
        fix_cumulative_formula(ws, first, last)
        write_helper_columns(ws, first, last)
        add_conditional_formatting(ws, first, last)
        add_data_validation(ws)
        add_editable_comments(ws, first, last)
        add_pareto_chart(ws, sheet_name, first, last)
        # Light tab colour to reinforce the reggae theme.
        if sheet_name.startswith("🟢"):
            ws.sheet_properties.tabColor = GREEN
        elif sheet_name.startswith("🟡"):
            ws.sheet_properties.tabColor = YELLOW
        elif sheet_name.startswith("🔴"):
            ws.sheet_properties.tabColor = RED

    build_dashboard(wb)

    wb.save(DST)
    print(f"Wrote {DST}")


if __name__ == "__main__":
    main()
