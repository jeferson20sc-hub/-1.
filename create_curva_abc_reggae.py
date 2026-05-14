#!/usr/bin/env python3
"""
Create a Professional Reggae-Themed ABC Curve Excel Dashboard
Colors: Green #009B3A, Gold #FFD700, Red #CC0000, Black #000000
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    GradientFill, numbers
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
import copy

# ===================== REGGAE COLOR PALETTE =====================
# Primary reggae colors
GREEN   = "009B3A"   # Jamaican green
GOLD    = "FFD700"   # Jamaican gold/yellow
RED     = "CC0000"   # Jamaican red
BLACK   = "000000"   # Jamaican black
WHITE   = "FFFFFF"
# Secondary shades
GREEN_DARK  = "006B28"
GREEN_LIGHT = "C8F5DB"
GREEN_MID   = "00C44A"
GOLD_DARK   = "C8A400"
GOLD_LIGHT  = "FFF5B0"
GOLD_MID    = "FFE44D"
RED_DARK    = "990000"
RED_LIGHT   = "FFD0D0"
RED_MID     = "FF4444"
GRAY_LIGHT  = "F5F5F5"
GRAY_MID    = "E0E0E0"
GRAY_DARK   = "B0B0B0"
DARK_BG     = "1A1A1A"  # near black for dashboard bg


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_font(bold=False, color=WHITE, size=11, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def make_border(style="thin", color=BLACK):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_border_thick():
    thick = Side(style="medium", color=BLACK)
    thin  = Side(style="thin",   color=BLACK)
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def make_border_outer_thick():
    thick = Side(style="thick", color=BLACK)
    thin  = Side(style="thin",  color=GRAY_DARK)
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def style_cell(cell, fill=None, font=None, border=None, alignment=None, num_fmt=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if border:    cell.border    = border
    if alignment: cell.alignment = alignment
    if num_fmt:   cell.number_format = num_fmt


# ===================== SHEET DATA DEFINITIONS =====================

SHEET_CONFIGS = [
    {
        "name": "Primeira",
        "title": "CURVA ABC REGGAE  —  PRIMEIRA",
        "subtitle": "Gestao de Estoque & Inventario | Classe A: ate 80%  |  Classe B: 80-95%  |  Classe C: 95-100%",
        "data_start": 10,
        "items": [
            # (Descricao,           Cod, Quant, Custo)
            ("Servidor Backup",       3,  2,  25000),
            ('Monitor 27" 4K',        5, 10,   3000),
            ("Notebook Administrativo",9,  6,   4500),
            ("Nobreak Profissional",   1,  5,   2000),
            ("Switch Gerenciavel",     7,  4,   1500),
            ("Mouse sem fio",          4, 40,     50),
            ("Patch Cord 1m",          8,100,     15),
            ("Teclado USB",            6, 80,     35),
            ("Cabo Rede Cat6 (m)",     2, 200,     8),
            ("Mousepad Slim",         10,120,     12),
        ],
    },
    {
        "name": "01",
        "title": "CURVA ABC REGGAE  —  ABA 01",
        "subtitle": "Analise de Estoque | Tecnologia da Informacao | Dashboard Premium",
        "data_start": 10,
        "items": [
            ("Notebook Executivo i7",     10, 35,  6500),
            ('Servidor de Dados Pro',      2, 12, 15000),
            ("Impressora Multifuncional", 17, 55,  1100),
            ("Toner Impressora Laser",    12, 68,   680),
            ("Memoria RAM 16GB",          19, 45,   380),
            ('Monitor LED 24"',            4, 20,   850),
            ("Webcam Full HD",            14, 82,   180),
            ("Roteador Wi-Fi 6",          11, 15,   650),
            ("SSD 480GB Sata",            15, 55,   190),
            ("Nobreak 1500VA",             6,  8,  1200),
            ("HD Externo 2TB",            20, 18,   420),
            ("Teclado Mecanico RGB",       5, 30,   250),
            ("Pen Drive 64GB",             9,160,    45),
            ("Headset com Microfone",     13, 64,   110),
            ("Pacote de Papel A4 (500fls)",7,237,    28),
            ("Filtro de Linha 5 Tom.",    18,110,    40),
            ("Cabo HDMI 2m",               1,150,    25),
            ("Adaptador USB-C",           16, 90,    35),
            ("Mouse Optico Simples",       3,120,    15),
            ("Patch Cord RJ45 1m",         8,200,    12),
        ],
    },
    {
        "name": "02",
        "title": "CURVA ABC REGGAE  —  ABA 02",
        "subtitle": "Analise Avancada | 30 SKUs | Pareto Automatico | Dashboard Executivo",
        "data_start": 10,
        "items": [
            ("Servidor Rack Dell Pro",     2, 15, 22500),
            ("Workstation Grafica",       17, 19, 12000),
            ("Notebook i7 32GB RAM",      10, 28,  7200),
            ("Switch Gerenciavel 48p",     8, 10,  4800),
            ("Placa de Video RTX",        22,  8,  5500),
            ("Storage NAS 40TB",          21,  2, 18500),
            ("Memoria RAM 16GB",          19, 45,   380),
            ("Toner Impressora Laser",    12, 40,   320),
            ('Monitor LED 24"',            4, 15,   850),
            ("Roteador Wi-Fi 6",          11, 25,   450),
            ("SSD 480GB Sata",            15, 55,   190),
            ("Nobreak 1500VA",             6,  8,  1200),
            ("HD Externo 2TB",            20, 18,   420),
            ("Teclado Mecanico RGB",       5, 30,   250),
            ("Pen Drive 64GB",             9,150,    45),
            ("Headset com Microfone",     13, 60,   110),
            ("Pacote de Papel A4",         7,200,    28),
            ("Teclado de Entrada",        29,100,    45),
            ("Filtro de Linha",           18,110,    40),
            ("Webcam Full HD",            14, 22,   180),
            ("Pilhas AA (Pacote)",        23,300,    12),
            ("Adaptador USB-C",           16, 90,    35),
            ("Suporte para Monitor",      28, 20,   140),
            ("Cabo HDMI 2m",               1, 85,    25),
            ("Mousepad Simples",          24,250,     8),
            ("Mouse Optico Simples",       3,120,    15),
            ("Organizador de Cabos",      25,180,    10),
            ("Ar comprimido (Lata)",      27, 35,    45),
            ("Pasta Termica",             26, 45,    35),
            ("Conector RJ45 (Cento)",     30, 50,    30),
        ],
    },
]


def build_abc_data(items):
    """Pre-compute ABC classification for styling and charts."""
    data = []
    for desc, cod, qty, cost in items:
        vt = qty * cost
        data.append({"desc": desc, "cod": cod, "qty": qty, "cost": cost, "vt": vt})

    # Sort by VT descending
    data.sort(key=lambda x: x["vt"], reverse=True)
    total = sum(d["vt"] for d in data)

    cumsum = 0
    for i, d in enumerate(data):
        cumsum += d["vt"]
        d["pct_ind"] = d["vt"] / total if total else 0
        d["pct_acc"] = cumsum / total if total else 0
        if d["pct_acc"] <= 0.80:
            d["abc"] = "A"
        elif d["pct_acc"] <= 0.95:
            d["abc"] = "B"
        else:
            d["abc"] = "C"
        d["rank"] = i + 1
    return data, total


def apply_thick_outer_border(ws, min_row, min_col, max_row, max_col, color=BLACK, style="medium"):
    """Apply thick outer border to a rectangular range."""
    thick = Side(style=style, color=color)
    thin  = Side(style="thin",  color=color)

    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            r, c = cell.row, cell.column
            top    = thick if r == min_row else thin
            bottom = thick if r == max_row else thin
            left   = thick if c == min_col else thin
            right  = thick if c == max_col else thin
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def add_kpi_card(ws, row, col, label, value_ref, bg_color, fg_color=WHITE, fmt=None):
    """Add a styled KPI card (2 rows x 2 cols)."""
    # Label cell
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.fill = make_fill(BLACK)
    lbl.font = make_font(bold=True, color=bg_color, size=9)
    lbl.alignment = center_align()
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

    # Value cell
    val = ws.cell(row=row+1, column=col, value=value_ref)
    val.fill = make_fill(bg_color)
    val.font = make_font(bold=True, color=fg_color, size=14)
    val.alignment = center_align()
    if fmt:
        val.number_format = fmt
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)

    apply_thick_outer_border(ws, row, col, row+1, col+1, color=bg_color, style="medium")


def create_sheet(wb, config):
    ws = wb.create_sheet(title=config["name"])
    items_raw = config["items"]
    data_start = config["data_start"]
    n = len(items_raw)
    data_end = data_start + n - 1

    # Pre-compute ABC
    abc_data, total_vt = build_abc_data(items_raw)

    # Reverse map: original row -> abc classification (for conditional coloring)
    orig_data = []
    for desc, cod, qty, cost in items_raw:
        vt = qty * cost
        orig_data.append({"desc": desc, "cod": cod, "qty": qty, "cost": cost, "vt": vt})
    # We'll also need the sorted order for chart data

    # ============================================================
    # COLUMN WIDTHS
    # ============================================================
    col_widths = {
        "A": 7,    # ABC class
        "B": 32,   # Description
        "C": 10,   # Code
        "D": 10,   # Quantity
        "E": 14,   # Unit cost
        "F": 16,   # V.T.
        "G": 13,   # % Individual
        "H": 13,   # % Accumulated
        "I": 2,    # spacer
        "J": 10,   # Class summary label
        "K": 16,   # Class summary VT
        "L": 10,   # Class summary %
        "M": 10,   # Class summary qty
        "N": 22,   # Class bar
        "O": 2,    # spacer
        "P": 2,    # spacer
        "Q": 2,    # spacer
        "R": 8,    # Rank
        "S": 6,    # ABC sorted
        "T": 38,   # Label
        "U": 16,   # VT sorted
        "V": 13,   # % Accumulated sorted
        "W": 16,   # VT A
        "X": 16,   # VT B
        "Y": 16,   # VT C
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ============================================================
    # ROW HEIGHTS
    # ============================================================
    ws.row_dimensions[1].height = 50   # Main title
    ws.row_dimensions[2].height = 22   # KPI labels row 1
    ws.row_dimensions[3].height = 32   # KPI values row 1
    ws.row_dimensions[4].height = 22   # KPI labels row 2
    ws.row_dimensions[5].height = 32   # KPI values row 2
    ws.row_dimensions[6].height = 18   # Instruction
    ws.row_dimensions[7].height = 18   # Sub-header 1
    ws.row_dimensions[8].height = 22   # Header row
    ws.row_dimensions[9].height = 6    # small spacer
    for r in range(data_start, data_end + 1):
        ws.row_dimensions[r].height = 19

    # ============================================================
    # ROW 1: MAIN TITLE (merged A1:N1)
    # ============================================================
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = config["title"]
    title_cell.fill  = make_fill(BLACK)
    title_cell.font  = make_font(bold=True, color=GREEN, size=20, name="Calibri")
    title_cell.alignment = center_align()

    # Reggae stripe: green, gold, red (cols I-N)
    stripe_colors = [GREEN, GOLD, RED, GREEN, GOLD, RED]
    for idx, col_letter in enumerate(["I", "J", "K", "L", "M", "N"]):
        c = ws[f"{col_letter}1"]
        c.fill = make_fill(stripe_colors[idx % 3])

    # Right side: legend area (cols J:N row 1 merged)
    ws.merge_cells("J1:N1")
    leg = ws["J1"]
    leg.value = "A = ate 80%   |   B = 80-95%   |   C = 95-100%"
    leg.fill  = make_fill(BLACK)
    leg.font  = make_font(bold=False, color=GOLD, size=10)
    leg.alignment = center_align()

    # Subtitle row above KPIs (using row 1 subtitle - actually add it in row 6)
    # We'll put subtitle in row 6 as instruction

    # ============================================================
    # ROWS 2-5: KPI DASHBOARD PANEL
    # ============================================================
    # KPI Panel background (cols A:H, rows 2-5)
    for r in range(2, 6):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.fill = make_fill(DARK_BG)

    # KPI cards definition
    kpi_row1 = 2
    kpi_row2 = 4

    # Row 2-3: KPI cards  (each card = 2 cols wide)
    # Card 1: V.T. TOTAL (cols A:B)
    kpi_lbl1 = ws.cell(row=kpi_row1, column=1, value="V.T. TOTAL")
    kpi_lbl1.fill = make_fill(DARK_BG)
    kpi_lbl1.font = make_font(bold=True, color=GREEN, size=9)
    kpi_lbl1.alignment = center_align()
    ws.merge_cells(start_row=kpi_row1, start_column=1, end_row=kpi_row1, end_column=2)

    kpi_val1 = ws.cell(row=kpi_row1+1, column=1,
                       value=f"=SUM($F${data_start}:$F${data_end})")
    kpi_val1.fill = make_fill(GREEN_DARK)
    kpi_val1.font = make_font(bold=True, color=WHITE, size=14)
    kpi_val1.alignment = center_align()
    kpi_val1.number_format = 'R$ #,##0.00'
    ws.merge_cells(start_row=kpi_row1+1, start_column=1, end_row=kpi_row1+1, end_column=2)

    # Card 2: V.T. CLASSE A (cols C:D)
    ws.cell(row=kpi_row1, column=3, value="V.T. CLASSE A").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row1, column=3).font = make_font(bold=True, color=GREEN, size=9)
    ws.cell(row=kpi_row1, column=3).alignment = center_align()
    ws.merge_cells(start_row=kpi_row1, start_column=3, end_row=kpi_row1, end_column=4)

    kpi_va = ws.cell(row=kpi_row1+1, column=3,
                     value=f'=SUMIF($A${data_start}:$A${data_end},"A",$F${data_start}:$F${data_end})')
    kpi_va.fill = make_fill(GREEN)
    kpi_va.font = make_font(bold=True, color=WHITE, size=14)
    kpi_va.alignment = center_align()
    kpi_va.number_format = 'R$ #,##0.00'
    ws.merge_cells(start_row=kpi_row1+1, start_column=3, end_row=kpi_row1+1, end_column=4)

    # Card 3: % ITENS A (cols E:F)
    ws.cell(row=kpi_row1, column=5, value="% ITENS A").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row1, column=5).font = make_font(bold=True, color=GOLD, size=9)
    ws.cell(row=kpi_row1, column=5).alignment = center_align()
    ws.merge_cells(start_row=kpi_row1, start_column=5, end_row=kpi_row1, end_column=6)

    kpi_pca = ws.cell(row=kpi_row1+1, column=5,
                      value=f'=IFERROR(COUNTIF($A${data_start}:$A${data_end},"A")/COUNTA($B${data_start}:$B${data_end}),0)')
    kpi_pca.fill = make_fill(GOLD_DARK)
    kpi_pca.font = make_font(bold=True, color=WHITE, size=14)
    kpi_pca.alignment = center_align()
    kpi_pca.number_format = '0.0%'
    ws.merge_cells(start_row=kpi_row1+1, start_column=5, end_row=kpi_row1+1, end_column=6)

    # Card 4: MAIOR V.T. (cols G:H)
    ws.cell(row=kpi_row1, column=7, value="MAIOR V.T.").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row1, column=7).font = make_font(bold=True, color=RED_MID, size=9)
    ws.cell(row=kpi_row1, column=7).alignment = center_align()
    ws.merge_cells(start_row=kpi_row1, start_column=7, end_row=kpi_row1, end_column=8)

    kpi_mvt = ws.cell(row=kpi_row1+1, column=7,
                      value=f'=MAX($F${data_start}:$F${data_end})')
    kpi_mvt.fill = make_fill(RED_DARK)
    kpi_mvt.font = make_font(bold=True, color=WHITE, size=14)
    kpi_mvt.alignment = center_align()
    kpi_mvt.number_format = 'R$ #,##0.00'
    ws.merge_cells(start_row=kpi_row1+1, start_column=7, end_row=kpi_row1+1, end_column=8)

    # Row 4-5: Second row of KPIs
    # Card 5: MENOR V.T. (A:B)
    ws.cell(row=kpi_row2, column=1, value="MENOR V.T.").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row2, column=1).font = make_font(bold=True, color=GREEN, size=9)
    ws.cell(row=kpi_row2, column=1).alignment = center_align()
    ws.merge_cells(start_row=kpi_row2, start_column=1, end_row=kpi_row2, end_column=2)

    kpi_min = ws.cell(row=kpi_row2+1, column=1,
                      value=f'=MIN($F${data_start}:$F${data_end})')
    kpi_min.fill = make_fill(GREEN_DARK)
    kpi_min.font = make_font(bold=True, color=WHITE, size=12)
    kpi_min.alignment = center_align()
    kpi_min.number_format = 'R$ #,##0.00'
    ws.merge_cells(start_row=kpi_row2+1, start_column=1, end_row=kpi_row2+1, end_column=2)

    # Card 6: STATUS (C:D)
    ws.cell(row=kpi_row2, column=3, value="STATUS").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row2, column=3).font = make_font(bold=True, color=GREEN, size=9)
    ws.cell(row=kpi_row2, column=3).alignment = center_align()
    ws.merge_cells(start_row=kpi_row2, start_column=3, end_row=kpi_row2, end_column=4)

    kpi_sts = ws.cell(row=kpi_row2+1, column=3,
                      value=f'=IF(IFERROR(MAX($H${data_start}:$H${data_end}),0)>1.0001,"VERIFICAR","CONSISTENTE")')
    kpi_sts.fill = make_fill(GREEN)
    kpi_sts.font = make_font(bold=True, color=WHITE, size=12)
    kpi_sts.alignment = center_align()
    ws.merge_cells(start_row=kpi_row2+1, start_column=3, end_row=kpi_row2+1, end_column=4)

    # Card 7: LIMITE A (E:F)
    ws.cell(row=kpi_row2, column=5, value="LIMITE A").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row2, column=5).font = make_font(bold=True, color=GOLD, size=9)
    ws.cell(row=kpi_row2, column=5).alignment = center_align()
    ws.merge_cells(start_row=kpi_row2, start_column=5, end_row=kpi_row2, end_column=6)

    kpi_la = ws.cell(row=kpi_row2+1, column=5, value=0.80)
    kpi_la.fill = make_fill(GOLD_DARK)
    kpi_la.font = make_font(bold=True, color=WHITE, size=12)
    kpi_la.alignment = center_align()
    kpi_la.number_format = '0%'
    ws.merge_cells(start_row=kpi_row2+1, start_column=5, end_row=kpi_row2+1, end_column=6)

    # Card 8: LIMITE B (G:H)
    ws.cell(row=kpi_row2, column=7, value="LIMITE B").fill = make_fill(DARK_BG)
    ws.cell(row=kpi_row2, column=7).font = make_font(bold=True, color=RED_MID, size=9)
    ws.cell(row=kpi_row2, column=7).alignment = center_align()
    ws.merge_cells(start_row=kpi_row2, start_column=7, end_row=kpi_row2, end_column=8)

    kpi_lb = ws.cell(row=kpi_row2+1, column=7, value=0.95)
    kpi_lb.fill = make_fill(RED_DARK)
    kpi_lb.font = make_font(bold=True, color=WHITE, size=12)
    kpi_lb.alignment = center_align()
    kpi_lb.number_format = '0%'
    ws.merge_cells(start_row=kpi_row2+1, start_column=7, end_row=kpi_row2+1, end_column=8)

    # Borders for KPI panel
    apply_thick_outer_border(ws, 2, 1, 5, 8, color=GREEN, style="medium")
    for r in range(2, 6):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            if cell.border.left.style is None:
                cell.border = make_border(style="hair", color=GREEN_DARK)

    # ============================================================
    # ROWS 2-5: RIGHT SIDE - Class Summary Panel (cols J:N)
    # ============================================================
    # Background
    for r in range(2, 6):
        for c in range(10, 15):
            ws.cell(row=r, column=c).fill = make_fill(DARK_BG)

    # Header row for summary
    ws.cell(row=7, column=10, value="Classe").fill  = make_fill(BLACK)
    ws.cell(row=7, column=11, value="Total V.T.").fill = make_fill(BLACK)
    ws.cell(row=7, column=12, value="% Total").fill = make_fill(BLACK)
    ws.cell(row=7, column=13, value="Qtd. Itens").fill = make_fill(BLACK)
    ws.cell(row=7, column=14, value="Barra Rapida").fill = make_fill(BLACK)

    for c in range(10, 15):
        cell = ws.cell(row=7, column=c)
        cell.font = make_font(bold=True, color=WHITE, size=9)
        cell.alignment = center_align()
        cell.border = make_border(color=GRAY_DARK)

    # Class A summary (row 8)
    cls_rows = [
        ("A", GREEN,    GOLD_LIGHT,  8),
        ("B", GOLD_DARK, BLACK,      9),
        ("C", RED_DARK,  WHITE,     10),
    ]

    LIMIT_A_REF = "E5"
    LIMIT_B_REF = "G5"

    for cls, bg, fg, r in cls_rows:
        # Class label
        c_lbl = ws.cell(row=r, column=10, value=cls)
        c_lbl.fill = make_fill(bg)
        c_lbl.font = make_font(bold=True, color=WHITE if cls != "B" else BLACK, size=16)
        c_lbl.alignment = center_align()
        c_lbl.border = make_border(color=BLACK)

        # Total VT
        c_vt = ws.cell(row=r, column=11,
                       value=f'=SUMIF($A${data_start}:$A${data_end},"{cls}",$F${data_start}:$F${data_end})')
        c_vt.fill = make_fill(bg)
        c_vt.font = make_font(bold=True, color=WHITE if cls != "B" else BLACK, size=11)
        c_vt.alignment = right_align()
        c_vt.number_format = 'R$ #,##0'
        c_vt.border = make_border(color=BLACK)

        # % Total
        c_pct = ws.cell(row=r, column=12,
                        value=f'=IFERROR(K{r}/SUM($F${data_start}:$F${data_end}),0)')
        c_pct.fill = make_fill(bg)
        c_pct.font = make_font(bold=True, color=WHITE if cls != "B" else BLACK, size=11)
        c_pct.alignment = center_align()
        c_pct.number_format = '0.0%'
        c_pct.border = make_border(color=BLACK)

        # Count items
        c_cnt = ws.cell(row=r, column=13,
                        value=f'=COUNTIF($A${data_start}:$A${data_end},"{cls}")')
        c_cnt.fill = make_fill(bg)
        c_cnt.font = make_font(bold=True, color=WHITE if cls != "B" else BLACK, size=11)
        c_cnt.alignment = center_align()
        c_cnt.border = make_border(color=BLACK)

        # Bar sparkline
        c_bar = ws.cell(row=r, column=14,
                        value=f'=REPT("█",ROUND(L{r}*20,0))')
        c_bar.fill = make_fill(bg)
        c_bar.font = make_font(bold=False, color=WHITE if cls != "B" else BLACK, size=8)
        c_bar.alignment = left_align()
        c_bar.border = make_border(color=BLACK)

    # Consistency alert row (row 11)
    ws.merge_cells("J11:N11")
    alert_cell = ws.cell(row=11, column=10,
                         value=f'=IF(IFERROR(MAX($H${data_start}:$H${data_end}),0)>1.0001,"ALERTA: acumulado > 100%","OK - CONSISTENTE")')
    alert_cell.fill = make_fill(DARK_BG)
    alert_cell.font = make_font(bold=True, color=GREEN, size=9)
    alert_cell.alignment = center_align()

    # ============================================================
    # ROW 6: INSTRUCTION BAR
    # ============================================================
    ws.merge_cells("A6:H6")
    instr = ws["A6"]
    instr.value = ("  " + config.get("subtitle", "") +
                   "  |  Edite somente Quant. e Custo Unit. — formulas, ABC e graficos atualizam automaticamente.")
    instr.fill  = make_fill(GREEN_DARK)
    instr.font  = make_font(bold=False, color=GOLD_LIGHT, size=8, italic=True)
    instr.alignment = left_align(wrap=False)

    # ============================================================
    # ROW 8: COLUMN HEADERS
    # ============================================================
    ws.row_dimensions[8].height = 26

    headers = [
        (1,  "ABC",            BLACK, GREEN,   center_align()),
        (2,  "Descricao do Item", BLACK, WHITE, left_align()),
        (3,  "Cod./ID",         BLACK, WHITE,  center_align()),
        (4,  "Quant.",          BLACK, GOLD,   center_align()),
        (5,  "Custo Unit. (R$)",BLACK, GOLD,   center_align()),
        (6,  "V.T. (R$)",       BLACK, GREEN,  center_align()),
        (7,  "% Individual",    BLACK, WHITE,  center_align()),
        (8,  "% Acumulado",     BLACK, WHITE,  center_align()),
    ]

    for col, label, bg, fg, align in headers:
        cell = ws.cell(row=8, column=col, value=label)
        cell.fill = make_fill(bg)
        cell.font = make_font(bold=True, color=fg, size=10)
        cell.alignment = align
        cell.border = make_border(color=GREEN_DARK)

    # Hidden helper headers (R-Y cols) - smaller styling
    helper_headers = [
        (18, "Rank",         DARK_BG, GRAY_MID),
        (19, "ABC Ord.",     DARK_BG, GRAY_MID),
        (20, "Rotulo",       DARK_BG, GRAY_MID),
        (21, "V.T. Ord.",    DARK_BG, GRAY_MID),
        (22, "% Acum. Ord.", DARK_BG, GRAY_MID),
        (23, "V.T. A",       GREEN,   WHITE),
        (24, "V.T. B",       GOLD_DARK, WHITE),
        (25, "V.T. C",       RED_DARK,  WHITE),
    ]
    for col, label, bg, fg in helper_headers:
        cell = ws.cell(row=8, column=col, value=label)
        cell.fill = make_fill(bg)
        cell.font = make_font(bold=True, color=fg, size=8)
        cell.alignment = center_align()
        cell.border = make_border(color=GRAY_DARK)

    # ============================================================
    # DATA ROWS (rows data_start to data_end)
    # ============================================================
    LIMIT_A_CELL = f"$E$5"  # 0.80
    LIMIT_B_CELL = f"$G$5"  # 0.95

    abc_class_for_styling = {}  # row -> abc class

    for i, (desc, cod, qty, cost) in enumerate(items_raw):
        row = data_start + i
        vt = qty * cost

        # Determine ABC class based on pre-computed sorted data
        # We need to find what class this item will be classified
        # Use the same cumulative logic the formula uses
        matching = [d for d in abc_data if d["desc"] == desc and d["cod"] == cod]
        if matching:
            item_abc = matching[0]["abc"]
        else:
            item_abc = "C"
        abc_class_for_styling[row] = item_abc

        # Row background color based on ABC class
        if item_abc == "A":
            row_fill   = make_fill(GREEN_LIGHT)
            abc_fill   = make_fill(GREEN)
            abc_font   = make_font(bold=True, color=WHITE, size=11)
        elif item_abc == "B":
            row_fill   = make_fill(GOLD_LIGHT)
            abc_fill   = make_fill(GOLD_DARK)
            abc_font   = make_font(bold=True, color=WHITE, size=11)
        else:
            row_fill   = make_fill(RED_LIGHT)
            abc_fill   = make_fill(RED_DARK)
            abc_font   = make_font(bold=True, color=WHITE, size=11)

        # Col A: ABC formula
        cell_a = ws.cell(row=row, column=1,
                         value=f'=IF($H{row}<={LIMIT_A_CELL},"A",IF($H{row}<={LIMIT_B_CELL},"B","C"))')
        cell_a.fill = abc_fill
        cell_a.font = abc_font
        cell_a.alignment = center_align()
        cell_a.border = make_border(color=BLACK)

        # Col B: Description
        cell_b = ws.cell(row=row, column=2, value=desc)
        cell_b.fill = row_fill
        cell_b.font = make_font(bold=False, color=BLACK, size=10)
        cell_b.alignment = left_align()
        cell_b.border = make_border(color=GRAY_DARK)

        # Col C: Code
        cell_c = ws.cell(row=row, column=3, value=cod)
        cell_c.fill = row_fill
        cell_c.font = make_font(bold=False, color=BLACK, size=10)
        cell_c.alignment = center_align()
        cell_c.border = make_border(color=GRAY_DARK)

        # Col D: Quantity (editable - light border)
        cell_d = ws.cell(row=row, column=4, value=qty)
        cell_d.fill = make_fill(WHITE)
        cell_d.font = make_font(bold=True, color=BLACK, size=10)
        cell_d.alignment = center_align()
        cell_d.border = Border(
            left=Side(style="medium", color=GREEN_DARK),
            right=Side(style="medium", color=GREEN_DARK),
            top=Side(style="thin", color=GRAY_DARK),
            bottom=Side(style="thin", color=GRAY_DARK),
        )
        cell_d.number_format = '#,##0'

        # Col E: Unit cost (editable)
        cell_e = ws.cell(row=row, column=5, value=cost)
        cell_e.fill = make_fill(WHITE)
        cell_e.font = make_font(bold=True, color=BLACK, size=10)
        cell_e.alignment = right_align()
        cell_e.border = Border(
            left=Side(style="medium", color=GREEN_DARK),
            right=Side(style="medium", color=GREEN_DARK),
            top=Side(style="thin", color=GRAY_DARK),
            bottom=Side(style="thin", color=GRAY_DARK),
        )
        cell_e.number_format = 'R$ #,##0.00'

        # Col F: V.T. formula
        cell_f = ws.cell(row=row, column=6,
                         value=f'=IFERROR(D{row}*E{row},0)')
        cell_f.fill = row_fill
        cell_f.font = make_font(bold=True, color=BLACK, size=10)
        cell_f.alignment = right_align()
        cell_f.border = make_border(color=GRAY_DARK)
        cell_f.number_format = 'R$ #,##0.00'

        # Col G: % Individual
        cell_g = ws.cell(row=row, column=7,
                         value=f'=IFERROR(F{row}/SUM($F${data_start}:$F${data_end}),0)')
        cell_g.fill = row_fill
        cell_g.font = make_font(bold=False, color=BLACK, size=10)
        cell_g.alignment = center_align()
        cell_g.border = make_border(color=GRAY_DARK)
        cell_g.number_format = '0.00%'

        # Col H: % Accumulated
        cell_h = ws.cell(row=row, column=8,
                         value=f'=IFERROR(SUMIF($F${data_start}:$F${data_end},">="&F{row},$F${data_start}:$F${data_end})/SUM($F${data_start}:$F${data_end}),0)')
        cell_h.fill = row_fill
        cell_h.font = make_font(bold=False, color=BLACK, size=10)
        cell_h.alignment = center_align()
        cell_h.border = make_border(color=GRAY_DARK)
        cell_h.number_format = '0.00%'

        # ---- Helper columns (R-Y) for chart ----
        # Col R: Rank
        ws.cell(row=row, column=18,
                value=f'=RANK.EQ(F{row},$F${data_start}:$F${data_end},0)+COUNTIF($F${data_start}:F{row},F{row})-1')
        ws.cell(row=row, column=18).fill = make_fill(DARK_BG)
        ws.cell(row=row, column=18).font = make_font(color=GRAY_MID, size=8)
        ws.cell(row=row, column=18).alignment = center_align()

        # Col S: ABC sorted
        ws.cell(row=row, column=19,
                value=f'=INDEX($A${data_start}:$A${data_end},MATCH(ROWS($S${data_start}:S{row}),$R${data_start}:$R${data_end},0))')
        ws.cell(row=row, column=19).fill = make_fill(DARK_BG)
        ws.cell(row=row, column=19).font = make_font(color=GRAY_MID, size=8)
        ws.cell(row=row, column=19).alignment = center_align()

        # Col T: Label sorted
        ws.cell(row=row, column=20,
                value=f'=INDEX($B${data_start}:$B${data_end},MATCH(ROWS($T${data_start}:T{row}),$R${data_start}:$R${data_end},0))&" ("&S{row}&")"')
        ws.cell(row=row, column=20).fill = make_fill(DARK_BG)
        ws.cell(row=row, column=20).font = make_font(color=GRAY_MID, size=8)
        ws.cell(row=row, column=20).alignment = left_align()

        # Col U: VT sorted
        ws.cell(row=row, column=21,
                value=f'=INDEX($F${data_start}:$F${data_end},MATCH(ROWS($U${data_start}:U{row}),$R${data_start}:$R${data_end},0))')
        ws.cell(row=row, column=21).fill = make_fill(DARK_BG)
        ws.cell(row=row, column=21).font = make_font(color=GRAY_MID, size=8)
        ws.cell(row=row, column=21).alignment = right_align()
        ws.cell(row=row, column=21).number_format = '#,##0'

        # Col V: % Accumulated sorted
        ws.cell(row=row, column=22,
                value=f'=INDEX($H${data_start}:$H${data_end},MATCH(ROWS($V${data_start}:V{row}),$R${data_start}:$R${data_end},0))')
        ws.cell(row=row, column=22).fill = make_fill(DARK_BG)
        ws.cell(row=row, column=22).font = make_font(color=GRAY_MID, size=8)
        ws.cell(row=row, column=22).alignment = center_align()
        ws.cell(row=row, column=22).number_format = '0.0%'

        # Col W: VT A
        ws.cell(row=row, column=23,
                value=f'=IF($S{row}="A",$U{row},NA())')
        ws.cell(row=row, column=23).fill = make_fill(GREEN_DARK)
        ws.cell(row=row, column=23).font = make_font(color=GREEN_LIGHT, size=8)
        ws.cell(row=row, column=23).number_format = '#,##0'

        # Col X: VT B
        ws.cell(row=row, column=24,
                value=f'=IF($S{row}="B",$U{row},NA())')
        ws.cell(row=row, column=24).fill = make_fill(GOLD_DARK)
        ws.cell(row=row, column=24).font = make_font(color=WHITE, size=8)
        ws.cell(row=row, column=24).number_format = '#,##0'

        # Col Y: VT C
        ws.cell(row=row, column=25,
                value=f'=IF($S{row}="C",$U{row},NA())')
        ws.cell(row=row, column=25).fill = make_fill(RED_DARK)
        ws.cell(row=row, column=25).font = make_font(color=WHITE, size=8)
        ws.cell(row=row, column=25).number_format = '#,##0'

    # ============================================================
    # TOTAL ROW
    # ============================================================
    total_row = data_end + 1
    ws.row_dimensions[total_row].height = 22

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    total_lbl = ws.cell(row=total_row, column=1, value="TOTAL GERAL")
    total_lbl.fill = make_fill(BLACK)
    total_lbl.font = make_font(bold=True, color=GOLD, size=11)
    total_lbl.alignment = center_align()

    total_qty = ws.cell(row=total_row, column=4,
                        value=f'=SUM(D{data_start}:D{data_end})')
    total_qty.fill = make_fill(BLACK)
    total_qty.font = make_font(bold=True, color=WHITE, size=11)
    total_qty.alignment = center_align()
    total_qty.number_format = '#,##0'

    ws.cell(row=total_row, column=5).fill = make_fill(BLACK)

    total_vt_cell = ws.cell(row=total_row, column=6,
                             value=f'=SUM(F{data_start}:F{data_end})')
    total_vt_cell.fill = make_fill(GREEN)
    total_vt_cell.font = make_font(bold=True, color=WHITE, size=11)
    total_vt_cell.alignment = right_align()
    total_vt_cell.number_format = 'R$ #,##0.00'

    total_pct = ws.cell(row=total_row, column=7, value=1.0)
    total_pct.fill = make_fill(BLACK)
    total_pct.font = make_font(bold=True, color=WHITE, size=11)
    total_pct.alignment = center_align()
    total_pct.number_format = '0%'

    total_acc = ws.cell(row=total_row, column=8, value=1.0)
    total_acc.fill = make_fill(BLACK)
    total_acc.font = make_font(bold=True, color=WHITE, size=11)
    total_acc.alignment = center_align()
    total_acc.number_format = '0%'

    for c in range(1, 9):
        cell = ws.cell(row=total_row, column=c)
        cell.border = Border(
            top=Side(style="medium", color=GREEN),
            bottom=Side(style="medium", color=GREEN),
            left=Side(style="thin", color=GREEN),
            right=Side(style="thin", color=GREEN),
        )

    # ============================================================
    # ADD CHARTS
    # ============================================================
    chart_row = total_row + 3

    # --- Chart 1: Bar Chart (V.T. by class A/B/C stacked) ---
    bar_chart = BarChart()
    bar_chart.type    = "col"
    bar_chart.grouping = "clustered"
    bar_chart.title   = f"Curva ABC — Valor Total por Item ({config['name']})"
    bar_chart.y_axis.title = "Valor Total (R$)"
    bar_chart.x_axis.title = "Itens (ordenados por V.T.)"
    bar_chart.style   = 10
    bar_chart.width   = 24
    bar_chart.height  = 14

    # Series A (green)
    ser_a = Series(
        Reference(ws, min_col=23, min_row=data_start, max_row=data_end),
        title="Classe A"
    )
    ser_a.graphicalProperties.solidFill = GREEN
    ser_a.graphicalProperties.line.solidFill = GREEN_DARK
    bar_chart.series.append(ser_a)

    # Series B (gold)
    ser_b = Series(
        Reference(ws, min_col=24, min_row=data_start, max_row=data_end),
        title="Classe B"
    )
    ser_b.graphicalProperties.solidFill = GOLD
    ser_b.graphicalProperties.line.solidFill = GOLD_DARK
    bar_chart.series.append(ser_b)

    # Series C (red)
    ser_c = Series(
        Reference(ws, min_col=25, min_row=data_start, max_row=data_end),
        title="Classe C"
    )
    ser_c.graphicalProperties.solidFill = RED
    ser_c.graphicalProperties.line.solidFill = RED_DARK
    bar_chart.series.append(ser_c)

    # X axis labels (item names from col T)
    cats = Reference(ws, min_col=20, min_row=data_start, max_row=data_end)
    bar_chart.set_categories(cats)
    bar_chart.shape = 4

    # Place chart
    chart_anchor = f"A{chart_row}"
    ws.add_chart(bar_chart, chart_anchor)

    # --- Chart 2: Line Chart (Pareto / Cumulative %) ---
    line_chart = LineChart()
    line_chart.title   = f"Curva de Pareto — % Acumulado ({config['name']})"
    line_chart.y_axis.title = "% Acumulado"
    line_chart.x_axis.title = "Itens"
    line_chart.style   = 10
    line_chart.width   = 24
    line_chart.height  = 14

    # Pareto line
    ser_pareto = Series(
        Reference(ws, min_col=22, min_row=data_start, max_row=data_end),
        title="% Acumulado"
    )
    ser_pareto.graphicalProperties.line.solidFill = GREEN
    ser_pareto.graphicalProperties.line.width      = 25000  # 2.5pt
    ser_pareto.smooth = True
    line_chart.series.append(ser_pareto)

    # 80% reference line (horizontal) - add as separate series with constant value
    line_chart.set_categories(Reference(ws, min_col=20, min_row=data_start, max_row=data_end))

    # Place line chart next to bar chart
    line_col = "N"
    line_anchor = f"{line_col}{chart_row}"
    ws.add_chart(line_chart, line_anchor)

    # ============================================================
    # REGGAE FOOTER BANNER
    # ============================================================
    footer_row = chart_row + 28
    ws.row_dimensions[footer_row].height = 18

    # Reggae flag stripes
    stripe_spans = [
        ("A", "B", GREEN),
        ("C", "D", GOLD),
        ("E", "F", RED),
        ("G", "H", GREEN),
    ]
    for start_col, end_col, color in stripe_spans:
        s = column_index_from_string(start_col)
        e = column_index_from_string(end_col)
        ws.merge_cells(start_row=footer_row, start_column=s,
                       end_row=footer_row, end_column=e)
        fc = ws.cell(row=footer_row, column=s)
        fc.fill = make_fill(color)
        fc.font = make_font(bold=True, color=WHITE if color != GOLD else BLACK, size=9)
        fc.alignment = center_align()

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer_cell = ws.cell(row=footer_row, column=1,
                          value=f"  CURVA ABC REGGAE  |  {config['title']}  |  One Love, One Data  |  2026")
    footer_cell.fill = make_fill(BLACK)
    footer_cell.font = make_font(bold=True, color=GREEN, size=9)
    footer_cell.alignment = center_align()

    # ============================================================
    # CONDITIONAL FORMATTING (visual stripes by ABC class)
    # ============================================================
    data_range = f"A{data_start}:H{data_end}"

    # Class A = Green
    rule_a = FormulaRule(
        formula=[f'$A{data_start}="A"'],
        fill=make_fill(GREEN_LIGHT),
        font=Font(color=BLACK)
    )
    ws.conditional_formatting.add(data_range, rule_a)

    # Class B = Gold
    rule_b = FormulaRule(
        formula=[f'$A{data_start}="B"'],
        fill=make_fill(GOLD_LIGHT),
        font=Font(color=BLACK)
    )
    ws.conditional_formatting.add(data_range, rule_b)

    # Class C = Red
    rule_c = FormulaRule(
        formula=[f'$A{data_start}="C"'],
        fill=make_fill(RED_LIGHT),
        font=Font(color=BLACK)
    )
    ws.conditional_formatting.add(data_range, rule_c)

    # ============================================================
    # SHEET PROTECTION
    # ============================================================
    ws.protection.sheet = True
    ws.protection.password = "abc2026."
    ws.protection.enable()
    # Allow selecting and formatting
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False

    # Unlock editable cells (D and E columns in data range)
    from openpyxl.styles.protection import Protection
    unlock_prot = Protection(locked=False)
    for row in range(data_start, data_end + 1):
        ws.cell(row=row, column=4).protection = unlock_prot  # Quant.
        ws.cell(row=row, column=5).protection = unlock_prot  # Custo Unit.

    # ============================================================
    # FREEZE PANES
    # ============================================================
    ws.freeze_panes = f"A{data_start}"

    # Tab color
    tab_colors = {"Primeira": GREEN, "01": GOLD_DARK, "02": RED_DARK}
    ws.sheet_properties.tabColor = tab_colors.get(config["name"], GREEN)

    return ws


def create_index_sheet(wb):
    """Create a navigation/index sheet as the first sheet."""
    ws = wb.create_sheet(title="INICIO", index=0)

    # Column widths
    for col_letter, width in {"A": 4, "B": 28, "C": 28, "D": 28, "E": 28, "F": 4}.items():
        ws.column_dimensions[col_letter].width = width

    # Row heights
    for r in range(1, 35):
        ws.row_dimensions[r].height = 20
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[8].height = 8
    ws.row_dimensions[14].height = 8
    ws.row_dimensions[20].height = 8

    # Background
    for r in range(1, 36):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = make_fill(DARK_BG)

    # Main title
    ws.merge_cells("B2:E2")
    t = ws["B2"]
    t.value = "CURVA ABC REGGAE"
    t.fill  = make_fill(BLACK)
    t.font  = Font(bold=True, color=GREEN, size=32, name="Calibri")
    t.alignment = center_align()

    ws.merge_cells("B3:E3")
    sub = ws["B3"]
    sub.value = "Dashboard Profissional de Analise de Estoque"
    sub.fill  = make_fill(BLACK)
    sub.font  = make_font(bold=False, color=GOLD, size=14, italic=True)
    sub.alignment = center_align()

    # Reggae stripes
    ws.merge_cells("B4:E4")
    for col, color in [(2, GREEN), (3, GOLD), (4, RED), (5, GREEN)]:
        ws.cell(row=4, column=col).fill = make_fill(color)
    stripe_cell = ws["B4"]
    stripe_cell.fill = make_fill(GREEN)

    # Divider
    for col in range(2, 6):
        c = ws.cell(row=4, column=col)
        colors_stripe = [GREEN, GOLD, RED, BLACK]
        c.fill = make_fill(colors_stripe[(col - 2) % 4])

    # Info cards
    info_items = [
        (6,  "Classificacao ABC", "A = ate 80% do valor total\nB = 80% a 95% do valor\nC = 95% a 100% do valor", GREEN),
        (10, "Abas Disponíveis",  "Primeira: 10 itens\nAba 01: 20 itens\nAba 02: 30 itens", GOLD_DARK),
        (16, "Como Usar",         "1. Va para a aba desejada\n2. Edite Quant. e Custo\n3. ABC calcula automaticamente", RED_DARK),
    ]

    for start_row, title_text, body_text, color in info_items:
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=5)
        title_cell = ws.cell(row=start_row, column=2, value=title_text)
        title_cell.fill = make_fill(color)
        title_cell.font = make_font(bold=True, color=WHITE, size=12)
        title_cell.alignment = center_align()
        title_cell.border = make_border(color=BLACK)

        for sub_row, line in enumerate(body_text.split("\n")):
            r = start_row + 1 + sub_row
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            body_cell = ws.cell(row=r, column=2, value=line)
            body_cell.fill = make_fill(DARK_BG)
            body_cell.font = make_font(bold=False, color=WHITE if color != GOLD_DARK else GOLD_LIGHT, size=10)
            body_cell.alignment = left_align()
            body_cell.border = make_border(color=color)

    # Footer
    ws.merge_cells("B32:E32")
    footer = ws["B32"]
    footer.value = "One Love, One Data  |  CURVA ABC REGGAE  |  2026  |  Senha: abc2026."
    footer.fill  = make_fill(BLACK)
    footer.font  = make_font(bold=True, color=GREEN, size=9)
    footer.alignment = center_align()

    ws.sheet_properties.tabColor = BLACK
    return ws


# ===================== MAIN =====================
def main():
    output_path = "/home/user/Zyth/curva_abc_reggae.xlsx"

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Create index sheet
    create_index_sheet(wb)

    # Create the 3 data sheets
    for config in SHEET_CONFIGS:
        print(f"Creating sheet: {config['name']}...")
        create_sheet(wb, config)

    # Workbook properties
    wb.properties.title    = "Curva ABC Reggae"
    wb.properties.subject  = "Analise de Estoque ABC - Reggae Colors"
    wb.properties.creator  = "CURVA ABC REGGAE Generator"
    wb.properties.keywords = "ABC, Pareto, Estoque, Reggae, Dashboard"

    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print(f"Done! File saved: {output_path}")

    import os
    size = os.path.getsize(output_path)
    print(f"File size: {size:,} bytes ({size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
